# views.py
import datetime
import time
import hashlib
import shutil
import os
import uuid
import re
import unicodedata
from io import BytesIO
from collections import OrderedDict

import pandas as pd
import numpy as np
from django.conf import settings
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.views import View
from .forms import ExcelUploadForm
from django.core.cache import cache

from django.views.decorators.cache import never_cache
import json, traceback, os
from datetime import date
from django.db.models import Q
from django.template.loader import render_to_string
from calendar import month_abbr, month_name
import calendar as calendar_module

from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.utils.text import slugify

from .models import MeetingPoint, ExcelSheetCache

try:
    from dateutil import parser as _dateutil_parser
except ImportError:  # pragma: no cover
    _dateutil_parser = None


def _strip_excel_date_str(val):
    """Normalize Excel-exported date strings (NBSP, BOM, bidi marks)."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    if not s or s.lower() in ("nat", "nan", "none", "<na>"):
        return ""
    for ch in ("\ufeff", "\u200e", "\u200f", "\u202a", "\u202c"):
        s = s.replace(ch, "")
    s = s.replace("\xa0", " ").strip()
    while "  " in s:
        s = s.replace("  ", " ")
    return s


def _try_dateutil_parse(val):
    """
    Last-resort parse for US-style Excel strings (e.g. 1/26/2026 12:15:00 PM)
    when pandas infer / strptime formats leave NaT.
    """
    if _dateutil_parser is None:
        return pd.NaT
    if val is None:
        return pd.NaT
    try:
        if isinstance(val, (float, int)) and not isinstance(val, bool):
            if pd.isna(val):
                return pd.NaT
    except TypeError:
        pass
    if isinstance(val, pd.Timestamp):
        return val if pd.notna(val) else pd.NaT
    if isinstance(val, datetime.datetime):
        return pd.Timestamp(val)
    if isinstance(val, datetime.date):
        return pd.Timestamp(datetime.datetime.combine(val, datetime.time.min))
    s = _strip_excel_date_str(val)
    if not s:
        return pd.NaT
    try:
        dt = _dateutil_parser.parse(s, dayfirst=False, yearfirst=False)
        return pd.Timestamp(dt)
    except (ValueError, TypeError, OverflowError, OSError):
        return pd.NaT


def _excel_dates_to_datetime(arg, **kwargs):
    """
    Parse Excel/date columns for the dashboard. Ambiguous strings such as 01/02/2024
    are read as month/day/year (US order): January 2, 2024 — not February 1.
    After the generic pass, any still-missing values in a Series are retried with
    explicit US (m/d/Y) and ISO formats so one column keeps one interpretation.
    ISO datetimes and Excel serials from the first pass stay as-is. Pass dayfirst=True
    to force European order for the whole column instead.
    Remaining NaT strings (e.g. ``1/26/2026 12:15:00 PM`` with odd spacing) are retried
    with US strptime formats first, then ``dateutil.parser`` as a last resort.
    """
    kw = dict(kwargs)
    kw.setdefault("errors", "coerce")
    kw.setdefault("dayfirst", False)
    out = pd.to_datetime(arg, **kw)
    if not isinstance(arg, pd.Series):
        try:
            if pd.isna(out):
                t2 = _try_dateutil_parse(arg)
                if pd.notna(t2):
                    return t2
        except (TypeError, ValueError):
            pass
        return out
    if kw.get("dayfirst", False) is not True:
        need = out.isna() & arg.notna()
        if need.any():
            s = (
                arg.loc[need]
                .map(lambda x: _strip_excel_date_str(x) if pd.notna(x) else "")
                .replace("", pd.NA)
            )
            s = s.dropna()
            if not s.empty:
                acc = pd.Series(pd.NaT, index=s.index, dtype="datetime64[ns]")
                for fmt in (
                    "%m/%d/%Y %I:%M:%S %p",
                    "%m/%d/%Y %I:%M %p",
                    "%m/%d/%Y %H:%M:%S",
                    "%m/%d/%Y %H:%M",
                    "%m/%d/%Y",
                    "%m/%d/%y %I:%M:%S %p",
                    "%m/%d/%y %H:%M:%S",
                    "%m/%d/%y",
                    "%Y-%m-%d %H:%M:%S",
                    "%Y-%m-%d",
                ):
                    part = pd.to_datetime(s, format=fmt, errors="coerce")
                    acc = acc.where(acc.notna(), part)
                out = out.copy()
                sub = out.reindex(acc.index)
                out.loc[acc.index] = acc.combine_first(sub)
        loose = out.isna() & arg.notna()
        if loose.any():
            extra = arg.loc[loose].map(_try_dateutil_parse)
            out = out.copy()
            out.loc[loose] = extra.where(extra.notna(), out.loc[loose])
    return out


def make_json_serializable(df):

    def convert_value(x):
        if isinstance(x, (pd.Timestamp, pd.Timedelta)):
            return x.isoformat()
        elif isinstance(x, (datetime.datetime, datetime.date, datetime.time)):
            return x.isoformat()
        elif isinstance(x, (np.int64, np.int32)):
            return int(x)
        elif isinstance(x, (np.float64, np.float32)):
            return float(x)
        elif isinstance(x, (np.ndarray, list, dict)):
            return str(x)
        else:
            return x

    return df.applymap(convert_value)


def _sanitize_for_json(obj):
    """Convert numpy/pandas types to native Python for JsonResponse."""
    if obj is None or isinstance(obj, (bool, str)):
        return obj
    if isinstance(obj, np.ndarray):
        return [_sanitize_for_json(v) for v in obj.tolist()]
    if isinstance(obj, (np.integer, np.int64, np.int32)):
        return int(obj)
    if isinstance(obj, (np.floating, np.float64, np.float32)):
        try:
            v = float(obj)
            if np.isnan(v) or np.isinf(v):
                return None
            return v
        except (ValueError, TypeError):
            return None
    if isinstance(obj, (pd.Timestamp, pd.Timedelta, datetime.datetime, datetime.date)):
        return obj.isoformat() if hasattr(obj, "isoformat") else str(obj)
    if isinstance(obj, (int, float)) and (obj != obj or abs(obj) == float("inf")):
        return None  # NaN or Inf
    try:
        if pd.isna(obj) and not isinstance(obj, (dict, list, tuple)):
            return None
    except (ValueError, TypeError):
        pass
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_sanitize_for_json(v) for v in obj]
    return obj


def _normalize_upload_to_latest_xlsx_and_update_cache(file_path, folder_path):
    """
    إذا كان الملف أصلاً .xlsx يُترك كما هو (رفع سريع).
    إذا كان .xlsm يُحوّل إلى latest.xlsx عبر ملف مؤقت.
    لا يتم تعبئة الكاش هنا؛ يتم مسح الكاش بعد الرفع وتعبئته عند أول قراءة لتاب.
    """
    if not file_path or not os.path.exists(file_path):
        return file_path
    out_path = os.path.join(folder_path, "latest.xlsx")
    ext = (os.path.splitext(file_path)[1] or "").lower()
    # ملف .xlsx محفوظ فعلياً كـ latest.xlsx — لا نعيد قراءة/كتابة (توفير وقت كبير)
    if ext == ".xlsx" and os.path.abspath(file_path) == os.path.abspath(out_path):
        print("✅ [DEBUG] الملف .xlsx — تم الاحتفاظ به كما هو (رفع سريع)")
        return file_path

    if ext != ".xlsm":
        return file_path

    temp_path = os.path.join(folder_path, "latest_temp_xlsx_upload.xlsx")
    try:
        xls = pd.ExcelFile(file_path, engine="openpyxl")
        sheet_names = list(xls.sheet_names)
        if not sheet_names:
            return file_path
        with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
            for sheet in sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet)
                df.to_excel(writer, sheet_name=sheet, index=False)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(temp_path)
            if wb.sheetnames and hasattr(wb[wb.sheetnames[0]], "sheet_state"):
                wb[wb.sheetnames[0]].sheet_state = "visible"
            wb.save(temp_path)
        except Exception:
            pass
        os.replace(temp_path, out_path)
        print("✅ [DEBUG] تم تحويل .xlsm إلى latest.xlsx")
        return out_path
    except Exception as e:
        print(f"⚠️ [DEBUG] normalize: {e}")
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass
        return file_path


_MERGE_KEY_EXACT = [
    "ORDER / SO",
    "ORDER/SO",
    "Order / SO",
    "SO",
    "Order Number",
    "Order Nbr",
    "Order No",
    "LPN Nbr",
    "LPN NBR",
    "Item Code",
    "ITEM CODE",
    "Batch Nbr",
    "Shipment ID",
    "AWB",
    "Tracking Number",
    "ASN",
]

_MERGE_KEY_SUBSTRINGS = [
    "order nbr",
    "order number",
    "order/so",
    "lpn nbr",
    "item code",
    "batch nbr",
    "shipment id",
    "tracking",
    "awb",
    "invoice",
    "asn",
]


def _find_sheet_merge_key_column(df):
    """عمود مميّز لدمج صفوف الشيت (الرفع الجديد يفوز عند التكرار)."""
    if df is None or df.empty:
        return None
    cols = list(df.columns)
    for cand in _MERGE_KEY_EXACT:
        cl = cand.lower()
        for c in cols:
            if str(c).strip().lower() == cl:
                return c
    for cand in _MERGE_KEY_EXACT:
        cl = cand.lower()
        for c in cols:
            if cl in str(c).strip().lower():
                return c
    for sub in _MERGE_KEY_SUBSTRINGS:
        for c in cols:
            if sub in str(c).strip().lower():
                return c
    return None


def _strip_empty_rows(df):
    """
    يحذف الصفوف اللي كل خلاياها فاضية (NaN / نص فاضي) ويعيد ترقيم الصفوف في الـ DataFrame.
    عند الحفظ بـ pandas لا تُكتب صفوف فارغة في الإكسل.
    """
    if df is None:
        return pd.DataFrame()
    original_cols = df.columns
    d = df.copy()
    if d.empty:
        return pd.DataFrame(columns=original_cols)
    d.columns = d.columns.astype(str)

    def _cell_blank(x):
        if x is None:
            return True
        try:
            if pd.isna(x):
                return True
        except (ValueError, TypeError):
            pass
        if isinstance(x, str) and not str(x).strip():
            return True
        return False

    keep = ~d.apply(lambda row: all(_cell_blank(v) for v in row), axis=1)
    d = d.loc[keep].reset_index(drop=True)
    if d.empty:
        return pd.DataFrame(columns=original_cols)
    return d


def _merge_two_sheet_dfs(old_df, new_df):
    if old_df is None or getattr(old_df, "empty", True):
        return new_df.copy() if new_df is not None else pd.DataFrame()
    if new_df is None or getattr(new_df, "empty", True):
        return old_df.copy()
    o = old_df.copy()
    n = new_df.copy()
    o.columns = o.columns.astype(str).str.strip()
    n.columns = n.columns.astype(str).str.strip()
    key = _find_sheet_merge_key_column(n) or _find_sheet_merge_key_column(o)
    all_cols = list(
        dict.fromkeys(list(n.columns) + [c for c in o.columns if c not in n.columns])
    )
    oa = o.reindex(columns=all_cols)
    na = n.reindex(columns=all_cols)
    if key and key in all_cols:
        combined = pd.concat([oa, na], ignore_index=True)
        k = combined[key].astype(str).str.strip()
        combined = combined.loc[
            ~k.isin(("", "nan", "none", "<na>", "nat", "NaT"))
        ].copy()
        combined = combined.drop_duplicates(subset=[key], keep="last")
    else:
        combined = pd.concat([oa, na], ignore_index=True).drop_duplicates()
    return combined


def _try_remove_libreoffice_lock_files(folder_path):
    """
    LibreOffice ينشئ ملفات مثل .~lock.latest.xlsx# في نفس مجلد الملف المفتوح.
    حذفها آمن (نص صغير) وقد يزيل تعارض الحذف/الاستبدال على لينكس.
    """
    for fname in (".~lock.latest.xlsx#", ".~lock.latest.xlsm#"):
        p = os.path.join(folder_path, fname)
        if not os.path.isfile(p):
            continue
        try:
            os.remove(p)
            print(f"🗑️ [upload] تم حذف ملف قفل LibreOffice: {fname}")
        except OSError as e:
            print(f"⚠️ [upload] تعذر حذف ملف القفل {fname}: {e}")


def _write_django_uploaded_file_to_disk(uploaded_file, dest_path):
    """
    حفظ موثوق لملف الرفع: الملفات الأكبر من حد الذاكرة تكون TemporaryUploadedFile
    ويُفضّل النسخ من temporary_file_path() بدل copyfileobj/chunks حتى لا يُكتب ملف فارغ.
    """
    tfp = getattr(uploaded_file, "temporary_file_path", None)
    if callable(tfp):
        try:
            src = uploaded_file.temporary_file_path()
            if src and os.path.isfile(src):
                try:
                    shutil.copy2(src, dest_path)
                    return
                except OSError as e:
                    if getattr(e, "errno", None) in (13, 1):
                        raise RuntimeError(
                            "لا يمكن الكتابة في مجلد الرفع (صلاحيات). أعد ملكية المجلد لمستخدم السيرفر، مثلاً: "
                            'sudo chown -R "$USER":"$USER" media/excel_uploads'
                        ) from e
                    raise
        except RuntimeError:
            raise
        except Exception as e:
            print(f"⚠️ [upload] نسخ من temporary_file_path فشل، نستخدم chunks: {e}")

    if hasattr(uploaded_file, "seek"):
        try:
            uploaded_file.seek(0)
        except (OSError, AttributeError, TypeError, ValueError):
            pass
    try:
        with open(dest_path, "wb") as out:
            for chunk in uploaded_file.chunks():
                out.write(chunk)
    except OSError as e:
        if getattr(e, "errno", None) in (13, 1):
            raise RuntimeError(
                "لا يمكن الكتابة في مجلد الرفع (صلاحيات). أعد ملكية المجلد لمستخدم السيرفر، مثلاً: "
                'sudo chown -R "$USER":"$USER" media/excel_uploads'
            ) from e
        raise


def _existing_main_workbook_before_upload(folder_path):
    """آخر ملف رئيسي قبل الاستبدال (latest.xlsx ثم latest.xlsm)."""
    for name in ("latest.xlsx", "latest.xlsm"):
        p = os.path.join(folder_path, name)
        if os.path.isfile(p):
            return p
    return None


def _merge_clean_excel_workbook(target_path, previous_path=None):
    """
    - إن وُجد previous_path: يدمج كل الشيتات ذات الاسم بين الملفين (الجديد يفوز عند تكرار المفتاح)،
      ويُلحق شيتات موجودة في القديم فقط.
    - دائماً: يحذف الصفوف الفارغة من كل شيت ويعيد كتابة الملف (صفوف متتالية بدون فراغات).
    """
    if not target_path or not os.path.isfile(target_path):
        return
    tmp_path = os.path.join(
        os.path.dirname(target_path) or ".",
        f"_workbook_merge_clean_{uuid.uuid4().hex}.xlsx",
    )
    try:
        new_xls = pd.ExcelFile(target_path, engine="openpyxl")
    except Exception as e:
        print(f"⚠️ [Excel merge/clean] لا يمكن قراءة الملف الجديد: {e}")
        return

    new_names = list(new_xls.sheet_names)
    old_by_name = {}
    if previous_path and os.path.isfile(previous_path):
        try:
            old_xls = pd.ExcelFile(previous_path, engine="openpyxl")
            for sn in old_xls.sheet_names:
                try:
                    odf = pd.read_excel(
                        previous_path, sheet_name=sn, engine="openpyxl", header=0
                    )
                    odf.columns = odf.columns.astype(str).str.strip()
                    old_by_name[str(sn)] = odf
                except Exception as se:
                    print(f"⚠️ [Excel merge/clean] تخطّي شيت قديم '{sn}': {se}")
        except Exception as e:
            print(f"⚠️ [Excel merge/clean] لا يمكن قراءة الملف القديم للدمج: {e}")

    new_name_set = {str(x) for x in new_names}
    ordered = list(new_names) + [s for s in old_by_name if s not in new_name_set]

    def _excel_safe_sheet_title(name):
        s = str(name or "Sheet").strip() or "Sheet"
        return s[:31]

    try:
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            for sn in ordered:
                snk = str(sn)
                if snk in new_name_set:
                    try:
                        new_df = pd.read_excel(
                            target_path, sheet_name=sn, engine="openpyxl", header=0
                        )
                        new_df.columns = new_df.columns.astype(str).str.strip()
                    except Exception as re:
                        print(f"⚠️ [Excel merge/clean] شيت جديد '{sn}': {re}")
                        continue
                else:
                    new_df = None

                if snk in new_name_set and snk in old_by_name:
                    try:
                        out = _merge_two_sheet_dfs(old_by_name[snk], new_df)
                    except Exception as me:
                        print(f"⚠️ [Excel merge/clean] دمج شيت '{sn}': {me}")
                        out = new_df.copy()
                elif snk in new_name_set:
                    out = new_df.copy()
                elif snk in old_by_name:
                    out = old_by_name[snk].copy()
                else:
                    continue

                out = _strip_empty_rows(out)
                out.to_excel(
                    writer, sheet_name=_excel_safe_sheet_title(sn), index=False
                )
        os.replace(tmp_path, target_path)
        print(
            f"✅ [Excel merge/clean] تمت المعالجة: شيتات={len(ordered)} "
            f"{'(مع دمج)' if old_by_name else '(تنظيف فقط)'}"
        )
    except Exception as e:
        print(f"⚠️ [Excel merge/clean] فشل إعادة الكتابة: {e}")
        if os.path.isfile(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


def _excel_full_data_requested(request):
    """الوضع الكامل: GET/POST full_data=1 أو true أو all."""
    if not request:
        return False
    v = (request.GET.get("full_data") or request.POST.get("full_data") or "").strip().lower()
    return v in ("1", "true", "yes", "all", "full")


def _excel_max_rows_for_request(request, *, force_full=False):
    """
    معاينة سريعة (EXCEL_PREVIEW_MAX_ROWS) إلا إذا طُلب full_data أو force_full (مثلاً Traceability).
    EXCEL_FULL_MAX_ROWS = None يعني بدون حد (كل الشيت).
    """
    if force_full:
        cap = getattr(settings, "EXCEL_FULL_MAX_ROWS", None)
        return cap
    if request is not None and _excel_full_data_requested(request):
        cap = getattr(settings, "EXCEL_FULL_MAX_ROWS", None)
        return cap
    try:
        return int(getattr(settings, "EXCEL_PREVIEW_MAX_ROWS", 200))
    except (TypeError, ValueError):
        return 200


def _read_excel_nrows_kw(max_rows):
    """وسيط لـ pandas.read_excel: لا تمرّر nrows إن كان الحد = None (كل الصفوف)."""
    if max_rows is None:
        return {}
    try:
        n = int(max_rows)
        if n <= 0:
            return {}
        return {"nrows": n}
    except (TypeError, ValueError):
        return {}


def _get_sheet_dataframe(
    excel_path, sheet_name, use_cache=True, max_rows=None, request=None, force_full=False
):
    """
    يرجع DataFrame للشيت مع كاش Django لتقليل قراءات الديسك.
    max_rows صريح يتجاوز الطلب؛ وإلا يُستخدم full_data / معاينة من request.
    """
    if not excel_path or not os.path.exists(excel_path):
        return None
    if max_rows is None:
        if request is not None:
            max_rows = _excel_max_rows_for_request(request, force_full=force_full)
        else:
            try:
                max_rows = int(getattr(settings, "EXCEL_PREVIEW_MAX_ROWS", 200))
            except (TypeError, ValueError):
                max_rows = 200
    try:
        from django.core.cache import cache as _dj_cache
        import hashlib as _hashlib

        _path_hash = _hashlib.md5((excel_path or "").encode()).hexdigest()[:12]
        _nkey = "all" if max_rows is None else int(max_rows)
        cache_key = f"excel_df::{_path_hash}::{sheet_name}::n{_nkey}"
        if use_cache:
            try:
                cached_df = _dj_cache.get(cache_key)
                if cached_df is not None:
                    return cached_df.copy()
            except Exception as e:
                print(f"⚠️ [Cache-MEM] قراءة الشيت من الكاش فشلت '{sheet_name}': {e}")

        read_kw = {"engine": "openpyxl", "header": 0}
        read_kw.update(_read_excel_nrows_kw(max_rows))
        df = pd.read_excel(excel_path, sheet_name=sheet_name, **read_kw)
        df.columns = [str(c).strip() for c in df.columns]

        if use_cache:
            try:
                _dj_cache.set(cache_key, df, 3600)
            except Exception as e:
                print(f"⚠️ [Cache-MEM] حفظ الشيت في الكاش فشل '{sheet_name}': {e}")

        return df
    except Exception:
        return None


def _get_excel_path_for_request(request):
    """يرجع مسار ملف الإكسل المرفوع من الجلسة أو المجلد الافتراضي."""
    if not request:
        return None
    folder = os.path.abspath(
        os.path.join(str(settings.MEDIA_ROOT), "excel_uploads")
    )
    if not os.path.isdir(folder):
        return None
    path = request.session.get("uploaded_excel_path")
    if path and os.path.isfile(path):
        return path
    # الملف الرئيسي لكل التابات (ماعدا Dashboard):
    # نجرّب أولاً ملف Nespresso الجديد ثم الأسماء القديمة للتوافق
    priority_files = [
        "all_sheet_nespresso.xlsx",
        "all_sheet_nespresso.xlsm",
        "all sheet nespresso.xlsx",
        "all sheet nespresso.xlsm",
        "latest.xlsm",
        "latest.xlsx",
        "all_sheet.xlsx",
        "all_sheet.xlsm",
        "all sheet.xlsx",
        "all sheet.xlsm",
    ]
    for name in priority_files:
        p = os.path.join(folder, name)
        if os.path.isfile(p):
            return p
    return None


def _excel_file_signature(excel_path):
    """mtime + حجم الملف — يتغيّر عند استبدال الملف فيُبطَل الكاش."""
    try:
        st = os.stat(excel_path)
        return f"{int(st.st_mtime)}_{st.st_size}"
    except OSError:
        return "0_0"


def _list_excel_sheet_names_openpyxl(excel_path):
    """أسماء الشيتات فقط: read_only يقرأ بنية الملف بسرعة دون تحميل كل الخلايا."""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        return [str(s).strip() for s in wb.sheetnames]
    finally:
        wb.close()


def _get_excel_sheet_names_cached(excel_path):
    """قائمة أسماء الشيتات مع كاش Django (تفادي تكرار فتح الملف على كل GET)."""
    if not excel_path or not os.path.isfile(excel_path):
        return []
    sig = _excel_file_signature(excel_path)
    path_key = hashlib.md5(os.path.abspath(excel_path).encode()).hexdigest()[:12]
    cache_key = f"excel_sheet_names::{path_key}::{sig}"
    cached = cache.get(cache_key)
    if cached is not None:
        return list(cached)
    try:
        names = _list_excel_sheet_names_openpyxl(excel_path)
    except Exception:
        try:
            xls = pd.ExcelFile(excel_path, engine="openpyxl")
            names = [str(s).strip() for s in xls.sheet_names]
        except Exception:
            names = []
    try:
        cache.set(cache_key, names, 86400)
    except Exception:
        pass
    return names


def _extract_months_from_excel_cached(excel_path, sheet_names):
    """استخراج الشهور من أو أول شيتين فيهما أعمدة تاريخ/شهر — مع كاش."""
    if not excel_path or not sheet_names:
        return []
    sig = _excel_file_signature(excel_path)
    path_key = hashlib.md5(os.path.abspath(excel_path).encode()).hexdigest()[:12]
    cache_key = f"excel_months::{path_key}::{sig}"
    cached = cache.get(cache_key)
    if cached is not None:
        return list(cached)

    all_months = []
    _max_rows = min(
        300, int(getattr(settings, "EXCEL_PREVIEW_MAX_ROWS", 200))
    )
    _max_sheets_for_months = 2
    _sheets_read = 0
    _month_order = [
        "Jan", "Feb", "Mar", "Apr", "May", "Jun",
        "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
    ]
    try:
        for sheet in sheet_names:
            if _sheets_read >= _max_sheets_for_months:
                break
            try:
                df = pd.read_excel(
                    excel_path, sheet_name=sheet, engine="openpyxl", nrows=_max_rows
                )
                df.columns = df.columns.str.strip().str.title()
                possible_date_cols = [
                    c for c in df.columns
                    if "date" in c.lower() or "month" in c.lower()
                ]
                if not possible_date_cols:
                    continue
                _sheets_read += 1
                col = possible_date_cols[0]
                df[col] = _excel_dates_to_datetime(df[col], errors="coerce")
                df["MonthName"] = df[col].dt.strftime("%b")
                seen = set(df["MonthName"].dropna().unique().tolist())
                all_months = [m for m in _month_order if m in seen]
                if all_months:
                    break
            except Exception:
                continue
        print(
            "📅 [INFO] الشهور (من شيت واحد أو اثنين):",
            all_months[:12] if len(all_months) > 12 else all_months,
        )
    except Exception as e:
        print("⚠️ [ERROR] أثناء استخراج الشهور:", e)
        all_months = []

    try:
        cache.set(cache_key, all_months, 86400)
    except Exception:
        pass
    return all_months


_MONTH_ABBR_ORDER = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]
_MONTH_ABBR_RANK = {m: i for i, m in enumerate(_MONTH_ABBR_ORDER)}


def _dt_series_to_month_abbr_en(dt_series):
    """
    Map datetimes to English Jan..Dec. Does not use strftime('%b'), which follows
    LC_TIME and can mismatch our fixed month lists on non-English servers.
    """
    if dt_series is None:
        return pd.Series(dtype=object)
    m = pd.to_numeric(dt_series.dt.month, errors="coerce")
    labels = np.array(_MONTH_ABBR_ORDER, dtype=object)
    out = pd.Series(pd.NA, index=dt_series.index, dtype=object)
    valid = m.notna() & (m >= 1) & (m <= 12)
    if valid.any():
        out.loc[valid] = labels[(m.loc[valid].astype(int).values - 1)]
    return out


def _sorted_unique_month_abbrs(months_set):
    return sorted((m for m in months_set if m in _MONTH_ABBR_RANK), key=lambda x: _MONTH_ABBR_RANK.get(x, 99))


def _resolve_workbook_sheet_name(excel_path, sheet_guess):
    if not excel_path or not sheet_guess or not os.path.isfile(excel_path):
        return None
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
        names = list(xls.sheet_names)
        if sheet_guess in names:
            return sheet_guess
        g = str(sheet_guess).strip().lower().replace(" ", "_")
        for s in names:
            if (s or "").strip().lower() == str(sheet_guess).strip().lower():
                return s
            if (s or "").strip().replace(" ", "_").lower() == g:
                return s
    except Exception:
        return None
    return None


def _inbound_sheet_name_for_months(excel_path):
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")

        def _n(s):
            return re.sub(r"[^a-z0-9]", "", (str(s) or "").strip().lower())

        for s in xls.sheet_names:
            if _n(s) == "inboundtab":
                return s
        for s in xls.sheet_names:
            if "ARAMCO Inbound Report" in (s or "").strip():
                return s
        for s in xls.sheet_names:
            if "inbound" in (s or "").lower():
                return s
    except Exception:
        pass
    return None


def _total_lead_time_sheet_for_months(excel_path):
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
        for s in xls.sheet_names:
            sl = (s or "").lower()
            if "total lead time preformance" in sl and "-r" not in sl:
                return s
    except Exception:
        pass
    return None


def _capacity_expiry_sheet_for_months(excel_path):
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")

        def _norm(s):
            return (
                str(s).strip().lower().replace(" ", "").replace("+", "").replace("_", "")
                if s
                else ""
            )

        for s in xls.sheet_names:
            if _norm(s) == "capacityexpirytab":
                return s
        for s in xls.sheet_names:
            if "capacity" in _norm(s) and "expiry" in _norm(s):
                return s
        for s in xls.sheet_names:
            if (s or "").strip().lower() == "expiry":
                return s
        for s in xls.sheet_names:
            if "expiry" in (s or "").lower():
                return s
    except Exception:
        pass
    return None


def _distinct_month_abbrs_from_sheet(
    excel_path,
    sheet_guess,
    *,
    preferred_columns=None,
    max_rows=None,
):
    """شهور مختصرة (Jan..Dec) من شيت واحد حسب أعمدة تاريخ أو قائمة أعمدة مفضلة."""
    sn = _resolve_workbook_sheet_name(excel_path, sheet_guess)
    if not sn:
        return []
    try:
        read_kw = {"engine": "openpyxl", "header": 0}
        mr = max_rows
        if mr is None:
            mr = getattr(settings, "EXCEL_FULL_MAX_ROWS", None)
        if mr is not None:
            try:
                ni = int(mr)
                if ni > 0:
                    read_kw["nrows"] = ni
            except (TypeError, ValueError):
                pass
        df = pd.read_excel(excel_path, sheet_name=sn, **read_kw)
        df.columns = df.columns.astype(str).str.strip()
        months = set()
        cols_try = []
        if preferred_columns:
            for pref in preferred_columns:
                pl = str(pref).lower()
                hit = None
                for c in df.columns:
                    if str(c).strip().lower() == pl:
                        hit = c
                        break
                if hit:
                    cols_try.append(hit)
                else:
                    for c in df.columns:
                        if pl in str(c).lower():
                            cols_try.append(c)
                            break
        for c in df.columns:
            if c in cols_try:
                continue
            cl = str(c).lower()
            if any(
                k in cl
                for k in (
                    "date",
                    "time",
                    "timestamp",
                    "month",
                )
            ):
                cols_try.append(c)
        seen = set()
        for c in cols_try:
            if c in seen or c not in df.columns:
                continue
            seen.add(c)
            try:
                ser = _excel_dates_to_datetime(df[c], errors="coerce")
            except Exception:
                continue
            if ser.notna().sum() == 0:
                continue
            for ab in _dt_series_to_month_abbr_en(ser).dropna().unique().tolist():
                if ab in _MONTH_ABBR_RANK:
                    months.add(ab)
        return _sorted_unique_month_abbrs(months)
    except Exception as e:
        print(f"⚠️ [_distinct_month_abbrs_from_sheet] {sheet_guess}: {e}")
        return []


# اسم ملف الداشبورد الثابت (شيت تاني للتاب Dashboard فقط)
DASHBOARD_EXCEL_FILENAME = "Aramco_Tamer3PL_KPI_Dashboard.xlsx"

# داتا Inbound الافتراضية (للكروت والشارت) — نفس فكرة chart_data في rejection
INBOUND_DEFAULT_KPI = {
    "number_of_vehicles": 12,
    "number_of_shipments": 287,
    "number_of_pallets": 1105,
    "total_quantity": 65400,
    "total_quantity_display": "65.4k",
}
# الداتا اللي بتظهر على شارت Pending Shipments (label, value, pct, color)
INBOUND_DEFAULT_PENDING_SHIPMENTS = [
    {"label": "In Transit", "value": "1%", "pct": 1, "color": "#87CEEB"},
    {"label": "Receiving Complete", "value": "96%", "pct": 96, "color": "#2E7D32"},
    {"label": "Verified", "value": "3%", "pct": 3, "color": "#1565C0"},
]

# داتا الشارتات الافتراضية للداشبورد (نفس فكرة chart_data في rejection — لو مفيش إكسل نستخدمها)
DASHBOARD_DEFAULT_CHART_DATA = {
    "outbound_chart_data": {
        "categories": ["Jan", "Feb", "Mar", "Apr", "May", "Jun"],
        "series": [40, 55, 48, 62, 58, 70],
    },
    "returns_chart_data": {
        "categories": ["Mar", "Apr", "May", "Jun", "Jul", "Aug"],
        "series": [280, 320, 300, 350, 380, 400],
    },
    "inventory_capacity_data": {"used": 78, "available": 22},
}


def _read_dashboard_charts_from_excel(excel_path, request=None):
    """
    يقرأ داتا الشارتات (Outbound, Returns, Inventory) من ملف الداشبورد لو الشيتات موجودة.
    ترجع ديكت باللي اتقرا فقط (لو مفيش داتا للشارت ترجع None للكاي) — عشان نعمل الشارتات دينامك.
    """
    result = {}
    _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
    except Exception:
        return result
    sheet_names = [str(s).strip() for s in xls.sheet_names]

    # Outbound: من شيت Outbound_Data أو Outbound — تجميع حسب شهر لو فيه عمود شهر/تاريخ
    for out_name in ["Outbound_Data", "Outbound Data", "Outbound"]:
        if not any(out_name.lower().replace(" ", "") in s.lower().replace(" ", "") for s in sheet_names):
            continue
        sheet_name = next((s for s in sheet_names if out_name.lower() in s.lower()), None)
        if not sheet_name:
            continue
        try:
            df = pd.read_excel(
                excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
            )
            if df.empty or len(df) < 2:
                break
            df.columns = [str(c).strip() for c in df.columns]
            cols_lower = {c.lower(): c for c in df.columns}
            month_col = None
            for c in cols_lower:
                if "month" in c or "date" in c:
                    month_col = cols_lower[c]
                    break
            if month_col:
                df["_m"] = _excel_dates_to_datetime(df[month_col], errors="coerce").dt.strftime("%b")
                by_month = df.groupby("_m").size().reindex(
                    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
                ).dropna()
                if not by_month.empty:
                    result["outbound_chart_data"] = {
                        "categories": by_month.index.tolist(),
                        "series": by_month.values.tolist(),
                    }
            break
        except Exception:
            break

    # Returns: من شيت Return أو Rejection
    for ret_name in ["Return", "Rejection", "Returns"]:
        sheet_name = next((s for s in sheet_names if ret_name.lower() in s.lower()), None)
        if not sheet_name:
            continue
        try:
            df = pd.read_excel(
                excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
            )
            if df.empty or len(df) < 2:
                break
            df.columns = [str(c).strip() for c in df.columns]
            month_col = next((c for c in df.columns if "month" in c.lower()), None)
            val_col = next(
                (c for c in df.columns if "order" in c.lower() or "booking" in c.lower() or "count" in c.lower()),
                df.columns[1] if len(df.columns) > 1 else None,
            )
            if month_col and val_col:
                summary = df[[month_col, val_col]].dropna()
                if not summary.empty:
                    try:
                        summary[val_col] = pd.to_numeric(summary[val_col].astype(str).str.replace("%", "", regex=False), errors="coerce")
                        summary = summary.dropna(subset=[val_col])
                        categories = summary[month_col].astype(str).tolist()
                        series = summary[val_col].astype(int).tolist()
                        if categories and series:
                            result["returns_chart_data"] = {"categories": categories, "series": series}
                    except Exception:
                        pass
            break
        except Exception:
            break

    # Inventory capacity: من شيت Inventory أو Capacity
    for inv_name in ["Inventory", "Capacity", "Warehouse"]:
        sheet_name = next((s for s in sheet_names if inv_name.lower() in s.lower()), None)
        if not sheet_name:
            continue
        try:
            df = pd.read_excel(
                excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
            )
            if df.empty:
                break
            df.columns = [str(c).strip() for c in df.columns]
            used_col = next((c for c in df.columns if "used" in c.lower() or "utilization" in c.lower()), None)
            if used_col:
                vals = pd.to_numeric(df[used_col], errors="coerce").dropna()
                if len(vals) > 0:
                    used = int(min(100, max(0, vals.mean())))
                    result["inventory_capacity_data"] = {"used": used, "available": 100 - used}
            break
        except Exception:
            break

    return result


def _get_dashboard_excel_path(request):
    """
    يرجّع مسار ملف إكسل الداشبورد (Aramco_Tamer3PL_KPI_Dashboard.xlsx) إن وُجد.
    مصدر الداتا لتاب Dashboard فقط؛ باقي التابات من الملف الرئيسي (all_sheet / latest).
    """
    if not request:
        return None
    folder = os.path.join(settings.MEDIA_ROOT, "excel_uploads")
    path = request.session.get("dashboard_excel_path")
    if path and os.path.isfile(path):
        return path
    p = os.path.join(folder, DASHBOARD_EXCEL_FILENAME)
    if os.path.isfile(p):
        try:
            request.session["dashboard_excel_path"] = p
            request.session.save()
        except Exception:
            pass
        return p
    return None


def _is_dashboard_excel_filename(name):
    """يعرف إذا الملف المرفوع هو ملف الداشبورد (شيت تاني)."""
    if not name:
        return False
    n = (name or "").strip().lower()
    return "kpi_dashboard" in n or "aramco_tamer3pl" in n


def _read_inbound_data_from_excel(excel_path, request=None):
    """
    يقرأ بيانات Inbound (KPI + Pending Shipments) من ملف الإكسل.
    الشيت: "Inbound" أو أول شيت اسمه يحتوي "inbound".
    أعمدة الـ KPI (حسب الطلب):
    - Vehicle_ID: كل يوم بيومه نشيل المتكرر (unique per day) ثم نجمع عدد المركبات لكل الأيام → Number of Vehicles
    - Shipment_ID: نفس المنطق يوم بيوم unique ثم جمع → Number of Shipments
    - Nbr_LPNs: مجموع كل القيم (27+18+13+...) → Number of Pallets (LPNs)
    - Total_Qty: مجموع كل القيم → Total Quantity
    عمود التاريخ: أي عمود اسمه يحتوي date/receipt/shipment date (للتجميع يوم بيوم).
    إن لم يوجد عمود تاريخ، نعتبر كل البيانات يوم واحد.
    Pending Shipments: إن وُجدت أعمدة Label/Status, Value, Pct, Color في نفس الشيت أو شيت آخر نستخدمها.
    """
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
    except Exception:
        return None
    sheet_name = None
    for name in xls.sheet_names:
        if "inbound" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = xls.sheet_names[0] if xls.sheet_names else None
    if not sheet_name:
        return None
    try:
        _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
        )
    except Exception:
        return None
    if df.empty or len(df) < 1:
        return None

    # تطبيع أسماء الأعمدة: strip + lower للبحث
    df.columns = [str(c).strip() for c in df.columns]
    cols_lower = {c.lower(): c for c in df.columns if c}

    def _col(*keys):
        for k in keys:
            if k.lower() in cols_lower:
                return cols_lower[k.lower()]
        return None

    # عمود التاريخ (للتجميع يوم بيوم)
    date_col = None
    for c in df.columns:
        cl = c.lower()
        if "date" in cl or "receipt" in cl or ("shipment" in cl and "date" in cl) or cl == "day":
            date_col = c
            break
    if not date_col and df.columns.size > 0:
        for c in df.columns:
            try:
                _excel_dates_to_datetime(df[c].dropna().head(20), errors="coerce")
                date_col = c
                break
            except Exception:
                continue

    vehicle_col = _col("Vehicle_ID", "Vehicle ID", "Vehicle_ID")
    shipment_col = _col("Shipment_ID", "Shipment ID", "Shipment_ID")
    lpn_col = _col("Nbr_LPNs", "Nbr LPNs", "LPNs")
    qty_col = _col("Total_Qty", "Total_Qty", "Total Qty", "Total_Qty")

    def _to_int(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return None

    n_vehicles = 0
    n_shipments = 0
    n_pallets = 0
    n_qty = 0

    if date_col and (vehicle_col or shipment_col):
        # تجميع يوم بيوم
        df_date = df.copy()
        df_date["_date"] = _excel_dates_to_datetime(df_date[date_col], errors="coerce")
        df_date = df_date.dropna(subset=["_date"])
        df_date["_day"] = df_date["_date"].dt.normalize()

        if vehicle_col:
            # كل يوم: عدد الـ Vehicle_ID المميزة، ثم نجمع كل الأيام
            per_day_vehicles = df_date.groupby("_day")[vehicle_col].nunique()
            n_vehicles = int(per_day_vehicles.sum())
        if shipment_col:
            # كل يوم: عدد الـ Shipment_ID المميزة، ثم نجمع كل الأيام
            per_day_shipments = df_date.groupby("_day")[shipment_col].nunique()
            n_shipments = int(per_day_shipments.sum())
    else:
        # بدون تاريخ: نعتبر كل الصفوف يوم واحد (unique للمركبات والشحنات)
        if vehicle_col:
            n_vehicles = int(df[vehicle_col].nunique())
        if shipment_col:
            n_shipments = int(df[shipment_col].nunique())

    if lpn_col:
        n_pallets = _to_int(df[lpn_col].sum()) or 0
    if qty_col:
        n_qty = _to_int(df[qty_col].sum()) or 0

    # لا نضع قيم يدوية — لو مفيش عمود القيمة تبقى 0 (الداتا من الشيت فقط)
    # if not vehicle_col: n_vehicles = 0  (already 0)
    # if not shipment_col: n_shipments = 0
    # if not lpn_col: n_pallets = 0
    # if not qty_col: n_qty = 0

    if n_qty >= 1000:
        qty_display = f"{n_qty / 1000:.1f}k".rstrip("0").rstrip(".")
        if not qty_display.endswith("k"):
            qty_display += "k"
    else:
        qty_display = str(n_qty)

    inbound_kpi = {
        "number_of_vehicles": n_vehicles,
        "number_of_shipments": n_shipments,
        "number_of_pallets": n_pallets,
        "total_quantity": n_qty,
        "total_quantity_display": qty_display,
    }

    # Pending Shipments: من عمود Status في نفس شيت Inbound — In Transit, Receiving Complete, Verified
    # يوم بيوم نجمع عدد الشحنات لكل حالة ثم نجمع التوتال، ثم النسبة = (عدد الحالة / التوتال) * 100
    pending = []
    status_col = _col("Status", "status")
    STATUS_LABELS = (
        ("in transit", "In Transit", "#87CEEB"),
        ("receiving complete", "Receiving Complete", "#2E7D32"),
        ("verified", "Verified", "#1565C0"),
    )
    if status_col:
        df_status = df.copy()
        # تطبيع Status: حروف صغيرة + إزالة مسافات زائدة لتحمل اختلافات الكتابة
        s = df_status[status_col].fillna("").astype(str).str.strip().str.lower()
        df_status["_status_norm"] = s.str.replace(r"\s+", " ", regex=True)
        if date_col:
            df_status["_date"] = _excel_dates_to_datetime(df_status[date_col], errors="coerce")
            df_status = df_status.dropna(subset=["_date"])
            df_status["_day"] = df_status["_date"].dt.normalize()
            # كل يوم: عدد الصفوف (شحنات) لكل حالة، ثم جمع كل الأيام
            count_in_transit = 0
            count_receiving_complete = 0
            count_verified = 0
            for _day, grp in df_status.groupby("_day"):
                count_in_transit += (grp["_status_norm"] == "in transit").sum()
                count_receiving_complete += (grp["_status_norm"] == "receiving complete").sum()
                count_verified += (grp["_status_norm"] == "verified").sum()
        else:
            count_in_transit = (df_status["_status_norm"] == "in transit").sum()
            count_receiving_complete = (df_status["_status_norm"] == "receiving complete").sum()
            count_verified = (df_status["_status_norm"] == "verified").sum()
        total_pending = count_in_transit + count_receiving_complete + count_verified
        if total_pending > 0:
            for key, label, color in STATUS_LABELS:
                if key == "in transit":
                    c = count_in_transit
                elif key == "receiving complete":
                    c = count_receiving_complete
                else:
                    c = count_verified
                pct = round((c / total_pending) * 100)
                pending.append({
                    "label": label,
                    "value": f"{pct}%",
                    "pct": pct,
                    "color": color,
                })
    # لو مفيش داتا Pending من الشيت نرجع قائمة فاضية (الداتا من الشيت فقط)
    if not pending:
        pending = []

    return {"inbound_kpi": inbound_kpi, "pending_shipments": pending}


def _read_outbound_data_from_excel(excel_path, request=None):
    """
    يقرأ بيانات Outbound من شيت Outbound_Data في ملف الداشبورد.
    - عمود Status: نفلتر "Released" → released_orders، "Picked" → picked_orders
    - عمود Order_ID: نحذف المتكرر (unique) ونحسب عدد الطلبات لكل حالة
    """
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
    except Exception:
        return None
    sheet_name = None
    for name in xls.sheet_names:
        if "outbound_data" in name.lower().replace(" ", "").replace("_", ""):
            sheet_name = name
            break
    if not sheet_name:
        for name in xls.sheet_names:
            if "outbound" in name.lower():
                sheet_name = name
                break
    if not sheet_name:
        return None
    try:
        _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
        )
    except Exception:
        return None
    if df.empty or len(df) < 1:
        return None

    df.columns = [str(c).strip() for c in df.columns]
    cols_lower = {c.lower(): c for c in df.columns if c}

    def _col(*keys):
        for k in keys:
            k_norm = k.lower().replace(" ", "").replace("_", "")
            for col in cols_lower:
                if col.replace(" ", "").replace("_", "") == k_norm:
                    return cols_lower[col]
            if k.lower() in cols_lower:
                return cols_lower[k.lower()]
        return None

    status_col = _col("Status", "status")
    order_col = _col("Order_ID", "Order ID", "Order_ID", "OrderID")
    if not status_col or not order_col:
        return None

    # تطبيع Status للمقارنة
    s = df[status_col].fillna("").astype(str).str.strip().str.lower()
    df["_status_norm"] = s.str.replace(r"\s+", " ", regex=True)

    released_mask = df["_status_norm"] == "released"
    picked_mask = df["_status_norm"] == "picked"

    released_orders = 0
    picked_orders = 0
    if released_mask.any():
        released_orders = df.loc[released_mask, order_col].dropna().astype(str).str.strip().nunique()
    if picked_mask.any():
        picked_orders = df.loc[picked_mask, order_col].dropna().astype(str).str.strip().nunique()

    # Number of Pallets (LPNs) من الشيت — عمود Pallets_number أو أي اسم معروف، نجمع القيم
    lpn_col = _col("Pallets_number", "Pallets number", "LPNs", "LPN", "Nbr_LPNs", "Nbr LPNs", "Pallets", "Number of Pallets")
    number_of_pallets = 0
    keys_from_sheet = ["released_orders", "picked_orders"]
    if lpn_col:
        number_of_pallets = int(pd.to_numeric(df[lpn_col], errors="coerce").fillna(0).sum())
        keys_from_sheet.append("number_of_pallets")

    return {
        "outbound_kpi": {
            "released_orders": int(released_orders),
            "picked_orders": int(picked_orders),
            "number_of_pallets": number_of_pallets,
        },
        "outbound_kpi_keys_from_sheet": keys_from_sheet,
    }


def _read_pods_data_from_excel(excel_path, request=None):
    """
    يقرأ من شيت PODs_Data: عمود POD_Status (On Time, Pending, Late)،
    Delivery_Date للشهور، POD_ID للعدد. يرجّع داتا لشارت خط: كل شهر ونسبة كل حالة %.
    """
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
    except Exception:
        return None
    sheet_name = None
    for name in xls.sheet_names:
        n = str(name).strip().lower().replace(" ", "").replace("_", "")
        if "podsdata" in n or "pods_data" in n or (n == "pods" and "data" in n):
            sheet_name = name
            break
    if not sheet_name:
        for name in xls.sheet_names:
            if "pod" in str(name).lower():
                sheet_name = name
                break
    if not sheet_name:
        return None
    try:
        _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
        )
    except Exception:
        return None
    if df.empty or len(df) < 1:
        return None

    df.columns = [str(c).strip() for c in df.columns]
    cols_lower = {c.lower(): c for c in df.columns if c}

    def _col(*keys):
        for k in keys:
            k_norm = k.lower().replace(" ", "").replace("_", "")
            for col in cols_lower:
                if col.replace(" ", "").replace("_", "") == k_norm:
                    return cols_lower[col]
            if k.lower() in cols_lower:
                return cols_lower[k.lower()]
        return None

    status_col = _col("POD_Status", "POD Status", "PODStatus")
    date_col = _col("Delivery_Date", "Delivery Date", "DeliveryDate", "Date")
    pod_id_col = _col("POD_ID", "POD ID", "PODID")
    if not status_col or not date_col:
        return None
    if not pod_id_col:
        pod_id_col = df.columns[0]

    s = df[status_col].fillna("").astype(str).str.strip().str.lower()
    df["_status_norm"] = s.str.replace(r"\s+", " ", regex=True)
    valid_statuses = {"on time", "pending", "late"}
    df = df[df["_status_norm"].isin(valid_statuses)].copy()
    if df.empty:
        return None

    df["_date"] = _excel_dates_to_datetime(df[date_col], errors="coerce")
    df = df.dropna(subset=["_date"])
    df["_month"] = df["_date"].dt.strftime("%b")
    month_order = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    months_in_data = df["_month"].unique().tolist()
    months_sorted = sorted(months_in_data, key=lambda m: month_order.index(m) if m in month_order else 99)

    series_on_time = []
    series_pending = []
    series_late = []
    for m in months_sorted:
        grp = df[df["_month"] == m]
        on_time = (grp["_status_norm"] == "on time").sum()
        pending = (grp["_status_norm"] == "pending").sum()
        late = (grp["_status_norm"] == "late").sum()
        total = on_time + pending + late
        if total == 0:
            series_on_time.append(0)
            series_pending.append(0)
            series_late.append(0)
        else:
            series_on_time.append(round(100.0 * on_time / total, 1))
            series_pending.append(round(100.0 * pending / total, 1))
            series_late.append(round(100.0 * late / total, 1))

    # تجميع النسب الإجمالية للـ pod_status_breakdown (من نفس الشيت)
    total_on = (df["_status_norm"] == "on time").sum()
    total_pend = (df["_status_norm"] == "pending").sum()
    total_late = (df["_status_norm"] == "late").sum()
    total_all = total_on + total_pend + total_late
    if total_all > 0:
        pct_on = int(round(100.0 * total_on / total_all))
        pct_pend = int(round(100.0 * total_pend / total_all))
        pct_late = int(round(100.0 * total_late / total_all))
    else:
        pct_on = pct_pend = pct_late = 0
    pod_status_breakdown = [
        {"label": "On Time", "pct": pct_on, "color": "#7FB7A6"},
        {"label": "Pending", "pct": pct_pend, "color": "#A8C8EB"},
        {"label": "Late", "pct": pct_late, "color": "#E8A8A2"},
    ]

    return {
        "categories": months_sorted,
        "series": [
            {"name": "On Time", "data": series_on_time},
            {"name": "Pending", "data": series_pending},
            {"name": "Late", "data": series_late},
        ],
        "pod_status_breakdown": pod_status_breakdown,
    }


def _read_returns_data_from_excel(excel_path, request=None):
    """
    يقرأ من شيت Returns_Data: عمود Return_Status (فلترة مثل PODs: On Time, Pending, Late)،
    Request_Date للشهور، Return_ID لعدد الشحنات (unique). يرجّع returns_kpi و returns_chart_data.
    """
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
    except Exception:
        return None
    sheet_name = None
    for name in xls.sheet_names:
        n = str(name).strip().lower().replace(" ", "").replace("_", "")
        if "returnsdata" in n or "returns_data" in n:
            sheet_name = name
            break
    if not sheet_name:
        for name in xls.sheet_names:
            if "returns" in str(name).lower() and "data" in str(name).lower():
                sheet_name = name
                break
    if not sheet_name:
        for name in xls.sheet_names:
            if "return" in str(name).lower():
                sheet_name = name
                break
    if not sheet_name:
        return None
    try:
        _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
        )
    except Exception:
        return None
    if df.empty or len(df) < 1:
        return None

    df.columns = [str(c).strip() for c in df.columns]
    cols_lower = {c.lower(): c for c in df.columns if c}

    def _col(*keys):
        for k in keys:
            k_norm = k.lower().replace(" ", "").replace("_", "")
            for col in cols_lower:
                if col.replace(" ", "").replace("_", "") == k_norm:
                    return cols_lower[col]
            if k.lower() in cols_lower:
                return cols_lower[k.lower()]
        return None

    status_col = _col("Return_Status", "Return Status", "ReturnStatus")
    date_col = _col("Request_Date", "Request Date", "RequestDate", "Date")
    return_id_col = _col("Return_ID", "Return ID", "ReturnID")
    nbr_skus_col = _col("Nbr_SKUs", "Nbr SKUs", "NbrSKUs")
    nbr_items_col = _col("Nbr_Items", "Nbr Items", "NbrItems")
    if not status_col or not date_col:
        return None
    if not return_id_col:
        return_id_col = df.columns[0]

    s = df[status_col].fillna("").astype(str).str.strip().str.lower()
    df["_status_norm"] = s.str.replace(r"\s+", " ", regex=True)
    valid_statuses = {"on time", "pending", "late"}
    df = df[df["_status_norm"].isin(valid_statuses)].copy()
    if df.empty:
        return None

    df["_date"] = _excel_dates_to_datetime(df[date_col], errors="coerce")
    df = df.dropna(subset=["_date"])
    df["_month"] = df["_date"].dt.strftime("%b")
    month_order = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    months_in_data = df["_month"].unique().tolist()
    months_sorted = sorted(months_in_data, key=lambda m: month_order.index(m) if m in month_order else 99)

    total_unique_returns = df[return_id_col].dropna().astype(str).str.strip().nunique()
    total_rows = len(df)

    total_skus_kpi = total_unique_returns
    total_lpns_kpi = total_rows
    if nbr_skus_col:
        total_skus_kpi = int(pd.to_numeric(df[nbr_skus_col], errors="coerce").fillna(0).sum())
    if nbr_items_col:
        total_lpns_kpi = int(pd.to_numeric(df[nbr_items_col], errors="coerce").fillna(0).sum())

    series_on_time = []
    series_pending = []
    series_late = []
    for m in months_sorted:
        grp = df[df["_month"] == m]
        on_time = (grp["_status_norm"] == "on time").sum()
        pending = (grp["_status_norm"] == "pending").sum()
        late = (grp["_status_norm"] == "late").sum()
        total = on_time + pending + late
        if total == 0:
            series_on_time.append(0)
            series_pending.append(0)
            series_late.append(0)
        else:
            series_on_time.append(round(100.0 * on_time / total, 1))
            series_pending.append(round(100.0 * pending / total, 1))
            series_late.append(round(100.0 * late / total, 1))

    return {
        "returns_kpi": {
            "total_skus": total_skus_kpi,
            "total_lpns": total_lpns_kpi,
        },
        "returns_chart_data": {
            "categories": months_sorted,
            "series": [
                {"name": "On Time", "data": series_on_time},
                {"name": "Pending", "data": series_pending},
                {"name": "Late", "data": series_late},
            ],
        },
    }


def _read_inventory_data_from_excel(excel_path, request=None):
    """
    يقرأ من شيت Inventory_Lots:
    - عمود LPNs: تجميع كل القيم (مجموع) = Total LPNs.
    - عمود Snapshot_Date: كل يوم بيومه (للاستخدام لاحقاً في شارت/تحليل).
    - عمود SKU: حذف المتكرر وعدّ القيم الفريدة فقط = Total SKUs.
    """
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
    except Exception:
        return None
    sheet_name = None
    for name in xls.sheet_names:
        n = str(name).strip().lower().replace(" ", "").replace("_", "")
        if "inventorylots" in n or "inventory_lots" in n:
            sheet_name = name
            break
    if not sheet_name:
        for name in xls.sheet_names:
            if "inventory" in str(name).lower() and "lot" in str(name).lower():
                sheet_name = name
                break
    if not sheet_name:
        for name in xls.sheet_names:
            if "inventory" in str(name).lower():
                sheet_name = name
                break
    if not sheet_name:
        return None
    try:
        _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
        )
    except Exception:
        return None
    if df.empty or len(df) < 1:
        return None

    df.columns = [str(c).strip() for c in df.columns]
    cols_lower = {c.lower(): c for c in df.columns if c}

    def _col(*keys):
        for k in keys:
            k_norm = k.lower().replace(" ", "").replace("_", "")
            for col in cols_lower:
                if col.replace(" ", "").replace("_", "") == k_norm:
                    return cols_lower[col]
            if k.lower() in cols_lower:
                return cols_lower[k.lower()]
        return None

    lpns_col = _col("LPNs", "LPN")
    snapshot_date_col = _col("Snapshot_Date", "Snapshot Date", "SnapshotDate", "Date")
    sku_col = _col("SKU", "Sku")

    total_lpns = 0
    total_skus = 0

    if lpns_col:
        total_lpns = int(pd.to_numeric(df[lpns_col], errors="coerce").fillna(0).sum())

    if sku_col:
        sku_series = df[sku_col].dropna().astype(str).str.strip()
        sku_series = sku_series[sku_series != ""]
        total_skus = int(sku_series.nunique())

    return {
        "inventory_kpi": {
            "total_skus": total_skus,
            "total_lpns": total_lpns,
            "utilization_pct": "",
        },
    }


def _read_inventory_snapshot_capacity_from_excel(excel_path, request=None):
    """
    يقرأ من شيت Inventory_Snapshot:
    - Used_Space_m3 → Used (مجموع ثم نسبة مئوية).
    - Available_Space_m3 → Available (مجموع ثم نسبة مئوية).
    يرجع inventory_capacity_data: { used: نسبة Used %, available: نسبة Available % }.
    """
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
    except Exception:
        return None
    sheet_name = None
    for name in xls.sheet_names:
        n = str(name).strip().lower().replace(" ", "").replace("_", "")
        if "inventorysnapshot" in n or "inventory_snapshot" in n:
            sheet_name = name
            break
    if not sheet_name:
        for name in xls.sheet_names:
            if "inventory" in str(name).lower() and "snapshot" in str(name).lower():
                sheet_name = name
                break
    if not sheet_name:
        return None
    try:
        _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
        )
    except Exception:
        return None
    if df.empty or len(df) < 1:
        return None

    df.columns = [str(c).strip() for c in df.columns]
    cols_lower = {c.lower(): c for c in df.columns if c}

    def _col(*keys):
        for k in keys:
            k_norm = k.lower().replace(" ", "").replace("_", "")
            for col in cols_lower:
                if col.replace(" ", "").replace("_", "") == k_norm:
                    return cols_lower[col]
            if k.lower() in cols_lower:
                return cols_lower[k.lower()]
        return None

    used_col = _col("Used_Space_m3", "Used Space m3", "UsedSpace_m3")
    avail_col = _col("Available_Space_m3", "Available Space m3", "AvailableSpace_m3")
    if not used_col or not avail_col:
        return None

    total_used = pd.to_numeric(df[used_col], errors="coerce").fillna(0).sum()
    total_avail = pd.to_numeric(df[avail_col], errors="coerce").fillna(0).sum()
    total = total_used + total_avail
    if total <= 0:
        return {"inventory_capacity_data": {"used": 0, "available": 0}}

    used_pct = round(100.0 * total_used / total, 1)
    available_pct = round(100.0 - used_pct, 1)
    return {
        "inventory_capacity_data": {
            "used": used_pct,
            "available": available_pct,
        },
    }


def _read_inventory_warehouse_table_from_excel(excel_path, request=None):
    """
    يقرأ من شيت Inventory_Snapshot جدول الـ Warehouse:
    - Warehouse من عمود Warehouse
    - SKUs من عمود Total_SKUs
    - Available Space من عمود Available_Space_m3
    - Utilization % من عمود Utilization_%
    كل صف كما هو من الشيت بدون جمع.
    """
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
    except Exception:
        return None
    sheet_name = None
    for name in xls.sheet_names:
        n = str(name).strip().lower().replace(" ", "").replace("_", "")
        if "inventorysnapshot" in n or "inventory_snapshot" in n:
            sheet_name = name
            break
    if not sheet_name:
        for name in xls.sheet_names:
            if "inventory" in str(name).lower() and "snapshot" in str(name).lower():
                sheet_name = name
                break
    if not sheet_name:
        return None
    try:
        _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
        df = pd.read_excel(
            excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
        )
    except Exception:
        return None
    if df.empty or len(df) < 1:
        return None

    df.columns = [str(c).strip() for c in df.columns]
    cols_lower = {c.lower(): c for c in df.columns if c}

    def _col(*keys):
        for k in keys:
            k_norm = k.lower().replace(" ", "").replace("_", "")
            for col in cols_lower:
                if col.replace(" ", "").replace("_", "") == k_norm:
                    return cols_lower[col]
            if k.lower() in cols_lower:
                return cols_lower[k.lower()]
        return None

    warehouse_col = _col("Warehouse")
    total_skus_col = _col("Total_SKUs", "Total SKUs", "TotalSKUs")
    avail_space_col = _col("Available_Space_m3", "Available Space m3", "AvailableSpace_m3")
    util_col = _col("Utilization_%", "Utilization %", "UtilizationPct", "Utilization")
    if not warehouse_col:
        return None

    def _val(col, r):
        if not col or col not in r.index:
            return ""
        v = r[col]
        if pd.isna(v):
            return ""
        if isinstance(v, (int, float)):
            return str(int(v)) if v == int(v) else str(v)
        return str(v).strip()

    def _util_pct(col, r):
        if not col or col not in r.index:
            return ""
        v = r[col]
        if pd.isna(v):
            return ""
        try:
            num = float(v)
            if 0 <= num <= 1:
                return f"{round(num * 100, 2)}%"
            return f"{round(num, 2)}%"
        except (TypeError, ValueError):
            s = str(v).strip()
            return f"{s}%" if s and not s.endswith("%") else s

    rows = []
    for _, r in df.iterrows():
        warehouse = "" if pd.isna(r[warehouse_col]) else str(r[warehouse_col]).strip()
        rows.append({
            "warehouse": warehouse,
            "skus": _val(total_skus_col, r),
            "available_space": _val(avail_space_col, r),
            "utilization_pct": _util_pct(util_col, r),
        })
    if not rows:
        return None
    return {"inventory_warehouse_table": rows}


def _read_returns_region_table_from_excel(excel_path, request=None):
    """
    يبني returns_region_table من Inventory_Lots + Inventory_Snapshot:
    - Region من عمود Warehouse في Inventory_Lots (فلترة بالـ Warehouse).
    - SKUs: عدد القيم الفريدة لعمود SKU لكل Warehouse بعد الفلترة بالتاريخ.
    - Available: مجموع LPNs لكل Warehouse بعد الفلترة بـ Snapshot_Date (آخر تاريخ).
    - Utilization %: (LPNs للمنطقة والتاريخ) / Capacity_m3 من Inventory_Snapshot لنفس المنطقة، كنسبة مئوية.
    """
    try:
        xls = pd.ExcelFile(excel_path, engine="openpyxl")
    except Exception:
        return None

    def _find_sheet(*names):
        for want in names:
            want_n = want.lower().replace(" ", "").replace("_", "")
            for s in xls.sheet_names:
                if want_n in str(s).lower().replace(" ", "").replace("_", ""):
                    return s
        return None

    lots_sheet = _find_sheet("Inventory_Lots", "Inventory Lots")
    snapshot_sheet = _find_sheet("Inventory_Snapshot", "Inventory Snapshot")
    if not lots_sheet:
        return None

    _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
    try:
        df_lots = pd.read_excel(
            excel_path, sheet_name=lots_sheet, engine="openpyxl", header=0, **_nr_kw
        )
    except Exception:
        return None
    if df_lots.empty:
        return None

    df_lots.columns = [str(c).strip() for c in df_lots.columns]
    cols_lots = {c.lower(): c for c in df_lots.columns if c}

    def _col_lots(*keys):
        for k in keys:
            k_norm = k.lower().replace(" ", "").replace("_", "")
            for col in cols_lots:
                if col.replace(" ", "").replace("_", "") == k_norm:
                    return cols_lots[col]
        return None

    wh_col = _col_lots("Warehouse")
    sku_col = _col_lots("SKU", "Sku")
    lpns_col = _col_lots("LPNs", "LPN")
    snap_col = _col_lots("Snapshot_Date", "Snapshot Date", "SnapshotDate", "Date")
    if not wh_col or not snap_col:
        return None
    if not lpns_col:
        lpns_col = df_lots.columns[1] if len(df_lots.columns) > 1 else None
    if not lpns_col:
        return None

    df_lots["_date"] = _excel_dates_to_datetime(df_lots[snap_col], errors="coerce")
    df_lots = df_lots.dropna(subset=["_date"])
    if df_lots.empty:
        return None

    latest_date = df_lots["_date"].max()
    df_filtered = df_lots[df_lots["_date"] == latest_date].copy()

    capacity_by_warehouse = {}
    if snapshot_sheet:
        try:
            df_snap = pd.read_excel(
                excel_path, sheet_name=snapshot_sheet, engine="openpyxl", header=0, **_nr_kw
            )
            if not df_snap.empty:
                df_snap.columns = [str(c).strip() for c in df_snap.columns]
                snap_cols = {c.lower(): c for c in df_snap.columns if c}
                snap_wh = next((snap_cols[c] for c in snap_cols if "warehouse" in c.replace(" ", "").replace("_", "")), None)
                cap_col = next((snap_cols[c] for c in snap_cols if "capacity_m3" in c.replace(" ", "").replace("_", "") or ("capacity" in c and "m3" in c)), None)
                if not cap_col:
                    used_c = next((snap_cols[c] for c in snap_cols if "used_space" in c.replace(" ", "").replace("_", "")), None)
                    avail_c = next((snap_cols[c] for c in snap_cols if "available_space" in c.replace(" ", "").replace("_", "")), None)
                    if used_c and avail_c:
                        df_snap["_cap"] = pd.to_numeric(df_snap[used_c], errors="coerce").fillna(0) + pd.to_numeric(df_snap[avail_c], errors="coerce").fillna(0)
                        cap_col = "_cap"
                if snap_wh and cap_col:
                    for _, r in df_snap.iterrows():
                        w = r.get(snap_wh)
                        if pd.isna(w):
                            continue
                        w = str(w).strip()
                        if not w:
                            continue
                        c = r.get(cap_col)
                        if cap_col == "_cap":
                            val = c
                        else:
                            val = pd.to_numeric(c, errors="coerce")
                        if pd.notna(val) and val > 0:
                            capacity_by_warehouse[w] = float(val)
        except Exception:
            pass

    df_filtered["_lpns_num"] = pd.to_numeric(df_filtered[lpns_col], errors="coerce").fillna(0)
    rows = []
    for wh, grp in df_filtered.groupby(wh_col, dropna=False):
        wh_name = "" if pd.isna(wh) else str(wh).strip()
        skus = grp[sku_col].dropna().astype(str).str.strip() if sku_col else pd.Series(dtype=object)
        skus = skus[skus != ""].nunique() if not skus.empty else 0
        available = int(grp["_lpns_num"].sum())
        cap = capacity_by_warehouse.get(wh_name) or capacity_by_warehouse.get(wh)
        if cap and cap > 0:
            util = round(100.0 * available / cap, 2)
            utilization_pct = f"{util}%"
        else:
            utilization_pct = "—"
        rows.append({
            "region": wh_name,
            "skus": str(int(skus)) if isinstance(skus, (int, float)) else str(skus),
            "available": str(available),
            "utilization_pct": utilization_pct,
        })

    if not rows:
        return None
    return {"returns_region_table": rows}


def get_dashboard_tab_context(request):
    """
    يبني سياق تاب الداشبورد (نفس بيانات Dashboard view).
    إذا وُجدت الفيو في تطبيق dashboard أو inbound يتم استخدامها، وإلا يُرجع سياق افتراضي.
    """
    try:
        for app_label in ["dashboard", "inbound"]:
            try:
                view_module = __import__(f"{app_label}.views", fromlist=["DashboardView"])
                ViewClass = getattr(view_module, "DashboardView", None)
                if ViewClass is not None:
                    view = ViewClass()
                    view.request = request
                    view.object = None
                    return view.get_context_data()
            except (ImportError, AttributeError):
                continue
    except Exception:
        pass
    # سياق افتراضي عند عدم وجود الموديلات/الفيو (مع داتا وهمية لـ Inbound)
    return {
        "title": "Dashboard",
        "breadcrumb": {"title": "Healthcare Dashboard", "parent": "Dashboard", "child": "Default"},
        "is_admin": False,
        "is_employee": False,
        "inbound_data": [],
        "transportation_outbound_data": [],
        "wh_outbound_data": [],
        "returns_data": [],
        "expiry_data": [],
        "damage_data": [],
        "inventory_data": [],
        "pallet_location_availability_data": [],
        "hse_data": [],
        "number_of_shipments": 0,
        "total_vehicles_daily": 0,
        "total_pallets": 0,
        "total_pending_shipments": 0,
        "total_number_of_shipments": 0,
        "total_quantity": 0,
        "total_number_of_line": 0,
        # Inbound KPI + داتا شارت Pending Shipments (من الديكت في الفيو)
        "inbound_kpi": INBOUND_DEFAULT_KPI.copy(),
        "pending_shipments": list(INBOUND_DEFAULT_PENDING_SHIPMENTS),
        "shipment_data": {"bulk": 0, "loose": 0, "cold": 0, "frozen": 0, "ambient": 0},
        "wh_total_released_order": 0,
        "wh_total_piked_order": 0,
        "wh_total_pending_pick_orders": 0,
        "wh_total_number_of_PODs_collected_on_time": 0,
        "wh_total_number_of_PODs_collected_Late": 0,
        "total_orders_items_returned": 0,
        "total_number_of_return_items_orders_updated_on_time": 0,
        "total_number_of_return_items_orders_updated_late": 0,
        "total_SKUs_expired": 0,
        "total_expired_SKUS_disposed": 0,
        "total_nearly_expired_1_to_3_months": 0,
        "total_nearly_expired_3_to_6_months": 0,
        "total_SKUs_expired_calculated": 0,
        "Total_QTYs_Damaged_by_WH": 0,
        "Total_Number_of_Damaged_during_receiving": 0,
        "Total_Araive_Damaged": 0,
        "Total_Locations_match": 0,
        "Total_Locations_not_match": 0,
        "last_shipment": None,
        "Total_Storage_Pallet": 0,
        "Total_Storage_pallet_empty": 0,
        "Total_Storage_Bin": 0,
        "Total_occupied_pallet_location": 0,
        "Total_Storage_Bin_empty": 0,
        "Total_occupied_Bin_location": 0,
        "Total_Incidents_on_the_side": 0,
        "total_no_of_employees": 0,
        "admin_data": [],
        "user_type": "Unknown",
        "years": [],
        "months": list(calendar_module.month_name)[1:],
        "days": list(range(1, 32)),
        "returns_region_table": [
            {"region": "Main warehouse", "skus": "2,538", "available": "1118", "utilization_pct": "71%"},
            {"region": "Dammam DC", "skus": "501", "available": "200", "utilization_pct": "—"},
            {"region": "Riyadh DC", "skus": "3,996", "available": "209", "utilization_pct": "—"},
            {"region": "Jeddah DC", "skus": "7,996", "available": "300", "utilization_pct": "—"},
        ],
    }


# إزالة الحروف العربية (وما شابهها) من وصف الصنف في نتائج Traceability — يزيل أيضًا نص UTF-8 المعروض كـ mojibake
_ARABIC_SCRIPT_RE = re.compile(
    r"[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF\uFB50-\uFDFF\uFE70-\uFEFF"
    r"\u061C\u200F\u200E]+"
)


def _keep_traceability_desc_char(c):
    """Keep printable ASCII, whitespace, real letters/numbers; drop modifier letters (e.g. ˆ U+02C6, Lm) and junk symbols."""
    if c in " \t\n\r":
        return True
    o = ord(c)
    if 32 <= o <= 126:
        return False if c == "^" else True
    cat = unicodedata.category(c)
    if cat == "Lm":
        return False
    if cat.startswith("L") or cat == "Nd":
        return True
    return False


def _clean_traceability_item_description(val):
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    if not s:
        return ""
    s = _ARABIC_SCRIPT_RE.sub("", s)
    # بقايا UTF-8 مقروء كـ Latin-1 (حروف في نطاق 0x80–0xFF كثيرة)
    compact = re.sub(r"\s", "", s)
    if compact:
        hi_latin = sum(1 for c in compact if "\x80" <= c <= "\xff")
        if hi_latin >= 3 and hi_latin * 2 >= len(compact):
            s = "".join(c for c in s if ord(c) < 128)
    s = "".join(c for c in s if _keep_traceability_desc_char(c))
    s = re.sub(r"[\^\u02c6\u0302]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


@method_decorator(csrf_exempt, name="dispatch")
class UploadExcelViewRoche(View):
    template_name = "index.html"
    excel_file_name = "all sheet.xlsm"
    correct_code = "1234"

    # تابات تحذف من الداشبورد (أضف أسماء الشيتات كما هي في الإكسل)
    EXCLUDE_TABS = []  # مثال: ["Sheet2", "تقارير قديمة", "Backup"]
    # أو: اعرض تابات معينة فقط (لو ضعت قائمة هنا، التابات الأخرى كلها تختفي)
    # لو عايزة تابات محددة فقط: ["Overview", "Dock to stock", ...] — وإلا اتركي None
    INCLUDE_ONLY_TABS = None
    # تابات افتراضية نعرضها بدون الاعتماد على شيت مباشر
    DASHBOARD_TAB_NAME = "Dashboard"
    DEFAULT_EXCEL_FILENAMES = [
        # ملفات Nespresso (جديدة)
        "all_sheet_nespresso.xlsx",
        "all_sheet_nespresso.xlsm",
        "all sheet nespresso.xlsx",
        "all sheet nespresso.xlsm",
        # الأسماء القديمة (للتوافق)
        "all sheet.xlsm",
        "all sheet.xlsx",
        "all_sheet.xlsm",
        "all_sheet.xlsx",
    ]

    MONTH_LOOKUP = {}
    MONTH_PREFIXES = set()
    for idx in range(1, 13):
        abbr = month_abbr[idx]
        full = month_name[idx]
        if abbr:
            MONTH_LOOKUP[abbr.lower()] = abbr
            MONTH_PREFIXES.add(abbr.lower())
        if full:
            MONTH_LOOKUP[full.lower()] = abbr
        MONTH_LOOKUP[str(idx)] = abbr
        MONTH_LOOKUP[f"{idx:02d}"] = abbr
    MONTH_LOOKUP["sept"] = "Sep"

    AGGREGATE_COLUMN_KEYWORDS = {
        "total",
        "grand total",
        "overall total",
        "sum",
        "ytd",
        "y.t.d.",
        "avg",
        "average",
        "target",
        "target (%)",
        "target %",
        "target%",
        "cumulative",
    }

    # اسم الملف الافتراضي إذا وُضع في excel_uploads بدون رفع (مثلاً all sheet.xlsm)
    def get_excel_path(self):
        folder_path = os.path.abspath(
            os.path.join(str(settings.MEDIA_ROOT), "excel_uploads")
        )
        os.makedirs(folder_path, exist_ok=True)
        # الملف الرئيسي لكل التابات (ما عدا Dashboard): all_sheet_nespresso.xlsx
        priority_files = [
            "all_sheet_nespresso.xlsx",
            "all_sheet_nespresso.xlsm",
            "all sheet nespresso.xlsx",
            "all sheet nespresso.xlsm",
            "all_sheet.xlsx",
            "all_sheet.xlsm",
            "all sheet.xlsx",
            "all sheet.xlsm",
            "latest.xlsm",
            "latest.xlsx",
        ] + self.DEFAULT_EXCEL_FILENAMES
        for name in priority_files:
            path = os.path.join(folder_path, name)
            if os.path.exists(path):
                return path
        return os.path.join(folder_path, "latest.xlsx")

    def get_uploaded_file_path(self, request):
        folder = os.path.abspath(
            os.path.join(str(settings.MEDIA_ROOT), "excel_uploads")
        )
        os.makedirs(folder, exist_ok=True)

        # أولوية: ملف الجلسة ثم all_sheet_nespresso ثم latest ثم all sheet
        if request:
            saved_path = request.session.get("uploaded_excel_path")
            if saved_path:
                if os.path.exists(saved_path):
                    return os.path.abspath(saved_path)
                try:
                    request.session.pop("uploaded_excel_path", None)
                    request.session.save()
                except Exception:
                    pass
        priority_files = [
            "all_sheet_nespresso.xlsx",
            "all_sheet_nespresso.xlsm",
            "all sheet nespresso.xlsx",
            "all sheet nespresso.xlsm",
            "latest.xlsm",
            "latest.xlsx",
        ] + self.DEFAULT_EXCEL_FILENAMES
        for name in priority_files:
            path = os.path.join(folder, name)
            if os.path.exists(path):
                if request:
                    try:
                        request.session["uploaded_excel_path"] = os.path.abspath(path)
                        request.session.save()
                    except Exception:
                        pass
                return os.path.abspath(path)
        # لا تُرجع مساراً وهمياً — كان يمنع سلسلة `or get_excel_path()` ويُضلل فحص exists
        return None

    def resolve_excel_file_path(self, request):
        """
        مسار ملف الإكسل الرئيسي فقط إذا الملف موجود فعلاً على القرص.
        يجمع الجلسة + أولوية أسماء الملفات في المجلد (مثل get_uploaded ثم get_excel).
        """
        p = self.get_uploaded_file_path(request)
        if p and os.path.isfile(p):
            return os.path.abspath(p)
        p2 = self.get_excel_path()
        if p2 and os.path.isfile(p2):
            return os.path.abspath(p2)
        return None

    def _available_months_for_tab(self, request, tab_key):
        """
        شهور تظهر في قائمة الفلتر حسب التاب/الشيت المعني فقط (وليس كل ملف الإكسل).
        tab_key يُفضّل أن يكون نص الطلب الأصلي للشيتات الخام؛ للتابات الافتراضية يُستخدم الاسم بعد lower().
        """
        excel_path = self.resolve_excel_file_path(request)
        if not excel_path:
            return []
        t = (tab_key or "").strip().lower()
        # الاسم كما في الرابط (يحافظ على حالة الأحرف لأسماء الشيتات الخام)
        tab_raw = (request.GET.get("tab") or "").strip()

        def _union(*lists):
            acc = set()
            for L in lists:
                for x in L or []:
                    if x:
                        acc.add(x)
            return _sorted_unique_month_abbrs(acc)

        if t == "all":
            m_in = []
            ins = _inbound_sheet_name_for_months(excel_path)
            if ins:
                m_in = _distinct_month_abbrs_from_sheet(
                    excel_path,
                    ins,
                    preferred_columns=[
                        "Create Timestamp",
                        "Creation Date",
                        "Last LPN Rcv TS",
                        "Create shipment D&T",
                        "Month",
                    ],
                )
            m_b2b = _distinct_month_abbrs_from_sheet(
                excel_path,
                "B2B_Outbound",
                preferred_columns=["Actual Delivery Date", "Creation Date & Time"],
            )
            m_b2c = _distinct_month_abbrs_from_sheet(
                excel_path,
                "B2C_Outbound",
                preferred_columns=[
                    "Delivered Date",
                    "CREATION DATE",
                    "Creation Date",
                    "Picked Date",
                ],
            )
            return _union(m_in, m_b2b, m_b2c)

        if "b2c outbound" in t:
            return _distinct_month_abbrs_from_sheet(
                excel_path,
                "B2C_Outbound",
                preferred_columns=[
                    "Delivered Date",
                    "DELIVERED DATE",
                    "CREATION DATE",
                    "Creation Date",
                    "Picked Date",
                    "Dispatch",
                ],
            )

        if "b2b outbound" in t:
            return _distinct_month_abbrs_from_sheet(
                excel_path,
                "B2B_Outbound",
                preferred_columns=["Actual Delivery Date", "Creation Date & Time"],
            )

        if t == "inbound" or ("dock" in t and "stock" in t):
            ins = _inbound_sheet_name_for_months(excel_path)
            if ins:
                return _distinct_month_abbrs_from_sheet(
                    excel_path,
                    ins,
                    preferred_columns=[
                        "Create Timestamp",
                        "Creation Date",
                        "Last LPN Rcv TS",
                        "Create shipment D&T",
                        "Month",
                    ],
                )
            return []

        if "return" in t or "refusal" in t or "rejections" in t:
            ins = _inbound_sheet_name_for_months(excel_path)
            if ins:
                return _distinct_month_abbrs_from_sheet(
                    excel_path,
                    ins,
                    preferred_columns=[
                        "Create Timestamp",
                        "Creation Date",
                        "Month",
                    ],
                )
            return []

        if ("capacity" in t and "expiry" in t) or t == "expiry":
            cap_sn = _capacity_expiry_sheet_for_months(excel_path)
            if cap_sn:
                return _distinct_month_abbrs_from_sheet(
                    excel_path,
                    cap_sn,
                    preferred_columns=[
                        "Expiry Date",
                        "Batch Date",
                        "Creation Date",
                        "Month",
                    ],
                )
            return []

        if "safety" in t and "kpi" in t:
            try:
                xls = pd.ExcelFile(excel_path, engine="openpyxl")
                sn = next(
                    (
                        s
                        for s in xls.sheet_names
                        if "safety" in (s or "").lower() and "kpi" in (s or "").lower()
                    ),
                    None,
                )
                if sn:
                    return _distinct_month_abbrs_from_sheet(
                        excel_path, sn, preferred_columns=None
                    )
            except Exception:
                pass
            return []

        if "traceability" in t and "kpi" in t:
            ins = _inbound_sheet_name_for_months(excel_path)
            mi = (
                _distinct_month_abbrs_from_sheet(
                    excel_path,
                    ins,
                    preferred_columns=["Create Timestamp", "Last LPN Rcv TS"],
                )
                if ins
                else []
            )
            m2 = _distinct_month_abbrs_from_sheet(
                excel_path,
                "B2C_Outbound",
                preferred_columns=["Delivered Date", "Creation Date"],
            )
            return _union(mi, m2)

        if "total lead time" in t and "-r" in t:
            try:
                xls = pd.ExcelFile(excel_path, engine="openpyxl")
                for s in xls.sheet_names:
                    sl = (s or "").lower()
                    if "total lead time" in sl and "preformance" in sl and "-r" in sl:
                        return _distinct_month_abbrs_from_sheet(
                            excel_path,
                            s,
                            preferred_columns=["Month", "OB Distribution Date"],
                        )
            except Exception:
                pass
            return []

        if "total lead time" in t or t == "outbound":
            tlsn = _total_lead_time_sheet_for_months(excel_path)
            if tlsn:
                return _distinct_month_abbrs_from_sheet(
                    excel_path,
                    tlsn,
                    preferred_columns=["Month", "OB Distribution Date"],
                )
            return []

        if "dashboard" in t or "meeting" in t:
            return []

        if "rejection" in t and "return" not in t:
            try:
                xls = pd.ExcelFile(excel_path, engine="openpyxl")
                for cand in ("Rejection", "Rejection breakdown"):
                    sn = _resolve_workbook_sheet_name(excel_path, cand)
                    if sn:
                        m = _distinct_month_abbrs_from_sheet(
                            excel_path, sn, preferred_columns=None
                        )
                        if m:
                            return m
            except Exception:
                pass
            return []

        # شيت خام: الاسم كما في الإكسل (الأصل من الطلب ثم tab_key)
        try:
            xls = pd.ExcelFile(excel_path, engine="openpyxl")
            names = set(xls.sheet_names)
            for candidate in (tab_raw, tab_key):
                if not candidate:
                    continue
                if candidate in names:
                    return _distinct_month_abbrs_from_sheet(
                        excel_path, candidate, preferred_columns=None
                    )
                sl = candidate.strip().lower()
                for s in xls.sheet_names:
                    if (s or "").strip().lower() == sl:
                        return _distinct_month_abbrs_from_sheet(
                            excel_path, s, preferred_columns=None
                        )
        except Exception:
            pass
        return []

    def _ajax_tab_json(self, request, tab_for_months, data):
        """JsonResponse للتابات مع available_months حسب الشيت."""
        if isinstance(data, dict):
            data = dict(data)
            if "available_months" not in data:
                if list(data.keys()) == ["error"]:
                    data["available_months"] = []
                else:
                    data["available_months"] = self._available_months_for_tab(
                        request, tab_for_months
                    )
        return JsonResponse(data, safe=False)

    @staticmethod
    def safe_format_value(val):
        if pd.isna(val) or val is pd.NaT:
            return ""
        elif isinstance(val, pd.Timestamp):
            if val.tzinfo is not None:
                val = val.tz_convert(None)
            return val.strftime("%Y-%m-%d %H:%M:%S")
        return val

    # ----------------------------------------------------
    # 🔧 Helper methods for month normalization & filtering
    # ----------------------------------------------------
    def normalize_month_label(self, month_value):
        if month_value is None:
            return None

        raw = str(month_value).strip()
        if not raw:
            return None

        lower = raw.lower()
        if lower in self.MONTH_LOOKUP:
            return self.MONTH_LOOKUP[lower]

        first_three = lower[:3]
        if first_three in self.MONTH_LOOKUP:
            return self.MONTH_LOOKUP[first_three]

        try:
            parsed = _excel_dates_to_datetime(raw, errors="coerce")
            if not pd.isna(parsed):
                return parsed.strftime("%b")
        except Exception:
            pass

        return raw[:3].capitalize()

    def _value_matches_month(self, value, month_lower):
        if value is None:
            return False
        normalized = self.normalize_month_label(value)
        return normalized is not None and normalized.lower() == month_lower

    def _column_matches_month(self, column, month_lower):
        if column is None:
            return False
        col_lower = str(column).strip().lower()
        if col_lower == month_lower:
            return True
        if col_lower.startswith(month_lower + " "):
            return True
        if col_lower.endswith(" " + month_lower):
            return True
        if col_lower.startswith(month_lower + "-") or col_lower.endswith(
            "-" + month_lower
        ):
            return True
        if col_lower.startswith(month_lower + "/") or col_lower.endswith(
            "/" + month_lower
        ):
            return True
        if col_lower.startswith(month_lower + "("):
            return True
        if col_lower.split(" ")[0] == month_lower:
            return True
        if col_lower.replace(".", "").startswith(month_lower):
            return True
        return False

    def _is_month_column(self, column):
        if column is None:
            return False
        col_lower = str(column).strip().lower()
        if col_lower in self.MONTH_LOOKUP:
            return True
        first_three = col_lower[:3]
        if first_three in self.MONTH_PREFIXES:
            return True
        col_split = col_lower.replace("/", " ").replace("-", " ").split()
        if col_split and col_split[0][:3] in self.MONTH_PREFIXES:
            return True
        return False

    def _is_aggregate_column(self, column):
        if column is None:
            return False
        col_lower = str(column).strip().lower()
        if col_lower in self.AGGREGATE_COLUMN_KEYWORDS:
            return True
        compact = col_lower.replace(" ", "")
        if compact in {"target%", "target(%)", "total%"}:
            return True
        if col_lower.isdigit():
            try:
                if int(col_lower) >= 1900:
                    return True
            except ValueError:
                pass
        return False

    def _append_missing_month_messages(self, tab_data, missing_months):
        if not missing_months:
            return

        message_table = {
            "title": "Missing Months",
            "columns": ["Message"],
            "data": [
                {"Message": f"No data available for month {month}."}
                for month in missing_months
            ],
        }

        if isinstance(tab_data.get("sub_tables"), list):
            tab_data["sub_tables"] = [
                sub
                for sub in tab_data["sub_tables"]
                if sub.get("title") != "Missing Months"
            ]
            tab_data["sub_tables"].append(message_table)
            return

        # في حال كان التاب عبارة عن جدول واحد فقط، نحوله إلى sub_tables
        columns = tab_data.pop("columns", None)
        data_rows = tab_data.pop("data", None)
        if columns is not None and data_rows is not None:
            existing_table = {
                "title": tab_data.get("name", "Data"),
                "columns": columns,
                "data": data_rows,
            }
            tab_data["sub_tables"] = [existing_table, message_table]
        else:
            tab_data["sub_tables"] = [message_table]

    def apply_month_filter_to_tab(
        self, tab_data, selected_month=None, selected_months=None
    ):
        if not tab_data:
            return None

        selected_months_norm = []
        if selected_months:
            if isinstance(selected_months, str):
                selected_months = [selected_months]
            seen = set()
            for month in selected_months:
                norm = self.normalize_month_label(month)
                if norm and norm.lower() not in seen:
                    seen.add(norm.lower())
                    selected_months_norm.append(norm)

        month_norm = self.normalize_month_label(selected_month)
        month_filters = []
        if selected_months_norm:
            month_filters = selected_months_norm
        elif month_norm:
            month_filters = [month_norm]
        else:
            tab_data.pop("selected_month", None)
            tab_data.pop("selected_months", None)
            return None

        month_filters_lower = [m.lower() for m in month_filters]
        matched_months = set()

        def matches_any_month(column):
            if not month_filters_lower:
                return False
            for month_lower in month_filters_lower:
                if self._column_matches_month(column, month_lower):
                    matched_months.add(month_lower)
                    return True
            return False

        def value_matches_month(value):
            if not month_filters_lower:
                return False
            normalized = self.normalize_month_label(value)
            if not normalized:
                return False
            val_lower = normalized.lower()
            if val_lower in month_filters_lower:
                matched_months.add(val_lower)
                return True
            return False

        def filter_columns(columns):
            filtered = []
            for col in columns:
                if self._is_month_column(col):
                    if matches_any_month(col):
                        filtered.append(col)
                elif self._is_aggregate_column(col) and not self._column_matches_month(
                    col,
                    month_filters_lower[0] if month_filters_lower else "",
                ):
                    continue
                else:
                    filtered.append(col)
            return filtered if filtered else columns

        def filter_rows(data_rows, columns):
            if not data_rows:
                return data_rows

            month_cols = [
                col
                for col in columns
                if str(col).strip().lower() in {"month", "month name", "monthname"}
            ]
            if not month_cols:
                return data_rows

            month_col = month_cols[0]
            scoped_rows = []
            for row in data_rows:
                value = None
                if isinstance(row, dict):
                    value = row.get(month_col)
                if value_matches_month(value):
                    scoped_rows.append(row)
            return scoped_rows if scoped_rows else data_rows

        if "sub_tables" in tab_data and isinstance(tab_data["sub_tables"], list):
            for sub in tab_data["sub_tables"]:
                if not isinstance(sub, dict):
                    continue
                # ✅ الحفاظ على chart_data في sub_table
                sub_chart_data = sub.get("chart_data", [])

                columns = sub.get("columns", [])
                if columns:
                    filtered_columns = filter_columns(columns)
                    if sub.get("data"):
                        new_data = []
                        for row in sub["data"]:
                            if isinstance(row, dict):
                                new_row = {
                                    col: row.get(col, "") for col in filtered_columns
                                }
                            else:
                                new_row = row
                            new_data.append(new_row)
                        sub["data"] = filter_rows(new_data, filtered_columns)
                    sub["columns"] = filtered_columns

                # ✅ إعادة إضافة chart_data إلى sub_table بعد التعديل (حتى لو كانت فارغة)
                sub["chart_data"] = sub_chart_data
        else:
            columns = tab_data.get("columns", [])
            data_rows = tab_data.get("data", [])
            if columns:
                filtered_columns = filter_columns(columns)
                if data_rows:
                    new_rows = []
                    for row in data_rows:
                        if isinstance(row, dict):
                            new_row = {
                                col: row.get(col, "") for col in filtered_columns
                            }
                        else:
                            new_row = row
                        new_rows.append(new_row)
                    tab_data["data"] = filter_rows(new_rows, filtered_columns)
                tab_data["columns"] = filtered_columns

        if "chart_data" in tab_data and isinstance(tab_data["chart_data"], list):
            for chart in tab_data["chart_data"]:
                if not isinstance(chart, dict):
                    continue
                points = chart.get("dataPoints")
                if not points:
                    continue
                filtered_points = []
                for point in points:
                    label_norm = self.normalize_month_label(point.get("label"))
                    if label_norm and label_norm.lower() in month_filters_lower:
                        matched_months.add(label_norm.lower())
                        filtered_points.append(point)
                if filtered_points:
                    chart["dataPoints"] = filtered_points

        if selected_months_norm:
            tab_data["selected_months"] = selected_months_norm
            return selected_months_norm[0]
        else:
            tab_data["selected_month"] = month_filters[0]
            return month_filters[0]

    @method_decorator(never_cache, name="get")
    def get(self, request):
        print("🟢 [GET] Loading main dashboard with Overview/All-in-One tabs")
        # لا نمسح الكاش هنا حتى يبقى التحميل التالي أسرع (يُمسح عند رفع ملف جديد)

        # مسح بيانات الجلسة أولاً لو الطلب clear_excel (حتى تظهر رسالة رفع الملف)
        action_param = request.GET.get("action", "").strip().lower()
        if action_param == "clear_excel":
            request.session.pop("uploaded_excel_path", None)
            request.session.pop("dashboard_excel_path", None)
            try:
                request.session.save()
            except Exception:
                pass
            from django.shortcuts import redirect
            return redirect(request.path or "/")

        # --------------------------
        # Resolve Excel path — نفتح الصفحة عادي لو في ملف (جلسة أو مجلد)، بدون إجبار على صفحة الرفع
        # --------------------------
        excel_path = self.resolve_excel_file_path(request)
        data_is_uploaded = bool(excel_path)
        print(
            f"🟢 [GET] data_is_uploaded={data_is_uploaded} "
            f"excel_path={'—' if not excel_path else excel_path}"
        )
        if not data_is_uploaded:
            form = ExcelUploadForm()
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "data_is_uploaded": False,
                    "excel_workbook_cache_sig": "",
                },
            )

        # --------------------------
        # مسح كاش الداتا (لو غيّرت الملف يدوياً أو حابب تعيد القراءة من الإكسل)
        # استخدم: ?clear_excel_cache=1 في الرابط
        # --------------------------
        if request.GET.get("clear_excel_cache"):
            try:
                deleted, _ = ExcelSheetCache.objects.all().delete()
                print(f"🗑️ [Cache] تم مسح كاش الداتا ({deleted} سجلات). الداتا هتُقرأ من الملف في الطلب الجاي.")
                messages.success(
                    request,
                    "تم مسح كاش الداتا. التابات هتقرأ من ملف الإكسل في المرة الجاية.",
                )
            except Exception as e:
                print(f"⚠️ [Cache] {e}")
            from django.http import HttpResponseRedirect
            from urllib.parse import urlencode
            q = {k: v for k, v in request.GET.items() if k != "clear_excel_cache"}
            return HttpResponseRedirect(request.path + ("?" + urlencode(q) if q else ""))

        # --------------------------
        # Read request parameters
        # --------------------------
        # الافتراضي = dashboard (تحميل أخف؛ تاب All-in-One اختياري من الرابط ?tab=all فقط)
        selected_tab = request.GET.get("tab", "").strip().lower() or "dashboard"
        selected_month = request.GET.get("month", "").strip()
        selected_quarter = request.GET.get("quarter", "").strip()
        action = request.GET.get("action", "").lower()
        status = request.GET.get("status")

        print(f"🔹 Selected tab: {selected_tab}")
        print(f"🔹 Selected month: {selected_month}")
        print(f"🔹 Selected quarter: {selected_quarter}")
        print(f"🔹 Action: {action}")

        print("🛰️ Quarter AJAX Triggered:", request.GET.get("quarter"))

        quarter_months = []
        quarter_error = None
        if selected_quarter:
            try:
                quarter_months = self._resolve_quarter_months(selected_quarter)
            except ValueError as exc:
                quarter_error = str(exc)

        effective_month = None if quarter_months else selected_month

        if action == "meeting_points_tab":
            return self.meeting_points_tab(request)

        # ✅ فلتر Meeting Points عبر معامل صريح (يعمل مع أي tab=… بما فيه dashboard)
        if (
            request.headers.get("X-Requested-With") == "XMLHttpRequest"
            and (request.GET.get("meeting_points_only") or "").strip() in ("1", "true", "yes")
            and request.GET.get("status")
        ):
            meeting_html = self.get_meeting_points_section_html(
                request, request.GET.get("status", "all")
            )
            return JsonResponse({"meeting_section_html": meeting_html}, safe=False)

        if action == "export_excel":
            if quarter_error:
                return HttpResponse(quarter_error, status=400)
            return self.export_dashboard_excel(
                request,
                selected_month=effective_month,
                selected_months=quarter_months or None,
            )

        # ====================== بحث Traceability (AJAX) ======================
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" and action == "traceability_search":
            return self.traceability_search(request)

        # ====================== طلبات AJAX ======================
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            print("⚡ [AJAX Request] Received request")

            if quarter_error:
                return JsonResponse({"error": quarter_error})

            tab_filter_map = {
                "dashboard": lambda: self.dashboard_tab(request),
                "all": lambda: self.filter_all_tabs(
                    request,
                    effective_month,
                    selected_months=quarter_months or None,
                ),
                "return & refusal": lambda: self.filter_rejections_combined(
                    request,
                    effective_month,
                    selected_months=quarter_months or None,
                ),
                "rejections": lambda: self.filter_rejections_combined(
                    request,
                    effective_month,
                    selected_months=quarter_months or None,
                ),
                "inbound": lambda: self.filter_dock_to_stock_combined(
                    request,
                    effective_month,
                    selected_months=quarter_months or None,
                ),
                "b2b outbound": lambda: self.filter_total_lead_time_performance(
                    request,
                    effective_month,
                    selected_months=quarter_months or None,
                ),
                "b2c outbound": lambda: self._render_b2c_outbound_tab(
                    request,
                    effective_month,
                    selected_months=quarter_months or None,
                ),
                "outbound": lambda: self.filter_total_lead_time_performance(
                    request,
                    effective_month,
                    selected_months=quarter_months or None,
                ),
                "total lead time performance": lambda: self.filter_total_lead_time_performance(
                    request,
                    effective_month,
                    selected_months=quarter_months or None,
                ),
                "safety kpi": lambda: self._render_safety_kpi_tab(request),
                "traceability kpi": lambda: self._traceability_kpi_tab_response(request),
                "meeting points": lambda: self.meeting_points_tab(request),
                "capacity + expiry": lambda: self.filter_capacity_expiry(
                    request,
                    effective_month,
                    selected_months=quarter_months or None,
                ),
                "expiry": lambda: self.filter_expiry(
                    request,
                    effective_month,
                    selected_months=quarter_months or None,
                ),
                "Expiry": lambda: self.filter_expiry(
                    request,
                    effective_month,
                    selected_months=quarter_months or None,
                ),
            }

            # Iterate available tab filters
            for key, func in tab_filter_map.items():
                if key in selected_tab:
                    print(f"📂 Executing tab filter: {key}")
                    try:
                        result = func()

                        # JsonResponse (مثلاً Meeting Points): ندمج available_months ثم نعيد إرسال JSON
                        if isinstance(result, JsonResponse):
                            try:
                                import json

                                payload = json.loads(
                                    result.content.decode("utf-8")
                                )
                                if isinstance(payload, dict):
                                    return self._ajax_tab_json(
                                        request, selected_tab, payload
                                    )
                            except Exception:
                                pass
                            return result

                        # HttpResponse غير JSON (HTML، إلخ)
                        if isinstance(result, HttpResponse):
                            print(
                                "ℹ️ Filter returned HttpResponse; returning as-is."
                            )
                            return result

                        # Dict/list response → JSON
                        if isinstance(result, dict):
                            return self._ajax_tab_json(request, selected_tab, result)
                        if isinstance(result, list):
                            return JsonResponse(result, safe=False)

                        # String response (likely HTML)
                        if isinstance(result, str):
                            return self._ajax_tab_json(
                                request, selected_tab, {"detail_html": result}
                            )

                        # Fallback conversion
                        return self._ajax_tab_json(
                            request, selected_tab, {"detail_html": str(result)}
                        )

                    except Exception as e:
                        import traceback

                        print("❌ Error while executing tab filter:", key)
                        traceback.print_exc()
                        return JsonResponse(
                            {"error": f"Error in '{key}': {str(e)}"},
                            status=200,
                        )

            # All-in-One (never cached)
            if selected_tab == "all":
                print("🔹 Loading All-in-One tab")
                all_result = self.filter_all_tabs(
                    request=request,
                    selected_month=effective_month,
                    selected_months=quarter_months or None,
                )
                return self._ajax_tab_json(request, selected_tab, all_result)

            # Remaining tabs
            if selected_tab in ["rejections", "return & refusal"]:
                return self._ajax_tab_json(
                    request,
                    selected_tab,
                    self.filter_rejections_combined(
                        request,
                        effective_month,
                        selected_months=quarter_months or None,
                    ),
                )
            # airport / seaport tabs تم إلغاؤها
            elif selected_tab in [
                "outbound",
                "total lead time performance",
                "total lead time preformance",
            ]:
                return self._ajax_tab_json(
                    request,
                    selected_tab,
                    self.filter_total_lead_time_performance(
                        request,
                        effective_month,
                        selected_months=quarter_months or None,
                    ),
                )
            elif selected_tab == "total lead time preformance -r":
                return self._ajax_tab_json(
                    request,
                    selected_tab,
                    self.filter_total_lead_time_roche(request, effective_month),
                )
            # data logger tab تم إلغاؤه
            elif "dock to stock - roche" in selected_tab:
                return self._ajax_tab_json(
                    request,
                    selected_tab,
                    self.filter_dock_to_stock_roche(request, effective_month),
                )
            elif (selected_tab or "").lower() == "inbound":
                return self._ajax_tab_json(
                    request,
                    selected_tab,
                    self.filter_dock_to_stock_combined(
                        request,
                        effective_month,
                        selected_months=quarter_months or None,
                    ),
                )
            elif (selected_tab or "").strip().lower() == "capacity + expiry":
                return self._ajax_tab_json(
                    request,
                    selected_tab,
                    self.filter_capacity_expiry(
                        request,
                        effective_month,
                        selected_months=quarter_months or None,
                    ),
                )
            elif (selected_tab or "").strip().lower() == "safety kpi":
                return self._ajax_tab_json(
                    request, selected_tab, self._render_safety_kpi_tab(request)
                )
            elif (selected_tab or "").strip().lower() == "traceability kpi":
                return self._ajax_tab_json(
                    request, selected_tab, self._traceability_kpi_tab_response(request)
                )
            elif "rejection" in selected_tab:
                return self._ajax_tab_json(
                    request,
                    selected_tab,
                    self.filter_rejection_data(request, effective_month),
                )
            elif "dock to stock" in selected_tab:
                return self._ajax_tab_json(
                    request,
                    selected_tab,
                    self.filter_dock_to_stock_combined(
                        request,
                        effective_month,
                        selected_months=quarter_months or None,
                    ),
                )
            elif "meeting points" in selected_tab:
                return self.meeting_points_tab(request)
            elif selected_tab:
                raw_data = self.render_raw_sheet(request, selected_tab)
                if isinstance(raw_data, dict):
                    return self._ajax_tab_json(request, selected_tab, raw_data)
                return JsonResponse(raw_data, safe=False)
            else:
                return JsonResponse({"error": "⚠️ Please select a tab first."})

        # ====================== الطلب العادي ======================
        # تبويبات الشيتات: openpyxl read_only + كاش (أسرع بكثير من pd.ExcelFile لكل GET)
        all_sheets = []
        try:
            all_sheets = _get_excel_sheet_names_cached(excel_path)
            if not all_sheets:
                raise ValueError("empty sheet list")

            MERGE_SHEETS = ["Urgent orders details", "Outbound details"]
            REJECTION_SHEETS = ["Rejection", "Rejection breakdown"]
            AIRPORT_SHEETS = ["Airport Clearance - Roche", "Airport Clearance - 3PL"]
            SEAPORT_SHEETS = ["Seaport clearance - 3pl", "Seaport clearance - Roche"]
            TOTAL_LEADTIME_SHEETS = [
                "Total lead time preformance",
                "Total lead time preformance -R",
            ]
            DOCK_TO_STOCK_SHEETS = ["Dock to stock", "Dock to stock - Roche"]
            # التابات اللي تحب تاخدها من الداشبورد: عدّل هنا (أسماء كما في الإكسل)
            EXCLUDE_SHEETS_BASE = ["Sheet2"]
            # لو حابب تحذف تابات إضافية: زود أسمائهم هنا (بالضبط كما في الإكسل)
            EXCLUDE_SHEETS_EXTRA = getattr(
                self.__class__, "EXCLUDE_TABS", []
            )  # أو عدّل EXCLUDE_TABS في أول الكلاس
            EXCLUDE_SHEETS = list(EXCLUDE_SHEETS_BASE) + list(EXCLUDE_SHEETS_EXTRA)

            include_only = getattr(self.__class__, "INCLUDE_ONLY_TABS", None)
            if include_only:
                # عرض التابات المذكورة فقط (الاسم كما في الإكسل)
                include_set = {s.strip() for s in include_only}
                filtered_tabs = [t for t in all_sheets if t in include_set]
            else:
                filtered_tabs = [
                    t
                    for t in all_sheets
                    if t not in MERGE_SHEETS
                    and t not in REJECTION_SHEETS
                    and t not in AIRPORT_SHEETS
                    and t not in SEAPORT_SHEETS
                    and t not in TOTAL_LEADTIME_SHEETS
                    and t not in DOCK_TO_STOCK_SHEETS
                    and t not in EXCLUDE_SHEETS
                ]

            virtual_tabs = [
                self.DASHBOARD_TAB_NAME,
                "Inbound",
                "B2B Outbound",
                "B2C Outbound",
                "Capacity + Expiry",
                "Return & Refusal",
                "Safety KPI",
                "Traceability KPI",
                "Meeting Points & Action",
            ]
            if include_only:
                include_set_v = {s.strip() for s in include_only}
                filtered_tabs += [v for v in virtual_tabs if v in include_set_v]
            else:
                filtered_tabs += virtual_tabs

            ordered_tabs = [
                self.DASHBOARD_TAB_NAME,
                "Inbound",
                "B2B Outbound",
                "B2C Outbound",
                "Capacity + Expiry",
                "Return & Refusal",
                "Safety KPI",
                "Traceability KPI",
                "Meeting Points & Action",
            ]

            filtered_tabs = [tab for tab in ordered_tabs if tab in filtered_tabs]
            excel_tabs = [{"original": name, "display": name} for name in filtered_tabs]

        except Exception as e:
            print(f"⚠️ [ERROR] تعذر قراءة الشيتات من الملف: {e}")
            excel_tabs = []
            all_sheets = []

        all_months = _extract_months_from_excel_cached(excel_path, all_sheets)
        months_for_select = self._available_months_for_tab(
            request, (selected_tab or "dashboard").strip().lower()
        )
        if not months_for_select:
            months_for_select = all_months

        meeting_points = MeetingPoint.objects.all().order_by("is_done", "-created_at")
        done_count = meeting_points.filter(is_done=True).count()
        total_count = meeting_points.count()

        # ✅ تحميل سريع: لا نحمّل الـ Overview على السيرفر (يُحمّل لاحقاً عبر AJAX)
        _overview_placeholder = (
            "<div id='overview-placeholder' class='text-center py-5'>"
            "<div class='spinner-border text-primary' role='status' style='width: 3rem; height: 3rem;'>"
            "<span class='visually-hidden'>Loading...</span></div>"
            "<p class='mt-3 text-muted'>Loading overview...</p>"
            "<p class='small text-muted'>First load may take a moment. Next loads will be faster.</p>"
            "</div>"
        )
        all_tab_data = {"detail_html": _overview_placeholder}

        render_context = {
            "data_is_uploaded": True,
            "months": months_for_select,
            "excel_tabs": excel_tabs,
            "active_tab": selected_tab or "dashboard",
            "tab_summaries": [],
            "form": ExcelUploadForm(),
            "meeting_points": meeting_points,
            "done_count": done_count,
            "total_count": total_count,
            "all_tab_data": all_tab_data,
            "raw_tab_data": None,
            # لتفريق كاش التابات في المتصفح عند استبدال ملف الإكسل (mtime+size)
            "excel_workbook_cache_sig": _excel_file_signature(excel_path),
        }
        # تحميل سريع: لا نحمّل بيانات الداشبورد على أول طلب GET (تُحمّل لاحقاً عبر AJAX عند فتح تاب Dashboard)
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
        if is_ajax and (selected_tab or "").lower() == "dashboard":
            try:
                dashboard_ctx = self._get_dashboard_include_context(request)
                render_context["dashboard_missing_data"] = dashboard_ctx.get("dashboard_missing_data", [])
                render_context.update(dashboard_ctx)
            except Exception as e:
                import traceback
                traceback.print_exc()
                print(f"⚠️ [Dashboard include context] {e}")
                render_context.setdefault("dashboard_missing_data", [])
        else:
            render_context["dashboard_missing_data"] = []
            if (selected_tab or "").lower() == "dashboard":
                render_context["load_dashboard_placeholder"] = True

        return render(request, self.template_name, render_context)

    def post(self, request):
        print("📥 [DEBUG] تم استدعاء post()")  # ✅ بداية الدالة

        entered_code = request.POST.get("upload_code", "").strip()
        print(f"🔑 [DEBUG] الكود المدخل: {entered_code}")

        # ✅ التحقق من الكود
        if entered_code != self.correct_code:
            print("❌ [DEBUG] الكود غير صحيح!")
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse(
                    {"error": "❌ Invalid code. Please try again."}, status=403
                )
            messages.error(request, "❌ Invalid code. Please try again.")
            return redirect(request.path)

        # ✅ التحقق من الملف المرفوع
        form = ExcelUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            print("⚠️ [DEBUG] النموذج غير صالح أو لم يتم رفع ملف.")
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse(
                    {"error": "⚠️ Please select an Excel file."}, status=400
                )
            return render(
                request,
                self.template_name,
                {
                    "form": form,
                    "data_is_uploaded": False,
                    "excel_workbook_cache_sig": "",
                },
            )

        # ✅ حفظ الملف (يدعم .xlsx و .xlsm مثل all sheet.xlsm)
        excel_file = form.cleaned_data["excel_file"]
        folder_path = os.path.abspath(
            os.path.join(str(settings.MEDIA_ROOT), "excel_uploads")
        )
        os.makedirs(folder_path, exist_ok=True)
        _try_remove_libreoffice_lock_files(folder_path)
        file_name = getattr(excel_file, "name", "") or ""
        is_dashboard_file = _is_dashboard_excel_filename(file_name)
        # رفع الملف الرئيسي: الافتراضي دمج مع النسخة السابقة؛ replace_workbook=1 يستبدل بالكامل بدون دمج
        _rw = (request.POST.get("replace_workbook") or "").strip().lower()
        replace_workbook = _rw in ("1", "true", "yes", "on")

        if is_dashboard_file:
            # ✅ ملف الداشبورد (شيت تاني): Aramco_Tamer3PL_KPI_Dashboard.xlsx — للتاب Dashboard فقط
            file_path = os.path.join(folder_path, DASHBOARD_EXCEL_FILENAME)
            print(f"📊 [DEBUG] رفع ملف الداشبورد: {file_name} → {file_path}")
        else:
            # ✅ الملف الرئيسي (all_sheet / latest) — لباقي التابات
            ext = os.path.splitext(file_name)[1] or ".xlsx"
            if ext.lower() not in (".xlsx", ".xlsm"):
                ext = ".xlsx"
            file_path = os.path.join(folder_path, "latest" + ext)

        merge_prev_copy = None
        try:
            _prev_src = (
                file_path
                if is_dashboard_file
                else _existing_main_workbook_before_upload(folder_path)
            )
            _do_merge_prev = (
                _prev_src
                and os.path.isfile(_prev_src)
                and (is_dashboard_file or not replace_workbook)
            )
            if _do_merge_prev:
                try:
                    _ext = os.path.splitext(_prev_src)[1] or ".xlsx"
                    merge_prev_copy = os.path.join(
                        folder_path, f"_merge_prev_{uuid.uuid4().hex}{_ext}"
                    )
                    shutil.copy2(_prev_src, merge_prev_copy)
                except Exception as _cp:
                    print(f"⚠️ [Excel merge/clean] نسخ الملف السابق: {_cp}")
                    merge_prev_copy = None
            elif not is_dashboard_file and replace_workbook:
                print(
                    "📄 [upload] replace_workbook: تخطي دمج الملف السابق — الملف المرفوع يصبح المصدر الوحيد لـ latest"
                )

            if not is_dashboard_file:
                # ✅ حذف أي ملف latest قديم (xlsx أو xlsm) لتفادي بقاء ملف بالامتداد الآخر
                for old_name in ("latest.xlsx", "latest.xlsm"):
                    old_path = os.path.join(folder_path, old_name)
                    if os.path.exists(old_path):
                        try:
                            os.chmod(old_path, 0o644)
                            os.remove(old_path)
                            print(f"🗑️ [DEBUG] تم حذف الملف القديم: {old_path}")
                        except Exception as e:
                            print(f"⚠️ [DEBUG] تحذير حذف {old_name}: {e}")
            # إذا نجحت الكتابة عبر ملف مؤقت (ملف قديم مقفول)، لا نعيد chunks() — قد تكون مستنفدة فيُفرّغ الملف
            skip_main_write = False
            if os.path.exists(file_path):
                try:
                    os.chmod(file_path, 0o644)
                    os.remove(file_path)
                    print(f"🗑️ [DEBUG] تم حذف الملف القديم: {file_path}")
                except PermissionError as pe:
                    print(
                        f"⚠️ [DEBUG] تحذير: لا يمكن حذف الملف القديم (PermissionError): {pe}"
                    )
                    temp_path = os.path.join(folder_path, "temp_upload.xlsx")
                    _write_django_uploaded_file_to_disk(excel_file, temp_path)
                    try:
                        os.replace(temp_path, file_path)
                        print(f"✅ [DEBUG] تم استبدال الملف باستخدام os.replace")
                        skip_main_write = True
                    except Exception as replace_error:
                        print(
                            f"⚠️ [DEBUG] تحذير: لا يمكن استبدال الملف: {replace_error}"
                        )
                        file_path = temp_path
                        skip_main_write = True
                except Exception as delete_error:
                    print(f"⚠️ [DEBUG] تحذير: خطأ في حذف الملف القديم: {delete_error}")

            # ✅ حفظ الملف الجديد (كل الشيتات وكل الصفوف)
            if not skip_main_write:
                _write_django_uploaded_file_to_disk(excel_file, file_path)

            try:
                os.chmod(file_path, 0o644)
            except Exception as chmod_error:
                print(f"⚠️ [DEBUG] تحذير: لا يمكن تغيير صلاحيات الملف: {chmod_error}")

            print(f"✅ [DEBUG] تم حفظ الملف بنجاح في: {file_path}")

            # ✅ الملف الرئيسي فقط: تحويل إلى latest.xlsx بكل الشيتات وتحديث كاش الداتا في DB
            if not is_dashboard_file:
                file_path = _normalize_upload_to_latest_xlsx_and_update_cache(
                    file_path, folder_path
                )
                # دائماً اجعل النسخة المرجعية على القرص = latest.xlsx (مسار الجلسة والـ GET يعتمدان عليه)
                canonical_latest = os.path.abspath(
                    os.path.join(folder_path, "latest.xlsx")
                )
                fp_abs = os.path.abspath(file_path)
                if os.path.isfile(fp_abs) and fp_abs != canonical_latest:
                    try:
                        shutil.copy2(fp_abs, canonical_latest)
                        file_path = canonical_latest
                        print(f"✅ [upload] تم توحيد الملف إلى latest.xlsx ({canonical_latest})")
                    except OSError as e:
                        print(f"⚠️ [upload] تعذر النسخ إلى latest.xlsx: {e}")

            file_path = os.path.abspath(file_path)
            if not os.path.isfile(file_path) or os.path.getsize(file_path) < 64:
                raise RuntimeError(
                    "الملف بعد الحفظ غير موجود أو فارغ — تحقق من صلاحيات المجلد أو مساحة القرص."
                )

            # دمج/تنظيف كامل فقط عند وجود نسخة سابقة؛ أول رفع يبقى الملف كما حُفظ (أقل عطل)
            if merge_prev_copy and os.path.isfile(merge_prev_copy):
                try:
                    _merge_clean_excel_workbook(file_path, merge_prev_copy)
                    file_path = os.path.abspath(file_path)
                    if not os.path.isfile(file_path) or os.path.getsize(file_path) < 64:
                        raise RuntimeError("الملف بعد الدمج غير صالح.")
                except Exception as _mc_err:
                    print(f"⚠️ [Excel merge/clean] {_mc_err}")

            # ✅ حفظ المسار في الجلسة حسب نوع الملف (داشبورد أو رئيسي)
            if is_dashboard_file:
                request.session["dashboard_excel_path"] = file_path
                print(f"💾 [DEBUG] تم حفظ مسار ملف الداشبورد في الجلسة: {file_path}")
            else:
                request.session["uploaded_excel_path"] = file_path
                print(f"💾 [DEBUG] تم حفظ مسار الملف الرئيسي في الجلسة: {file_path}")
            request.session.save()

            # ✅ مسح الكاش بعد رفع ملف جديد (Django cache + كاش الشيتات في DB)
            try:
                cache.clear()
                print(f"🗑️ [DEBUG] تم مسح الكاش")
            except Exception as cache_error:
                print(f"⚠️ [DEBUG] تحذير: لا يمكن مسح الكاش: {cache_error}")
            if not is_dashboard_file:
                try:
                    n, _ = ExcelSheetCache.objects.all().delete()
                    print(f"🗑️ [DEBUG] تم مسح كاش الشيتات ({n} سجلات) — التابات ستقرأ من الملف الجديد عند الفتح")
                except Exception as e:
                    print(f"⚠️ [DEBUG] مسح كاش الشيتات: {e}")
                # تسخين خفيف مرة واحدة بعد الرفع (أسماء الشيتات + الشهور) لتسريع أول GET والتابات التالية على السيرفر
                try:
                    _wn = _get_excel_sheet_names_cached(file_path)
                    _extract_months_from_excel_cached(file_path, _wn or [])
                    print("✅ [upload] تم تسخين كاش أسماء الشيتات والشهور على السيرفر")
                except Exception as _warm_err:
                    print(f"⚠️ [upload] تسخين الكاش: {_warm_err}")

            # ✅ إرجاع response
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                from urllib.parse import urlencode

                _q = urlencode(
                    {
                        "tab": "dashboard",
                        "full_data": "0",
                        "_upload": str(int(time.time())),
                    }
                )
                _rel = reverse("dashboard:upload_excel") + "?" + _q
                _redirect = request.build_absolute_uri(_rel)
                print(
                    f"📤 [POST] رفع ناجح — الحجم={os.path.getsize(file_path)} بايت — {file_path}"
                )
                return JsonResponse(
                    {
                        "success": True,
                        "message": "✅ File uploaded successfully!",
                        "redirect_url": _redirect,
                    }
                )
            messages.success(request, "✅ File uploaded successfully!")
            print(
                f"📤 [POST] رفع ناجح — الحجم={os.path.getsize(file_path)} بايت — {file_path}"
            )
            return redirect(request.path)
        except Exception as e:
            import traceback

            error_trace = traceback.format_exc()
            print(f"❌ [DEBUG] خطأ في حفظ الملف: {e}")
            print(f"❌ [DEBUG] Traceback:\n{error_trace}")
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse(
                    {"error": f"❌ Error saving file: {str(e)}"}, status=500
                )
            messages.error(request, f"❌ Error saving file: {str(e)}")
            return redirect(request.path)
        finally:
            if merge_prev_copy and os.path.isfile(merge_prev_copy):
                try:
                    os.remove(merge_prev_copy)
                except Exception:
                    pass

    def export_dashboard_excel(
        self, request, selected_month=None, selected_months=None
    ):
        """
        تحميل الملف الأصلي للإكسل (all_sheet) — نفس الملف المستخدم لكل التابات.
        أولوية: ملف الجلسة المرفوع ثم latest ثم all_sheet في المجلد.
        """
        # استخدام نفس مصدر الملف الذي تُقرأ منه كل التابات (all_sheet / ملف مرفوع)
        excel_path = self.resolve_excel_file_path(request)
        if not excel_path:
            html = (
                "<!DOCTYPE html><html><head><meta charset='utf-8'><title>File not found</title></head><body style='font-family:sans-serif;padding:2rem;'>"
                "<h2>Excel file not found</h2>"
                "<p>Please upload the Excel file first (use <strong>Upload File</strong> on the main page).</p>"
                "<p><a href='javascript:window.close()'>Close this tab</a></p>"
                "</body></html>"
            )
            return HttpResponse(html, status=404, content_type="text/html")

        try:
            # اسم الملف للتنزيل: اسم الملف الأصلي إن أمكن
            download_name = os.path.basename(excel_path)
            if not download_name or download_name == "latest.xlsx":
                download_name = "All_Sheets.xlsx"

            # تحديد نوع المحتوى حسب الامتداد
            ext = os.path.splitext(download_name)[1].lower()
            if ext == ".xlsm":
                content_type = "application/vnd.ms-excel.sheet.macroEnabled.12"
            else:
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            with open(excel_path, "rb") as f:
                file_data = f.read()

            response = HttpResponse(file_data, content_type=content_type)
            response["Content-Disposition"] = (
                f'attachment; filename="{download_name}"'
            )
            return response

        except Exception as e:
            import traceback
            traceback.print_exc()
            return HttpResponse(f"❌ حدث خطأ عند تحميل الملف: {str(e)}", status=500)

    def render_raw_sheet(self, request, sheet_name):
        """عرض أي شيت كجدول خام إذا مفيش فلتر خاص"""
        print(f"🟢 [DEBUG] ✅ دخل على render_raw_sheet() - التاب: {sheet_name}")

        # 📁 جلب مسار ملف الإكسل
        excel_file_path = self.resolve_excel_file_path(request)
        if not excel_file_path:
            print("⚠️ [ERROR] لم يتم العثور على ملف Excel.")
            return {
                "detail_html": "<p class='text-danger'>⚠️ Excel file not found.</p>",
                "count": 0,
            }

        try:
            sheet_names = _get_excel_sheet_names_cached(excel_file_path)
            matching_sheet = next(
                (
                    s
                    for s in sheet_names
                    if s.lower().strip() == sheet_name.lower().strip()
                ),
                None,
            )

            if not matching_sheet:
                print(
                    f"⚠️ [WARNING] التاب '{sheet_name}' غير موجود. الشيتات المتاحة: {sheet_names}"
                )
                return {
                    "detail_html": f"<p class='text-danger'>❌ Tab '{sheet_name}' does not exist in the file.</p>",
                    "count": 0,
                }

            # 🧾 قراءة الشيت المطابق (حد 500 صف لتسريع التحميل)
            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
            df = pd.read_excel(
                excel_file_path,
                sheet_name=matching_sheet,
                engine="openpyxl",
                header=0,
                **_nr_kw,
            )

            # 🧹 تنظيف الأعمدة
            df.columns = df.columns.str.strip().str.title()

            # 🗓️ فلترة حسب الشهر إذا تم اختياره
            selected_month = request.GET.get("month")
            if selected_month:
                date_cols = [c for c in df.columns if "Date" in c]
                if date_cols:
                    df[date_cols[0]] = _excel_dates_to_datetime(df[date_cols[0]], errors="coerce")
                    df["Month"] = df[date_cols[0]].dt.strftime("%b")
                    df = df[df["Month"] == selected_month]

            # 🧩 طباعة حالة البيانات
            if df.empty:
                print(
                    f"⚠️ [WARNING] الشيت '{matching_sheet}' فاضي أو غير موجود بعد الفلترة!"
                )
            else:
                print(
                    f"✅ [INFO] الشيت '{matching_sheet}' اتقرأ بنجاح وفيه {len(df)} صفوف."
                )
                print(f"📋 [COLUMNS] الأعمدة: {list(df.columns)}")

            # 🔢 تجهيز أول 50 صف فقط للعرض
            data = df.head(50).to_dict(orient="records")
            for row in data:
                for col, val in row.items():
                    row[col] = self.safe_format_value(val)

            # 🧩 توليد HTML من التمبلت
            tab_data = {
                "name": matching_sheet,
                "columns": df.columns.tolist(),
                "data": data,
            }
            month_norm = self.apply_month_filter_to_tab(tab_data, selected_month)

            preview_mode = not _excel_full_data_requested(request)
            cap = _excel_max_rows_for_request(request)
            banner = ""
            if preview_mode and cap:
                banner = (
                    "<div class=\"alert alert-info small py-2 mb-3\" role=\"alert\">"
                    f"Fast load: first <strong>{cap}</strong> rows from Excel. "
                    "Use <strong>Load full data</strong> above for more rows (may be slower).</div>"
                )
            html = banner + render_to_string(
                "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
                {"tab": tab_data, "selected_month": month_norm},
            )

            # 📤 إرجاع النتيجة للواجهة
            return {
                "detail_html": html,
                "count": len(df),
                "tab_data": tab_data,
                "excel_preview_mode": preview_mode,
                "excel_row_cap": cap,
            }

        except Exception as e:
            print(f"❌ [ERROR] أثناء قراءة الشيت '{sheet_name}': {e}")
            return {
                "detail_html": f"<p class='text-danger'>⚠️ Error while reading sheet: {e}</p>",
                "count": 0,
            }

    def filter_by_month(self, request, selected_month):
        import pandas as pd
        from django.template.loader import render_to_string

        try:
            excel_file_path = self.resolve_excel_file_path(request)
            if not excel_file_path:
                return {"error": "⚠️ Excel file not found."}
            xls = pd.ExcelFile(excel_file_path, engine="openpyxl")

            # 🧩 تحديد اسم الشيت المطلوب تلقائيًا
            # نحاول نختار شيت يحتوي على "Data logger" أو "Dock to stock"
            possible_sheets = [
                s
                for s in xls.sheet_names
                if any(key in s.lower() for key in ["data logger", "dock to stock"])
            ]

            if not possible_sheets:
                print(
                    "⚠️ لم يتم العثور على أي شيت يحتوي على Data logger أو Dock to stock"
                )
                return {
                    "error": "⚠️ No sheet containing Data logger or Dock to stock was found."
                }

            sheet_name = possible_sheets[0]  # ناخد أول واحد مطابق
            print(f"📄 قراءة الشيت: {sheet_name}")
            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
            df = pd.read_excel(
                excel_file_path,
                sheet_name=sheet_name,
                engine="openpyxl",
                header=0,
                **_nr_kw,
            )
        except Exception as e:
            return {"error": f"⚠️ Unable to read the tab: {e}"}

        # تنظيف الأعمدة
        df.columns = df.columns.str.strip()

        # التحقق من عمود التاريخ
        if "Month" not in df.columns:
            return {"error": "⚠️ Column 'Month' is missing."}

        # تحويل/تطبيع عمود الشهر لقبول كل الصيغ (تاريخ، اختصار، اسم كامل، رقم 1-12)
        import calendar

        month_raw = df["Month"]
        # حاول تحويله لتاريخ؛ اللي يفشل هنرجّعه نصياً
        parsed = _excel_dates_to_datetime(month_raw, errors="coerce")
        month_abbr_from_dates = parsed.dt.strftime("%b")

        # طبّع النصوص: أول 3 حروف من اسم الشهر (Jan/February -> Feb)، والأرقام 1-12 إلى اختصار
        def normalize_month_val(v):
            if pd.isna(v):
                return None
            s = str(v).strip()
            # أرقام
            if s.isdigit():
                n = int(s)
                if 1 <= n <= 12:
                    return calendar.month_abbr[n]
            # أسماء كاملة أو مختصرة
            # جرّب اسم كامل
            for i, mname in enumerate(calendar.month_name):
                if i == 0:
                    continue
                if s.lower() == mname.lower():
                    return calendar.month_abbr[i]
            # جرّب اختصار جاهز أو نص عام -> أول 3 أحرف بحالة Capitalize
            return s[:3].capitalize()

        month_abbr_fallback = month_raw.apply(normalize_month_val)
        # استخدم من التاريخ حيث متاح وإلا fallback
        df["Month"] = month_abbr_from_dates.where(~parsed.isna(), month_abbr_fallback)

        # توحيد تمثيل الشهر المختار (أمان لحالات الإدخال المختلفة)
        selected_month_norm = (
            str(selected_month).strip().capitalize() if selected_month else None
        )

        # حفظ الشهر في الجلسة ليستخدمه باقي التابات عند الاستعلامات اللاحقة
        try:
            if selected_month_norm:
                request.session["selected_month"] = selected_month_norm
        except Exception:
            # في حال عدم توفر الجلسة (مثلاً في طلبات غير مرتبطة بمستخدم)، نتجاوز بهدوء
            pass

        # فلترة الشهر المختار أولاً
        month_df = df[df["Month"] == selected_month_norm]

        if month_df.empty:
            return {
                "error": f"⚠️ لا توجد بيانات متاحة للشهر {selected_month_norm}.",
                "month": selected_month_norm,
                "sheet_name": sheet_name,
            }

        # البحث عن عمود KPI بشكل مرن (ممكن يكون اسمه مختلف)
        kpi_miss_col = None
        possible_kpi_names = [
            "kpi miss in",
            "kpi miss",
            "kpi",
            "miss",
            "clearance handling kpi",
            "transit kpi",
        ]

        for kpi_name in possible_kpi_names:
            kpi_miss_col = next(
                (col for col in df.columns if str(col).strip().lower() == kpi_name),
                None,
            )
            if kpi_miss_col:
                break

        # حساب الإحصائيات
        total = len(month_df.drop_duplicates())

        # لو وجدنا عمود KPI، نحسب Miss
        if kpi_miss_col:
            miss_df = month_df[month_df[kpi_miss_col].astype(str).str.lower() == "miss"]
            miss_count = len(miss_df)
            valid = total - miss_count
        else:
            # لو مفيش عمود KPI، نعرض كل البيانات بدون فلترة Miss
            miss_df = pd.DataFrame()  # جدول فاضي
            miss_count = 0
            valid = total
            print(
                f"⚠️ لم يتم العثور على عمود KPI، سيتم عرض جميع البيانات للشهر {selected_month_norm}"
            )

        # تحويل النتائج إلى HTML (للحفاظ على التوافق مع أي استخدام حالي)
        dedup_html = month_df.to_html(
            classes="table table-bordered table-hover text-center",
            index=False,
            border=0,
        )
        miss_html = miss_df.to_html(
            classes="table table-bordered table-hover text-center text-danger",
            index=False,
            border=0,
        )

        print(
            f"📆 فلترة الشهر {selected_month}: إجمالي={total}, Miss={miss_count}, Valid={valid}"
        )

        hit_pct = int(round((valid / total) * 100)) if total else 0

        # تجهيز البيانات للتمبلت القياسي (جداول + شارت)
        month_df_display = month_df.fillna("").astype(str)
        sub_tables = [
            {
                "title": f"{sheet_name} – {selected_month_norm} (كل السجلات)",
                "columns": month_df_display.columns.tolist(),
                "data": month_df_display.to_dict(orient="records"),
            }
        ]

        if miss_count > 0:
            miss_df_display = miss_df.fillna("").astype(str)
            sub_tables.append(
                {
                    "title": f"{sheet_name} – {selected_month_norm} (السجلات المتأخرة)",
                    "columns": miss_df_display.columns.tolist(),
                    "data": miss_df_display.to_dict(orient="records"),
                }
            )

        summary_table = [
            {"المؤشر": "إجمالي الشحنات", "القيمة": int(total)},
            {"المؤشر": "شحنات صحيحة", "القيمة": int(valid)},
            {"المؤشر": "شحنات Miss", "القيمة": int(miss_count)},
            {"المؤشر": "Hit %", "القيمة": f"{hit_pct}%"},
        ]
        sub_tables.append(
            {
                "title": f"{sheet_name} – {selected_month_norm} (ملخص الأداء)",
                "columns": ["المؤشر", "القيمة"],
                "data": summary_table,
            }
        )

        chart_title = f"{sheet_name} – {selected_month_norm} Performance"
        chart_data = [
            {
                "title": chart_title,
                "type": "column",
                "name": "Valid Shipments",
                "color": "#4caf50",
                "showInLegend": True,
                "dataPoints": [{"label": selected_month_norm, "y": int(valid)}],
                "related_table": sub_tables[0]["title"],
            },
            {
                "title": chart_title,
                "type": "column",
                "name": "Miss Shipments",
                "color": "#f44336",
                "showInLegend": True,
                "dataPoints": [{"label": selected_month_norm, "y": int(miss_count)}],
                "related_table": sub_tables[0]["title"],
            },
            {
                "title": chart_title,
                "type": "line",
                "name": "Hit %",
                "color": "#1976d2",
                "showInLegend": True,
                "dataPoints": [{"label": selected_month_norm, "y": hit_pct}],
                "related_table": sub_tables[-1]["title"],
            },
        ]

        tab_data = {
            "name": f"{sheet_name} ({selected_month_norm})",
            "sub_tables": sub_tables,
            "chart_data": chart_data,
            "chart_title": chart_title,
        }
        month_norm_filtered = self.apply_month_filter_to_tab(
            tab_data, selected_month_norm
        )

        combined_html = render_to_string(
            "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
            {"tab": tab_data, "selected_month": month_norm_filtered},
        )

        return {
            "month": selected_month_norm,
            "selected_month": selected_month_norm,
            "sheet_name": sheet_name,
            "total_shipments": total,
            "miss_count": miss_count,
            "valid_shipments": valid,
            "hit_pct": hit_pct,
            "dedup_html": dedup_html,
            "miss_html": miss_html,
            "html": combined_html,
            "detail_html": combined_html,
            "chart_data": chart_data,
            "chart_title": chart_title,
            "tab_data": tab_data,
        }

    def _resolve_quarter_months(self, selected_quarter):
        if not selected_quarter:
            return []

        import re

        quarter_pattern = re.compile(r"^Q([1-4])(?:[-\s]?(\d{4}))?$", re.IGNORECASE)
        match = quarter_pattern.match(str(selected_quarter).strip())
        if not match:
            raise ValueError(f"⚠️ كورتر غير معروف: {selected_quarter}")

        quarter_number = int(match.group(1))
        quarter_months_map = {
            1: ["Jan", "Feb", "Mar"],
            2: ["Apr", "May", "Jun"],
            3: ["Jul", "Aug", "Sep"],
            4: ["Oct", "Nov", "Dec"],
        }

        months = quarter_months_map.get(quarter_number, [])
        if not months:
            raise ValueError(f"⚠️ لا توجد شهور معرّفة للكوارتر {selected_quarter}.")
        return months

    def filter_by_quarter(self, request, selected_quarter):
        from django.template.loader import render_to_string
        import re

        if not selected_quarter:
            return {"error": "⚠️ Please select a valid quarter."}

        quarter_pattern = re.compile(r"^Q([1-4])(?:[-\s]?(\d{4}))?$", re.IGNORECASE)
        match = quarter_pattern.match(str(selected_quarter).strip())
        if not match:
            return {"error": f"⚠️ Unknown quarter: {selected_quarter}"}

        quarter_number = int(match.group(1))
        quarter_months_map = {
            1: ["Jan", "Feb", "Mar"],
            2: ["Apr", "May", "Jun"],
            3: ["Jul", "Aug", "Sep"],
            4: ["Oct", "Nov", "Dec"],
        }

        display_month_list = quarter_months_map.get(quarter_number, [])
        if not display_month_list:
            return {
                "error": f"⚠️ No months were defined for quarter {selected_quarter}."
            }

        try:
            total_lead_time_result = self.filter_total_lead_time_performance(
                request, selected_months=display_month_list
            )
        except Exception as exc:
            import traceback

            traceback.print_exc()
            total_lead_time_result = {
                "detail_html": f"<p class='text-danger text-center p-4'>⚠️ Error while loading Total Lead Time Performance: {exc}</p>"
            }

        section_html = (
            total_lead_time_result.get("detail_html")
            or total_lead_time_result.get("html")
            or "<p class='text-warning text-center p-4'>⚠️ No data available for this quarter.</p>"
        )

        section_wrapper = f"""
        <section class="quarter-section mb-5">
            <div class="d-flex justify-content-between align-items-center mb-3">
                <h4 class="mb-0 text-primary">Total Lead Time Performance – Quarter {selected_quarter}</h4>
                <span class="badge bg-light text-dark px-3 py-2">{', '.join(display_month_list)}</span>
            </div>
            {section_html}
        </section>
        """

        header_html = f"""
        <div class="quarter-header text-center mb-4">
            <h3 class="fw-bold text-primary mb-1">Quarter {selected_quarter}</h3>
            <p class="text-muted mb-0">Months in scope: {', '.join(display_month_list)}</p>
        </div>
        """

        combined_html = (
            f"<div class='quarter-wrapper'>{header_html}{section_wrapper}</div>"
        )

        return {
            "quarter": selected_quarter,
            "months": ", ".join(display_month_list),
            "detail_html": combined_html,
            "html": combined_html,
            "chart_data": total_lead_time_result.get("chart_data", []),
            "chart_title": total_lead_time_result.get("chart_title"),
            "hit_pct": total_lead_time_result.get("hit_pct"),
        }

    def filter_all_tabs(self, request=None, selected_month=None, selected_months=None):
        try:
            month_for_filters = selected_month if not selected_months else None
            # ✅ تحديد الفلتر الحالي
            status_filter = "all"
            if request is not None and hasattr(request, "GET"):
                status_filter = request.GET.get("status", "all")

            # ✅ الحصول على مسار ملف Excel
            excel_path = self.resolve_excel_file_path(request)
            if not excel_path:
                html = render_to_string(
                    "components/ui-kits/tab-bootstrap/components/dashboard-overview.html",
                    {"message": "⚠️ لم يتم العثور على ملف Excel."},
                )
                return {"detail_html": html}

            # ✅ تخزين مؤقت لنتيجة الـ overview (10 دقائق) — أول فتح لتاب All-in-One بيكون أثقل، الباقي من الكاش
            import hashlib
            _path_hash = hashlib.md5((excel_path or "").encode()).hexdigest()[:12]
            _month = (selected_month or "") + "_" + (str(selected_months) if selected_months else "")
            _full_flag = "1" if request and _excel_full_data_requested(request) else "0"
            # v2: per-tab multiple chart series + merged chart_data_pods in overview_tab
            _cache_key = f"tlp_overview_{_path_hash}_{_month}_{status_filter}_{_full_flag}_v2"
            overview_data = cache.get(_cache_key)
            if overview_data is None:
                overview_data = self.overview_tab(
                    request=request,
                    selected_month=month_for_filters,
                    selected_months=selected_months,
                    from_all_in_one=True,
                )
                cache.set(_cache_key, overview_data, 600)

            if not overview_data or "tab_cards" not in overview_data:
                html = render_to_string(
                    "components/ui-kits/tab-bootstrap/components/dashboard-overview.html",
                    {"message": "⚠️ لا توجد بيانات متاحة من overview_tab."},
                )
                return {"detail_html": html}

            # ✅ قائمة التابات المحذوفة
            excluded_tabs = [
                "return & refusal",
                "airport clearance",
                "seaport clearance",
                "data logger measurement",
            ]

            clean_tabs = []
            for tab in overview_data.get("tab_cards", []):
                name = tab.get("name", "غير معروف")
                name_lower = name.strip().lower()

                # ✅ حذف التابات المطلوبة
                if name_lower in excluded_tabs:
                    continue

                # ✅ النسبة الفعلية
                try:
                    hit = float(tab.get("hit_pct", 0))
                except Exception:
                    hit = 0
                hit = int(round(max(0, min(hit, 100))))

                # ✅ التارجت
                try:
                    target = float(tab.get("target_pct", 100))
                except Exception:
                    target = 100

                # ✅ نستخدم أي chart_data و chart_type راجعين من الـ overview كما هي
                chart_data = tab.get("chart_data", []) or []
                chart_type = tab.get("chart_type", "bar")

                clean_tabs.append(
                    {
                        "name": name,
                        "hit_pct": hit,
                        "target_pct": int(target),
                        "count": tab.get("count", 0),
                        "chart_type": chart_type,
                        "chart_data": chart_data,
                    }
                )

            # ✅ ترتيب التابات حسب الأولوية (Safety KPI و Traceability KPI من tabs_order)
            desired_order = [
                "Inbound",
                "B2B Outbound",
                "B2C Outbound",
                "Return & Refusal",
                "Safety KPI",
                "Traceability KPI",
            ]
            clean_tabs.sort(
                key=lambda x: (
                    desired_order.index(x["name"])
                    if x["name"] in desired_order
                    else len(desired_order)
                )
            )

            # ✅ بيانات الميتنج - جلب كل النقاط (مثل meeting_points_tab)
            meeting_points = MeetingPoint.objects.all().order_by(
                "is_done", "-created_at"
            )

            if status_filter == "done":
                meeting_points = meeting_points.filter(is_done=True)
            elif status_filter == "pending":
                meeting_points = meeting_points.filter(is_done=False)

            meeting_data = [
                {
                    "id": p.id,
                    "description": p.description,
                    "assigned_to": getattr(p, "assigned_to", "") or "",
                    "status": "Done" if p.is_done else "Pending",
                    "created_at": p.created_at,
                    "target_date": p.target_date,
                }
                for p in meeting_points
            ]

            tabs_for_display = clean_tabs

            html = render_to_string(
                "components/ui-kits/tab-bootstrap/components/dashboard-overview.html",
                {
                    "tabs": tabs_for_display,
                    "tabs_json": json.dumps(tabs_for_display),
                    "meeting_data": meeting_data,
                    "status_filter": status_filter,
                },
                request=request,
            )

            return {"detail_html": html}

        except Exception as e:
            traceback.print_exc()
            return {
                "detail_html": f"<div class='alert alert-danger'>⚠️ Error: {e}</div>"
            }

    def get_meeting_points_section_html(self, request, status_filter="all"):
        """
        ✅ دالة مساعدة لإرجاع HTML قسم Meeting Points فقط
        """
        try:
            meeting_points = MeetingPoint.objects.all().order_by(
                "is_done", "-created_at"
            )

            if status_filter == "done":
                meeting_points = meeting_points.filter(is_done=True)
            elif status_filter == "pending":
                meeting_points = meeting_points.filter(is_done=False)

            meeting_data = [
                {
                    "id": p.id,
                    "description": p.description,
                    "assigned_to": getattr(p, "assigned_to", "") or "",
                    "status": "Done" if p.is_done else "Pending",
                    "created_at": p.created_at,
                    "target_date": p.target_date,
                }
                for p in meeting_points
            ]

            # ✅ إرجاع HTML قسم Meeting Points فقط
            html = render_to_string(
                "components/ui-kits/tab-bootstrap/components/meeting_points_section.html",
                {
                    "meeting_data": meeting_data,
                    "status_filter": status_filter,
                },
                request=request,
            )
            return html
        except Exception as e:
            import traceback

            traceback.print_exc()
            return f"<div class='alert alert-danger'>⚠️ Error: {e}</div>"

    def filter_total_lead_time_detail(self, request, selected_month=None):
        try:
            # تحميل الملف من الجلسة
            excel_path = request.session.get("uploaded_excel_path")
            if not excel_path or not os.path.exists(excel_path):
                return {"error": "⚠️ Excel file was not found in the session."}

            # قراءة الشيت المطلوب
            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
            df = pd.read_excel(
                excel_path,
                sheet_name="Total lead time preformance",
                engine="openpyxl",
                header=0,
                **_nr_kw,
            )
            df.columns = df.columns.str.strip().str.lower()

            # التأكد من الأعمدة المطلوبة
            required_cols = [
                "month",
                "outbound delivery",
                "kpi",
                "reason group",
                "miss reason",
            ]
            for col in required_cols:
                if col not in df.columns:
                    return {"error": f"⚠️ Column '{col}' does not exist in the sheet."}

            # تحويل التاريخ إلى الشهر
            df["month"] = (
                _excel_dates_to_datetime(df["month"], errors="coerce")
                .dt.strftime("%b")
                .str.capitalize()
            )

            # استخراج الشهور الموجودة فعليًا في الملف (بترتيب زمني)
            existing_months = df["month"].dropna().unique().tolist()
            existing_months = sorted(
                existing_months, key=lambda x: _excel_dates_to_datetime(x, format="%b").month
            )

            if not existing_months:
                return {"error": "⚠️ No valid months were found in the file."}

            # إزالة التكرارات حسب رقم الشحنة
            df = df.drop_duplicates(subset=["outbound delivery"])

            # تنظيف النصوص
            df["reason group"] = df["reason group"].astype(str).str.strip().str.lower()
            df["kpi"] = df["kpi"].astype(str).str.strip().str.lower()

            # بيانات Miss الخاصة بـ 3PL فقط
            df_miss_3pl = df[
                (df["kpi"] == "miss") & (df["reason group"] == "3pl")
            ].copy()

            # 🔹 تنظيف السبب فقط (بدون تغيير الحروف الأصلية)
            df_miss_3pl["miss reason"] = (
                df_miss_3pl["miss reason"]
                .astype(str)
                .str.strip()
                .str.replace(r"\s+", " ", regex=True)  # إزالة المسافات المكررة
            )

            # معالجة اختلاف الحروف أثناء التجميع (case-insensitive grouping)
            df_miss_3pl["_miss_reason_key"] = df_miss_3pl["miss reason"].str.lower()

            # بيانات On Time Delivery (Hit)
            df_hit = df[df["kpi"] != "miss"].copy()

            # تجميع Miss حسب السبب والشهر (باستخدام المفتاح الموحد للحروف)
            miss_grouped = (
                df_miss_3pl.groupby(["_miss_reason_key", "month"], as_index=False)
                .agg(
                    {
                        "miss reason": "first",
                        "month": "first",
                        "_miss_reason_key": "count",
                    }
                )
                .rename(columns={"_miss_reason_key": "count"})
            )

            # Pivot الجدول
            miss_pivot = miss_grouped.pivot_table(
                index="miss reason", columns="month", values="count", fill_value=0
            )

            # تأكد أن كل الشهور الموجودة في الملف موجودة في الجدول
            for m in existing_months:
                if m not in miss_pivot.columns:
                    miss_pivot[m] = 0
            miss_pivot = miss_pivot[existing_months]

            # حساب On Time Delivery لكل شهر
            hit_counts = (
                df_hit.groupby("month").size().reindex(existing_months, fill_value=0)
            )

            # بناء الجدول النهائي
            final_df = miss_pivot.copy()
            final_df.loc["On Time Delivery"] = hit_counts
            final_df = final_df.fillna(0)

            # ترتيب الصفوف بحيث On Time في الأعلى
            final_df = final_df.reindex(
                ["On Time Delivery"]
                + [r for r in final_df.index if r != "On Time Delivery"]
            )

            # إضافة عمود الإجمالي
            final_df["Total"] = final_df.sum(axis=1)

            # صف الإجمالي النهائي
            final_df.loc["TOTAL"] = final_df.sum(numeric_only=True)

            # 🟦 إنشاء جدول HTML
            html_table = """
            <table class='table table-bordered text-center align-middle mb-0'>
                <thead class='table-warning'>
                    <tr><th colspan='{colspan}'>Reason From 3PL Side</th></tr>
                </thead>
                <thead class='table-primary'>
                    <tr>
                        <th>KPI</th>
                        {month_headers}
                        <th>2025</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            """

            # رؤوس الأعمدة
            month_headers = "".join([f"<th>{m}</th>" for m in existing_months])

            # الصفوف
            rows_html = ""
            for reason, row in final_df.iterrows():
                rows_html += f"<tr><td>{reason}</td>"
                for m in existing_months:
                    rows_html += f"<td>{int(row[m])}</td>"
                rows_html += f"<td class='fw-bold'>{int(row['Total'])}</td></tr>"

            # استبدال القيم في القالب
            html_table = html_table.format(
                colspan=len(existing_months) + 2,
                month_headers=month_headers,
                table_rows=rows_html,
            )

            # وضع الجدول داخل واجهة مرتبة
            html_output = f"""
            <div class='container-fluid'>
                <h5 class='text-center text-primary mb-3'>KPI Summary - 3PL Performance</h5>
                <div class='card shadow'>
                    <div class='card-body'>
                        {html_table}
                    </div>
                </div>
            </div>
            """

            return {"detail_html": html_output, "months": existing_months}

        except Exception as e:
            import traceback

            print("❌ خطأ أثناء تحليل البيانات:", str(e))
            print(traceback.format_exc())
            return {"error": f"⚠️ Error while analyzing data: {e}"}

    def filter_rejection_data(self, request, month=None):
        print("🟣 [DEBUG] filter_rejection_data CALLED ✅ month:", month)

        excel_path = request.session.get("uploaded_excel_path")

        if not excel_path or not os.path.exists(excel_path):
            return {"error": "⚠️ Excel file not found."}

        try:
            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
            df = pd.read_excel(
                excel_path, sheet_name="Rejection", engine="openpyxl", header=0, **_nr_kw
            )
            print("🟢 [DEBUG] الأعمدة:", df.columns.tolist())
            print(df.head(3))
        except Exception as e:
            return {"error": f"⚠️ Unable to read the 'Rejection' sheet: {e}"}

        df.columns = df.columns.str.strip().str.title()
        required = ["Month", "Total Number Of Orders", "Booking Orders"]
        if not all(col in df.columns for col in required):
            return {
                "error": "⚠️ Required columns (Month, Total Number Of Orders, Booking Orders) are missing."
            }

        if month:
            df = df[df["Month"].astype(str).str.contains(month, case=False, na=False)]

        if df.empty:
            return {"error": "⚠️ No data available."}

        # ✅ خدي القيم زي ما هي من الإكسل (من العمود Booking Orders)
        summary = df[["Month", "Booking Orders"]].copy()

        # 🧠 تنظيف القيم — شيل علامة % لو موجودة وحوّليها لأرقام
        summary["Booking Orders"] = (
            summary["Booking Orders"]
            .astype(str)
            .str.replace("%", "", regex=False)
            .astype(float)
        )

        # 🎯 البيانات للشارت مباشرة
        chart_data = [
            {"month": row["Month"], "percentage": row["Booking Orders"]}
            for _, row in summary.iterrows()
        ]

        html = df.to_html(
            index=False,
            classes="table table-bordered table-striped text-center align-middle",
            border=0,
        )

        print("📊 DEBUG - chart_data:", chart_data)  # <-- شوفيها في التيرمنال
        return {"detail_html": html, "chart_data": chart_data}

    def filter_dock_to_stock_roche(self, request, selected_month=None):
        print("🟢 [DEBUG] ✅ دخل على filter_dock_to_stock_roche()")

        excel_path = request.session.get("uploaded_excel_path")
        if not excel_path or not os.path.exists(excel_path):
            return {"error": "⚠️ Excel file not found."}

        try:
            import pandas as pd
            from django.template.loader import render_to_string

            sheet_name = "Dock to stock - Roche"
            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
            df = pd.read_excel(
                excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
            )
            df.columns = df.columns.astype(str).str.strip()

            if df.empty:
                return {"error": "⚠️ Sheet 'Dock to stock - Roche' is empty."}

            # أول عمود هو الشهر
            month_col = df.columns[0]
            # باقي الأعمدة هي الأسباب (KPIs)
            kpi_cols = df.columns[1:]

            # تحويل البيانات بحيث تكون الأسباب صفوف والشهور أعمدة
            melted_df = df.melt(id_vars=[month_col], var_name="KPI", value_name="Value")

            # Pivot فعلي (KPI كصفوف والشهور كأعمدة)
            pivot_df = melted_df.pivot_table(
                index="KPI", columns=month_col, values="Value", aggfunc="sum"
            ).reset_index()
            pivot_df = pivot_df.rename_axis(None, axis=1)

            # ترتيب الأعمدة حسب تسلسل الشهور الموجود في الشيت الأصلي
            month_order = list(df[month_col].unique())
            ordered_cols = ["KPI"] + month_order
            pivot_df = pivot_df.reindex(columns=ordered_cols)

            # ✅ حذف أي عمود اسمه "Total" (اللي بيتولد من الشيت أو من الخطأ)
            if "Total" in pivot_df.columns:
                pivot_df = pivot_df.drop(columns=["Total"])

            # ✅ إضافة عمود "2025" فقط بعد الشهور
            pivot_df["2025"] = pivot_df.iloc[:, 1:].sum(axis=1)

            # ✅ إضافة صف Total (اللي بيكون تحت الجدول)
            total_row = {"KPI": "Total"}
            for col in pivot_df.columns[1:]:  # تجاهل عمود KPI
                total_row[col] = pivot_df[col].sum()
            pivot_df = pd.concat(
                [pivot_df, pd.DataFrame([total_row])], ignore_index=True
            )

            print("✅ [DEBUG] جدول KPI النهائي بعد التعديل:")
            print(pivot_df.to_string(index=False))

            # تجهيز البيانات للعرض
            columns = list(pivot_df.columns)
            table_data = pivot_df.fillna("").to_dict(orient="records")

            tab = {
                "name": "Dock to Stock - Roche",
                "columns": columns,
                "data": table_data,
            }

            month_norm = self.apply_month_filter_to_tab(tab, selected_month)

            html = render_to_string(
                "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
                {
                    "tab": tab,
                    "table_title": "Dock to Stock - Roche (KPI Summary)",
                    "selected_month": month_norm,
                },
            )

            return {
                "detail_html": html,
                "chart_title": "Dock to Stock - Roche",
            }

        except Exception as e:
            print(f"❌ [ERROR] {e}")
            return {"error": f"⚠️ Error while reading data: {e}"}

    def filter_dock_to_stock_3pl(
        self, request, selected_month=None, selected_months=None
    ):
        try:
            print("🟢 [DEBUG] ✅ دخل على filter_dock_to_stock_3pl()")
            file_path = self.resolve_excel_file_path(request)
            print(f"📁 [DEBUG] مسار الملف المستخدم: {file_path}")

            if not file_path:
                return {"error": "⚠️ File not found."}

            # 🧩 قراءة الشيت
            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
            df = pd.read_excel(
                file_path,
                sheet_name="Dock to stock",
                engine="openpyxl",
                header=0,
                **_nr_kw,
            )
            print(f"📄 [DEBUG] أول 10 صفوف من الشيت Dock to stock:\n{df.head(10)}")

            # ✅ التحقق من وجود الأعمدة المطلوبة
            if "Delv #" not in df.columns or "Month" not in df.columns:
                return {
                    "error": "⚠️ Columns 'Delv #' or 'Month' are missing in the sheet."
                }

            # 🧮 استخراج الشهر من العمود Month
            df["Month"] = _excel_dates_to_datetime(df["Month"], errors="coerce").dt.strftime("%b")

            selected_months_norm = []
            if selected_months:
                if isinstance(selected_months, str):
                    selected_months = [selected_months]
                seen = set()
                for m in selected_months:
                    norm = self.normalize_month_label(m)
                    if norm and norm not in seen:
                        seen.add(norm)
                        selected_months_norm.append(norm)

            selected_month_norm = (
                self.normalize_month_label(selected_month)
                if selected_month and not selected_months_norm
                else None
            )
            if selected_months_norm:
                df = df[
                    df["Month"]
                    .str.lower()
                    .isin([m.lower() for m in selected_months_norm])
                ]
                if df.empty:
                    return {
                        "detail_html": "<p class='text-warning text-center'>⚠️ No data available for the selected quarter months.</p>",
                        "chart_data": [],
                    }
            elif selected_month_norm:
                df = df[df["Month"].str.lower() == selected_month_norm.lower()]
                if df.empty:
                    return {
                        "detail_html": "<p class='text-warning text-center'>⚠️ No data available for this month.</p>",
                        "chart_data": [],
                    }

            # 🧱 حذف الصفوف اللي مافيهاش شهر
            df = df.dropna(subset=["Month"])

            # ✅ حساب عدد الشحنات الفريدة (hit) لكل شهر من العمود Delv #
            hits_per_month = (
                df.drop_duplicates(subset=["Delv #"])
                .groupby("Month")["Delv #"]
                .count()
                .reset_index(name="Hits")
            )

            print("📊 [DEBUG] نتائج عدد الشحنات الفريدة لكل شهر:")
            print(hits_per_month)

            # ✅ حساب إجمالي الشحنات (Total) لكل شهر قبل حذف المكرر
            total_per_month = (
                df.groupby("Month")["Delv #"]
                .count()
                .reset_index(name="Total_Shipments")
            )

            # ✅ دمج نتائج الـ hits مع الإجمالي
            merged = pd.merge(hits_per_month, total_per_month, on="Month", how="left")

            # ✅ حساب نسبة التارجت لكل شهر
            merged["Target_%"] = (
                merged["Hits"] / merged["Total_Shipments"] * 100
            ).round(2)

            print("📈 [DEBUG] بعد حساب نسبة التارجت:")
            print(merged)

            # ✅ تجهيز جدول KPI بصيغة نهائية
            kpi_name = "On Time Receiving"
            df_kpi = pd.DataFrame({"KPI": [kpi_name]})

            for _, row in merged.iterrows():
                month = row["Month"]
                hits = int(row["Hits"])
                df_kpi[month] = hits

            # ✅ إضافة صف جديد Total
            total_row = {"KPI": "Total"}
            for col in df_kpi.columns[1:]:  # تجاهل عمود KPI
                total_row[col] = df_kpi[col].sum()
            df_kpi = pd.concat([df_kpi, pd.DataFrame([total_row])], ignore_index=True)

            # ✅ إضافة عمود جديد "2025" يمثل مجموع كل الشهور
            df_kpi["2025"] = df_kpi.iloc[:, 1:].sum(axis=1)

            # ✅ إضافة صف جديد لنسبة التارجت
            target_row = {"KPI": "Target (%)"}
            for _, row in merged.iterrows():
                month = row["Month"]
                target_row[month] = row["Target_%"]
            target_row["2025"] = round(merged["Target_%"].mean(), 2)
            df_kpi = pd.concat([df_kpi, pd.DataFrame([target_row])], ignore_index=True)

            print("✅ [DEBUG] جدول KPI النهائي بعد الإضافات:")
            print(df_kpi.to_string(index=False))

            if selected_months_norm:
                desired_cols = ["KPI"] + [
                    m for m in selected_months_norm if m in df_kpi.columns
                ]
                if "2025" in df_kpi.columns:
                    desired_cols.append("2025")
                df_kpi = df_kpi[[col for col in desired_cols if col in df_kpi.columns]]
            elif selected_month_norm:
                keep_cols = ["KPI", selected_month_norm]
                if "2025" in df_kpi.columns:
                    keep_cols.append("2025")
                df_kpi = df_kpi[[col for col in keep_cols if col in df_kpi.columns]]

            # 🧾 تحويل الجدول إلى HTML
            html_table = df_kpi.to_html(
                classes="table table-bordered text-center table-striped", index=False
            )

            # 🔹 الإرجاع لعرض الجدول في الواجهة
            return {
                "detail_html": html_table,
                "chart_data": df_kpi.to_dict(orient="records"),
            }

        except Exception as e:
            print(f"❌ [EXCEPTION] خطأ أثناء تنفيذ الدالة: {e}")
            return {"error": str(e)}

    def filter_total_lead_time_detail(self, request, selected_month=None):
        excel_path = request.session.get("uploaded_excel_path")
        if not excel_path or not os.path.exists(excel_path):
            return {
                "detail_html": "<p class='text-danger'>⚠️ Excel file not found.</p>",
                "count": 0,
            }

        try:
            # قراءة الشيت المطلوب
            xls = pd.ExcelFile(excel_path, engine="openpyxl")
            sheet_name = next(
                (
                    s
                    for s in xls.sheet_names
                    if "total lead time preformance" in s.lower()
                ),
                None,
            )
            if not sheet_name:
                return {
                    "detail_html": "<p class='text-danger'>❌ Tab 'Total lead time preformance' does not exist in the file.</p>",
                    "count": 0,
                }

            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
            df = pd.read_excel(
                excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
            )
            df.columns = df.columns.str.strip().str.lower()

            # التحقق من الأعمدة المطلوبة
            required_cols = [
                "month",
                "outbound delivery",
                "kpi",
                "reason group",
                "miss reason",
            ]
            if not all(col in df.columns for col in required_cols):
                html = render_to_string(
                    "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
                    {
                        "tabs": [
                            {
                                "name": sheet_name,
                                "columns": df.columns.tolist(),
                                "data": df.head(50).to_dict(orient="records"),
                            }
                        ]
                    },
                )
                return {"detail_html": html, "count": len(df)}

            # تحويل التاريخ إلى شهر
            df["month"] = (
                _excel_dates_to_datetime(df["month"], errors="coerce")
                .dt.strftime("%b")
                .str.capitalize()
            )

            # استخراج الشهور الموجودة فعليًا
            existing_months = sorted(
                df["month"].dropna().unique().tolist(),
                key=lambda x: _excel_dates_to_datetime(x, format="%b").month,
            )
            if not existing_months:
                return {
                    "detail_html": "<p class='text-danger'>⚠️ No valid months were found in the file.</p>",
                    "count": 0,
                }

            # تنظيف النصوص
            df["reason group"] = df["reason group"].astype(str).str.strip().str.lower()
            df["kpi"] = df["kpi"].astype(str).str.strip().str.lower()
            df["miss reason"] = (
                df["miss reason"]
                .astype(str)
                .str.strip()
                .str.replace(r"\s+", " ", regex=True)
            )

            # بيانات Miss الخاصة بـ 3PL فقط
            df_miss_3pl = df[
                (df["kpi"] == "miss") & (df["reason group"] == "3pl")
            ].copy()
            df_miss_3pl["_reason_key"] = df_miss_3pl["miss reason"].str.lower()

            # بيانات Hit (On Time Delivery)
            df_hit = df[df["kpi"] != "miss"].copy()

            # تجميع Miss حسب السبب والشهر
            miss_grouped = df_miss_3pl.groupby(
                ["_reason_key", "month"], as_index=False
            ).agg({"miss reason": "first"})
            miss_grouped["count"] = (
                df_miss_3pl.groupby(["_reason_key", "month"]).size().values
            )

            miss_pivot = miss_grouped.pivot_table(
                index="miss reason", columns="month", values="count", fill_value=0
            )

            # إضافة أعمدة الشهور الناقصة
            for m in existing_months:
                if m not in miss_pivot.columns:
                    miss_pivot[m] = 0
            miss_pivot = miss_pivot[existing_months]

            # حساب On Time Delivery
            hit_counts = (
                df_hit.groupby("month").size().reindex(existing_months, fill_value=0)
            )

            # بناء الجدول النهائي
            final_df = miss_pivot.copy()
            final_df.loc["On Time Delivery"] = hit_counts
            final_df = final_df.fillna(0)

            # تحويل كل القيم لأعداد صحيحة
            final_df = final_df.astype(int)

            # إضافة عمود الإجمالي (2025 بدل TOTAL)
            final_df["2025"] = final_df.sum(axis=1)

            # صف الإجمالي النهائي
            total_row = final_df.sum(numeric_only=True)
            total_row.name = "TOTAL"
            final_df = pd.concat([final_df, pd.DataFrame([total_row])])

            # ترتيب الأعمدة
            final_df.reset_index(inplace=True)
            # final_df.rename(columns={"miss reason": "KPI"}, inplace=True)
            final_df.rename(columns={"index": "KPI"}, inplace=True)

            # ✅ تجهيز البيانات للتمبلت الديناميكي
            tab_data = {
                "name": "KPI Summary - 3PL Performance",
                "sub_tables": [
                    {
                        "title": "Reason From 3PL Side",
                        "columns": ["KPI"] + existing_months + ["2025"],
                        "data": final_df.to_dict(orient="records"),
                    }
                ],
            }

            month_norm = self.apply_month_filter_to_tab(tab_data, selected_month)
            html = render_to_string(
                "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
                {"tab": tab_data, "selected_month": month_norm},
            )

            return {"detail_html": html, "count": len(df), "tab_data": tab_data}

        except Exception as e:
            import traceback

            print(traceback.format_exc())
            return {
                "detail_html": f"<p class='text-danger'>⚠️ Error while reading data: {e}</p>",
                "count": 0,
            }

    def filter_total_lead_time_roche(self, request, selected_month=None):
        """
        🔹 قراءة شيت "Total lead time preformance -R" من التمبلت المرفوع
        🔹 استخراج أسباب التأخير وترتيبها حسب الشهور
        🔹 عرضها بتصميم الجدول الموحد
        """
        print("🟢 [DEBUG] ✅ دخل على filter_total_lead_time_roche()")

        excel_path = request.session.get("uploaded_excel_path")
        if not excel_path or not os.path.exists(excel_path):
            return {"error": "⚠️ Excel file not found."}

        try:
            # فتح ملف الإكسل
            xls = pd.ExcelFile(excel_path, engine="openpyxl")

            # 🔍 البحث عن الشيت المطلوب
            sheet_name = next(
                (s for s in xls.sheet_names if "preformance -r" in s.lower()), None
            )
            if not sheet_name:
                return {
                    "error": "⚠️ No sheet containing 'Total lead time preformance -R' was found."
                }

            # قراءة الشيت (حد صفوف لتسريع التحميل)
            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
            df = pd.read_excel(
                excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
            )
            df.columns = df.columns.str.strip()

            # التحقق من وجود الأعمدة المطلوبة
            if "Month" not in df.columns:
                return {"error": "⚠️ Column named 'Month' was not found in the sheet."}

            # ترتيب الشهور بالترتيب الزمني
            month_order = [
                "Jan",
                "Feb",
                "Mar",
                "Apr",
                "May",
                "Jun",
                "Jul",
                "Aug",
                "Sep",
                "Oct",
                "Nov",
                "Dec",
            ]
            df["Month"] = pd.Categorical(
                df["Month"], categories=month_order, ordered=True
            )
            df = df.sort_values("Month")

            # تحويل البيانات إلى شكل طويل (Melt)
            df_melted = df.melt(id_vars=["Month"], var_name="KPI", value_name="Count")

            # تجميع البيانات حسب السبب والشهر
            pivot_df = (
                df_melted.groupby(["KPI", "Month"])["Count"]
                .sum()
                .unstack(fill_value=0)
                .reindex(columns=month_order, fill_value=0)
            )

            # إضافة عمود الإجمالي السنوي
            pivot_df["2025"] = pivot_df.sum(axis=1)

            # صف الإجمالي الكلي
            total_row = pivot_df.sum(numeric_only=True)
            total_row.name = "TOTAL"
            pivot_df = pd.concat([pivot_df, pd.DataFrame([total_row])])

            # ✅ إعادة تسمية العمود الأول إلى KPI
            pivot_df.reset_index(inplace=True)
            pivot_df.rename(columns={"index": "KPI"}, inplace=True)

            # حذف الشهور الفارغة تمامًا (بدون بيانات)
            pivot_df = pivot_df.loc[:, (pivot_df != 0).any(axis=0)]

            # ✅ تجهيز بيانات الجدول لتمبلت الـ HTML
            tab = {
                "name": "Total Lead Time Performance - Roche Side",
                "columns": list(pivot_df.columns),
                "data": pivot_df.to_dict(orient="records"),
            }

            month_norm = self.apply_month_filter_to_tab(tab, selected_month)
            html = render_to_string(
                "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
                {
                    "tab": tab,
                    "table_title": "Roche Lead Time 2025",
                    "selected_month": month_norm,
                },
            )

            return {
                "detail_html": html,
                "message": "✅ تم عرض بيانات Roche Lead Time بنجاح.",
            }

        except Exception as e:
            import traceback

            print(traceback.format_exc())
            return {"error": f"⚠️ Error while reading Roche Lead Time data: {e}"}

    def filter_outbound(self, request, selected_month=None):
        """
        🔹 عرض تاب Outbound بخطوات أفقية من تمبلت خارجي
        """
        try:
            # ✅ الخطوات مع ألوان وخلفيات مختلفة
            raw_steps = [
                {
                    "title": "GI Issue<br>Pick & Pack",
                    "icon": "bi-receipt",
                    "bg": "#9fc0e4",
                    "text_color": "#fff",
                    "border": "4px solid #9fc0e4",
                    "sub_color": "#eee",
                },
                {
                    "title": "Prepare Docs<br>Invoice, PO and Market place",
                    "icon": "bi-box-seam",
                    "bg": "#e8f1fb",
                    "text_color": "#007fa3",
                    "border": "4px solid #9fc0e4",
                    "sub_color": "#000",
                },
                {
                    "title": "Dispatch Time<br>from Docs Ready till left from WH",
                    "icon": "bi-arrow-left-right",
                    "bg": "#9fc0e4",
                    "text_color": "#fff",
                    "border": "4px solid #9fc0e4",
                    "sub_color": "#eee",
                },
                {
                    "title": "Delivery<br>Deliver to Customer",
                    "icon": "bi-file-earmark-text",
                    "bg": "#e8f1fb",
                    "text_color": "#007fa3",
                    "border": "4px solid #9fc0e4",
                    "sub_color": "#000",
                },
            ]

            steps = []
            for step in raw_steps:
                # نقسم النص على <br>
                parts = step["title"].split("<br>")
                styled_title = ""
                for i, part in enumerate(parts):
                    # لو دا السطر الأخير → نستخدم sub_color
                    color = (
                        step["sub_color"] if i == len(parts) - 1 else step["text_color"]
                    )
                    styled_title += f"<span class='step-line d-block' style='color:{color};'>{part.strip()}</span>"

                steps.append(
                    {
                        "title": styled_title,
                        "icon": step["icon"],
                        "bg": step["bg"],
                        "text_color": step["text_color"],
                        "border": step["border"],
                    }
                )

            # ✅ تمرير البيانات إلى التمبلت
            html = render_to_string(
                "forms-table/table/bootstrap-table/basic-table/components/workflow.html",
                {
                    "table_title": "Outbound workflow",
                    "table_text": "Process Stages",
                    "table_span": "Way Of Calculation",
                    "table_text_bottom": "The KPI was calculated based full lead time Order creation to deliver the order to the customer Based on SLA for each city",
                    "process_steps_text": "=NETWORKDAYS(Order Date, Delivery Date,7)-1",
                    "steps": steps,
                    "workflow_type": "outbound",
                },
            )

            return {
                "detail_html": html,
                "message": "✅ Outbound steps displayed successfully.",
            }

        except Exception as e:
            import traceback

            print(traceback.format_exc())
            return {"error": f"⚠️ Error while rendering the Outbound tab: {e}"}

    def filter_outbound_shipments(
        self, request, selected_month=None, selected_months=None
    ):
        """
        🔹 B2B Outbound: يقرأ من شيت B2B_Outbound فقط.
        🔹 جدول B2B: Channel = B2B أو Trade (نفس دلو الـ B2B في الإكسل)، ORDER STATUS ≠ Cancelled، Creation → Actual Delivery ≤48h = Hit.
        🔹 جدول BTQ: Channel = BTQ، ORDER STATUS ≠ Cancelled، نفس 48h.
        🔹 الشارت: عمودين (B2B و BTQ) مع تسمية "الشهر — اسم الجدول".
        - يُحمّل شيت B2B_Outbound بنفس حد الصفوف الكامل مثل B2C (force_full) حتى تظهر كل الشهور
          وليس أول 200 صف فقط (كانت تُظهر شهرين أو أقل في المعاينة).
        - عمود Month للتجميع: شهر الإنشاء؛ لو التاريخ ناقص نستخدم شهر Actual Delivery حتى لا يختفي شهر (مثل يناير).
        """
        try:
            import os

            excel_path = self.resolve_excel_file_path(request)
            if not excel_path:
                return {
                    "detail_html": "<p class='text-danger'>⚠️ Excel file not found.</p>",
                    "sub_tables": [],
                    "chart_data": [],
                    "stats": {},
                }

            xls = pd.ExcelFile(excel_path, engine="openpyxl")
            sheet_name = "B2B_Outbound"
            if sheet_name not in xls.sheet_names:
                b2b_sheet = next(
                    (s for s in xls.sheet_names if s.strip().replace(" ", "_") == sheet_name or s.strip().lower() == sheet_name.lower()),
                    None,
                )
                if b2b_sheet:
                    sheet_name = b2b_sheet
                else:
                    return {
                        "detail_html": f"<p class='text-warning'>⚠️ Sheet '{sheet_name}' not found.</p>",
                        "sub_tables": [],
                        "chart_data": [],
                        "stats": {},
                    }

            _b2b_max_rows = _excel_max_rows_for_request(request, force_full=True)
            _nr_kw = _read_excel_nrows_kw(_b2b_max_rows)
            df = pd.read_excel(
                excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
            )
            df.columns = df.columns.astype(str).str.strip()

            def find_col(d, candidates):
                for c in d.columns:
                    if str(c).strip().lower() in [x.lower() for x in candidates]:
                        return c
                for cand in candidates:
                    for c in d.columns:
                        if cand.lower() in str(c).lower():
                            return c
                return None

            so_col = find_col(df, ["SO", "Order Number", "Order Nbr", "Order No"])
            channel_col = find_col(df, ["Channel", "channel"])
            order_status_col = find_col(df, ["ORDER STATUS", "Order Status", "OrderStatus", "Status"])
            creation_col = find_col(df, ["Creation Date & Time", "Creation Date and Time", "Create Timestamp", "Creation DateTime", "Create Date"])
            actual_delivery_col = find_col(df, ["Actual Delivery Date", "Actual Delivery", "Delivery Date"])
            pod_date_col = find_col(df, ["POD Date", "PODDate", "POD date"])

            if not so_col or not channel_col or not order_status_col or not creation_col or not actual_delivery_col:
                missing = [x for x, c in [("SO", so_col), ("Channel", channel_col), ("ORDER STATUS", order_status_col), ("Creation Date & Time", creation_col), ("Actual Delivery Date", actual_delivery_col)] if not c]
                return {"detail_html": f"<p class='text-danger'>⚠️ B2B_Outbound: missing columns: {', '.join(missing)}.</p>", "sub_tables": [], "chart_data": [], "stats": {}}

            df = df.rename(columns={so_col: "SO", channel_col: "Channel", order_status_col: "ORDER STATUS", creation_col: "Creation Date & Time", actual_delivery_col: "Actual Delivery Date"})
            if pod_date_col:
                df = df.rename(columns={pod_date_col: "POD Date"})
            else:
                df["POD Date"] = pd.NaT

            df["Channel"] = df["Channel"].astype(str).str.strip()
            df["ORDER STATUS"] = df["ORDER STATUS"].astype(str).str.strip()
            df = df[~df["ORDER STATUS"].str.upper().str.contains("CANCELLED", na=False)]
            df["Creation Date & Time"] = _excel_dates_to_datetime(df["Creation Date & Time"], errors="coerce")
            df["Actual Delivery Date"] = _excel_dates_to_datetime(df["Actual Delivery Date"], errors="coerce")
            df["POD Date"] = _excel_dates_to_datetime(df["POD Date"], errors="coerce")
            # Month for KPIs / filters / charts = creation month (matches dashboard month filter & one-month files)
            month_order = [
                "Jan",
                "Feb",
                "Mar",
                "Apr",
                "May",
                "Jun",
                "Jul",
                "Aug",
                "Sep",
                "Oct",
                "Nov",
                "Dec",
            ]
            # English abbr + creation-first; if creation is missing/invalid, fall back to Actual Delivery
            # so months that exist on the sheet (e.g. Jan) are not dropped from KPIs/charts.
            _m_cre = _dt_series_to_month_abbr_en(df["Creation Date & Time"])
            _m_del = _dt_series_to_month_abbr_en(df["Actual Delivery Date"])
            df["Month"] = _m_cre.where(_m_cre.notna(), _m_del)
            # Keep only real calendar months (drops NaT / bad parses so pivots don't stretch to empty months)
            df = df[df["Month"].isin(month_order)]
            month_order_value = {m: i for i, m in enumerate(month_order)}

            def _compute_48h(df_part):
                hours = (df_part["Actual Delivery Date"] - df_part["Creation Date & Time"]).dt.total_seconds() / 3600.0
                is_hit = (hours <= 48) & hours.notna()
                return df_part.assign(Hours_48=hours, is_hit=is_hit)

            _chu = df["Channel"].str.upper()
            # Trade (مثل Amazon / AFAQ في الشيت) يُعامل مع B2B في الـ KPI والجدول الخام
            df_b2b = df[_chu.isin(["B2B", "TRADE"])].copy()
            df_btq = df[_chu == "BTQ"].copy()
            df_b2b = _compute_48h(df_b2b)
            df_btq = _compute_48h(df_btq)

            selected_months_norm = []
            if selected_months:
                if isinstance(selected_months, str):
                    selected_months = [selected_months]
                for m in selected_months:
                    norm = self.normalize_month_label(m)
                    if norm and norm not in selected_months_norm:
                        selected_months_norm.append(norm)
            selected_month_norm = self.normalize_month_label(selected_month) if selected_month and not selected_months_norm else None
            if selected_months_norm:
                df_b2b = df_b2b[df_b2b["Month"].fillna("").str.lower().isin([m.lower() for m in selected_months_norm])]
                df_btq = df_btq[df_btq["Month"].fillna("").str.lower().isin([m.lower() for m in selected_months_norm])]
            elif selected_month_norm:
                df_b2b = df_b2b[df_b2b["Month"].fillna("").str.lower() == selected_month_norm.lower()]
                df_btq = df_btq[df_btq["Month"].fillna("").str.lower() == selected_month_norm.lower()]

            def _fmt_dt(x):
                if pd.isna(x) or x is pd.NaT:
                    return ""
                try:
                    return pd.Timestamp(x).strftime("%Y-%m-%d %H:%M")
                except Exception:
                    return ""

            def _to_blank(val):
                if val is None:
                    return ""
                if isinstance(val, float) and (pd.isna(val) or (val != val)):
                    return ""
                s = str(val).strip()
                if s.lower() in ("nan", "nat", "none", "<nat>"):
                    return ""
                return s

            sub_tables = []
            chart_data = []

            # ——— جدول B2B KPI (48h) ———
            if not df_b2b.empty:
                b2b_summary = df_b2b.groupby("Month").agg(Total_Shipments=("SO", "nunique"), Hits=("is_hit", "sum")).reset_index()
                b2b_summary["Misses"] = b2b_summary["Total_Shipments"] - b2b_summary["Hits"]
                b2b_summary["Hit %"] = (b2b_summary["Hits"] / b2b_summary["Total_Shipments"].replace(0, np.nan) * 100).fillna(0).round(2)
                b2b_summary = b2b_summary.sort_values(by="Month", key=lambda c: c.map(month_order_value))
                ordered_b2b = b2b_summary["Month"].tolist()
                agg_col_b2b = None
                if len(ordered_b2b) >= 2 and not df_b2b.empty:
                    try:
                        agg_col_b2b = str(
                            int(df_b2b["Creation Date & Time"].dropna().dt.year.max())
                        )
                    except Exception:
                        agg_col_b2b = None
                pivot_b2b = ["KPI"] + ordered_b2b + ([agg_col_b2b] if agg_col_b2b else [])
                hit_pct_b2b = {"KPI": "Hit %"}
                total_b2b = {"KPI": "Total Shipments"}
                hit_b2b = {"KPI": "Hit (≤48h)"}
                miss_b2b = {"KPI": "Miss (>48h)"}
                for m in ordered_b2b:
                    r = b2b_summary[b2b_summary["Month"] == m].iloc[0]
                    t, h = int(r["Total_Shipments"]), int(r["Hits"])
                    total_b2b[m], hit_b2b[m], miss_b2b[m] = t, h, int(r["Misses"])
                    hit_pct_b2b[m] = int(round(h * 100 / t)) if t > 0 else 0
                if agg_col_b2b and agg_col_b2b in pivot_b2b:
                    t2025 = int(b2b_summary["Total_Shipments"].sum())
                    h2025 = int(b2b_summary["Hits"].sum())
                    total_b2b[agg_col_b2b], hit_b2b[agg_col_b2b], miss_b2b[agg_col_b2b] = (
                        t2025,
                        h2025,
                        t2025 - h2025,
                    )
                    hit_pct_b2b[agg_col_b2b] = (
                        int(round(h2025 * 100 / t2025)) if t2025 > 0 else 0
                    )
                sub_tables.append({"id": "sub-table-b2b-hit-summary", "title": "B2B & Trade KPI (Creation → Delivery ≤ 48h)", "columns": pivot_b2b, "data": [hit_pct_b2b, hit_b2b, miss_b2b, total_b2b], "chart_data": [], "full_width": False, "side_by_side": True})
                chart_data.append({
                    "type": "column",
                    "name": "B2B & Trade Hit % (≤48h)",
                    "color": "#9F8170",
                    "related_table": "sub-table-b2b-hit-summary",
                    "dataPoints": [{"label": f"{m} — B2B & Trade", "y": hit_pct_b2b.get(m, 0)} for m in ordered_b2b],
                })

            # ——— جدول BTQ KPI (48h) ———
            if not df_btq.empty:
                btq_summary = df_btq.groupby("Month").agg(Total_Shipments=("SO", "nunique"), Hits=("is_hit", "sum")).reset_index()
                btq_summary["Misses"] = btq_summary["Total_Shipments"] - btq_summary["Hits"]
                btq_summary["Hit %"] = (btq_summary["Hits"] / btq_summary["Total_Shipments"].replace(0, np.nan) * 100).fillna(0).round(2)
                btq_summary = btq_summary.sort_values(by="Month", key=lambda c: c.map(month_order_value))
                ordered_btq = btq_summary["Month"].tolist()
                agg_col_btq = None
                if len(ordered_btq) >= 2 and not df_btq.empty:
                    try:
                        agg_col_btq = str(
                            int(df_btq["Creation Date & Time"].dropna().dt.year.max())
                        )
                    except Exception:
                        agg_col_btq = None
                pivot_btq = ["KPI"] + ordered_btq + ([agg_col_btq] if agg_col_btq else [])
                hit_pct_btq = {"KPI": "Hit %"}
                total_btq = {"KPI": "Total Shipments"}
                hit_btq = {"KPI": "Hit (≤48h)"}
                miss_btq = {"KPI": "Miss (>48h)"}
                for m in ordered_btq:
                    r = btq_summary[btq_summary["Month"] == m].iloc[0]
                    t, h = int(r["Total_Shipments"]), int(r["Hits"])
                    total_btq[m], hit_btq[m], miss_btq[m] = t, h, int(r["Misses"])
                    hit_pct_btq[m] = int(round(h * 100 / t)) if t > 0 else 0
                if agg_col_btq and agg_col_btq in pivot_btq:
                    t2025 = int(btq_summary["Total_Shipments"].sum())
                    h2025 = int(btq_summary["Hits"].sum())
                    total_btq[agg_col_btq], hit_btq[agg_col_btq], miss_btq[agg_col_btq] = (
                        t2025,
                        h2025,
                        t2025 - h2025,
                    )
                    hit_pct_btq[agg_col_btq] = (
                        int(round(h2025 * 100 / t2025)) if t2025 > 0 else 0
                    )
                sub_tables.append({"id": "sub-table-btq-hit-summary", "title": "BTQ KPI (Creation → Delivery ≤ 48h)", "columns": pivot_btq, "data": [hit_pct_btq, hit_btq, miss_btq, total_btq], "chart_data": [], "full_width": False, "side_by_side": True})
                chart_data.append({
                    "type": "column",
                    "name": "BTQ Hit % (≤48h)",
                    "color": "#81613E",
                    "related_table": "sub-table-btq-hit-summary",
                    "dataPoints": [{"label": f"{m} — BTQ", "y": hit_pct_btq.get(m, 0)} for m in ordered_btq],
                })

            # ——— PODs B2B و PODs BTQ: Actual Delivery → POD Date ≤18 يوم = Hit ———
            chart_data_pods = []
            df_pods_b2b = df[_chu.isin(["B2B", "TRADE"])].copy()
            df_pods_b2b = df_pods_b2b[df_pods_b2b["POD Date"].notna()]
            df_pods_btq = df[df["Channel"].str.upper() == "BTQ"].copy()
            df_pods_btq = df_pods_btq[df_pods_btq["POD Date"].notna()]

            if not df_pods_b2b.empty:
                days_pod_b2b = (df_pods_b2b["POD Date"] - df_pods_b2b["Actual Delivery Date"]).dt.total_seconds() / (24 * 3600.0)
                df_pods_b2b["PODs_is_hit"] = (days_pod_b2b <= 18) & days_pod_b2b.notna()
            if not df_pods_btq.empty:
                days_pod_btq = (df_pods_btq["POD Date"] - df_pods_btq["Actual Delivery Date"]).dt.total_seconds() / (24 * 3600.0)
                df_pods_btq["PODs_is_hit"] = (days_pod_btq <= 18) & days_pod_btq.notna()

            if selected_months_norm:
                if not df_pods_b2b.empty:
                    df_pods_b2b = df_pods_b2b[df_pods_b2b["Month"].fillna("").str.lower().isin([m.lower() for m in selected_months_norm])]
                if not df_pods_btq.empty:
                    df_pods_btq = df_pods_btq[df_pods_btq["Month"].fillna("").str.lower().isin([m.lower() for m in selected_months_norm])]
            elif selected_month_norm:
                if not df_pods_b2b.empty:
                    df_pods_b2b = df_pods_b2b[df_pods_b2b["Month"].fillna("").str.lower() == selected_month_norm.lower()]
                if not df_pods_btq.empty:
                    df_pods_btq = df_pods_btq[df_pods_btq["Month"].fillna("").str.lower() == selected_month_norm.lower()]

            if not df_pods_b2b.empty:
                pods_b2b_summary = df_pods_b2b.groupby("Month").agg(Total_Shipments=("SO", "nunique"), Hits=("PODs_is_hit", "sum")).reset_index()
                pods_b2b_summary["Misses"] = pods_b2b_summary["Total_Shipments"] - pods_b2b_summary["Hits"]
                pods_b2b_summary["Hit %"] = (pods_b2b_summary["Hits"] / pods_b2b_summary["Total_Shipments"].replace(0, np.nan) * 100).fillna(0).round(2)
                pods_b2b_summary = pods_b2b_summary.sort_values(by="Month", key=lambda c: c.map(month_order_value))
                ordered_pb2b = pods_b2b_summary["Month"].tolist()
                agg_col_pb2b = None
                if len(ordered_pb2b) >= 2 and not df_pods_b2b.empty:
                    try:
                        agg_col_pb2b = str(
                            int(
                                df_pods_b2b["Creation Date & Time"]
                                .dropna()
                                .dt.year.max()
                            )
                        )
                    except Exception:
                        agg_col_pb2b = None
                pivot_pb2b = ["KPI"] + ordered_pb2b + ([agg_col_pb2b] if agg_col_pb2b else [])
                hit_pct_pb2b = {"KPI": "Hit %"}
                total_pb2b = {"KPI": "Total Shipments"}
                hit_pb2b = {"KPI": "Hit (≤18d)"}
                miss_pb2b = {"KPI": "Miss (>18d)"}
                for m in ordered_pb2b:
                    r = pods_b2b_summary[pods_b2b_summary["Month"] == m].iloc[0]
                    t, h = int(r["Total_Shipments"]), int(r["Hits"])
                    total_pb2b[m], hit_pb2b[m], miss_pb2b[m] = t, h, int(r["Misses"])
                    hit_pct_pb2b[m] = int(round(h * 100 / t)) if t > 0 else 0
                if agg_col_pb2b and agg_col_pb2b in pivot_pb2b:
                    t2025 = int(pods_b2b_summary["Total_Shipments"].sum())
                    h2025 = int(pods_b2b_summary["Hits"].sum())
                    total_pb2b[agg_col_pb2b], hit_pb2b[agg_col_pb2b], miss_pb2b[agg_col_pb2b] = (
                        t2025,
                        h2025,
                        t2025 - h2025,
                    )
                    hit_pct_pb2b[agg_col_pb2b] = (
                        int(round(h2025 * 100 / t2025)) if t2025 > 0 else 0
                    )
                sub_tables.append({"id": "sub-table-pods-b2b-hit-summary", "title": "PODs B2B & Trade KPI (Delivery → POD ≤ 18 days)", "columns": pivot_pb2b, "data": [hit_pct_pb2b, hit_pb2b, miss_pb2b, total_pb2b], "chart_data": [], "full_width": False, "side_by_side": True})
                chart_data_pods.append({"type": "column", "name": "PODs B2B & Trade Hit % (≤18d)", "color": "#9F8170", "related_table": "sub-table-pods-b2b-hit-summary", "dataPoints": [{"label": f"{m} — PODs B2B & Trade", "y": hit_pct_pb2b.get(m, 0)} for m in ordered_pb2b]})

            if not df_pods_btq.empty:
                pods_btq_summary = df_pods_btq.groupby("Month").agg(Total_Shipments=("SO", "nunique"), Hits=("PODs_is_hit", "sum")).reset_index()
                pods_btq_summary["Misses"] = pods_btq_summary["Total_Shipments"] - pods_btq_summary["Hits"]
                pods_btq_summary["Hit %"] = (pods_btq_summary["Hits"] / pods_btq_summary["Total_Shipments"].replace(0, np.nan) * 100).fillna(0).round(2)
                pods_btq_summary = pods_btq_summary.sort_values(by="Month", key=lambda c: c.map(month_order_value))
                ordered_pbtq = pods_btq_summary["Month"].tolist()
                agg_col_pbtq = None
                if len(ordered_pbtq) >= 2 and not df_pods_btq.empty:
                    try:
                        agg_col_pbtq = str(
                            int(
                                df_pods_btq["Creation Date & Time"]
                                .dropna()
                                .dt.year.max()
                            )
                        )
                    except Exception:
                        agg_col_pbtq = None
                pivot_pbtq = ["KPI"] + ordered_pbtq + ([agg_col_pbtq] if agg_col_pbtq else [])
                hit_pct_pbtq = {"KPI": "Hit %"}
                total_pbtq = {"KPI": "Total Shipments"}
                hit_pbtq = {"KPI": "Hit (≤18d)"}
                miss_pbtq = {"KPI": "Miss (>18d)"}
                for m in ordered_pbtq:
                    r = pods_btq_summary[pods_btq_summary["Month"] == m].iloc[0]
                    t, h = int(r["Total_Shipments"]), int(r["Hits"])
                    total_pbtq[m], hit_pbtq[m], miss_pbtq[m] = t, h, int(r["Misses"])
                    hit_pct_pbtq[m] = int(round(h * 100 / t)) if t > 0 else 0
                if agg_col_pbtq and agg_col_pbtq in pivot_pbtq:
                    t2025 = int(pods_btq_summary["Total_Shipments"].sum())
                    h2025 = int(pods_btq_summary["Hits"].sum())
                    total_pbtq[agg_col_pbtq], hit_pbtq[agg_col_pbtq], miss_pbtq[agg_col_pbtq] = (
                        t2025,
                        h2025,
                        t2025 - h2025,
                    )
                    hit_pct_pbtq[agg_col_pbtq] = (
                        int(round(h2025 * 100 / t2025)) if t2025 > 0 else 0
                    )
                sub_tables.append({"id": "sub-table-pods-btq-hit-summary", "title": "PODs BTQ KPI (Delivery → POD ≤ 18 days)", "columns": pivot_pbtq, "data": [hit_pct_pbtq, hit_pbtq, miss_pbtq, total_pbtq], "chart_data": [], "full_width": False, "side_by_side": True})
                chart_data_pods.append({"type": "column", "name": "PODs BTQ Hit % (≤18d)", "color": "#81613E", "related_table": "sub-table-pods-btq-hit-summary", "dataPoints": [{"label": f"{m} — PODs BTQ", "y": hit_pct_pbtq.get(m, 0)} for m in ordered_pbtq]})

            if not sub_tables:
                return {"detail_html": "<p class='text-warning'>⚠️ No B2B or BTQ records for the selected period.</p>", "sub_tables": [], "chart_data": [], "chart_data_pods": [], "stats": {}}

            df_all = pd.concat([df_b2b, df_btq], ignore_index=True) if not df_b2b.empty and not df_btq.empty else (df_b2b if not df_b2b.empty else df_btq)
            overall_total = int(df_all["SO"].nunique())
            overall_hits = int(df_all["is_hit"].sum()) if "is_hit" in df_all.columns else (int(df_b2b["is_hit"].sum()) if not df_b2b.empty else int(df_btq["is_hit"].sum()))
            overall_hit_pct = round((overall_hits / overall_total) * 100, 2) if overall_total else 0

            raw_sheet_cols = ["SO", "Channel", "ORDER STATUS", "Creation Date & Time", "Actual Delivery Date", "POD Date", "Month"]
            raw_sheet_cols = [c for c in raw_sheet_cols if c in df_all.columns]
            _seen_b2b_m = {
                str(m).strip()
                for m in df_all["Month"].dropna()
                if str(m).strip() in month_order
            }
            b2b_raw_month_filter_options = [m for m in month_order if m in _seen_b2b_m]
            raw_df = df_all[raw_sheet_cols].copy().sort_values(
                "Actual Delivery Date", ascending=False
            ).head(500)
            raw_df["Creation Date & Time"] = raw_df["Creation Date & Time"].apply(_fmt_dt)
            raw_df["Actual Delivery Date"] = raw_df["Actual Delivery Date"].apply(_fmt_dt)
            if "POD Date" in raw_df.columns:
                raw_df["POD Date"] = raw_df["POD Date"].apply(_fmt_dt)
            raw_excel_rows = []
            for row in raw_df.to_dict(orient="records"):
                rec = {k: _to_blank(row.get(k)) for k in raw_sheet_cols}
                _fm = str(row.get("Month") or "").strip()
                rec["_filter_month"] = _fm if _fm in month_order else ""
                raw_excel_rows.append(rec)
            raw_excel_table = {
                "id": "sub-table-b2b-raw-sheet",
                "title": "B2B_Outbound (Sheet Data)",
                "columns": [{"name": c, "key": c, "group": "sheet"} for c in raw_sheet_cols],
                "data": raw_excel_rows,
                "full_width": True,
                "month_filter_options": b2b_raw_month_filter_options,
            }

            return {
                "detail_html": "",
                "sub_tables": sub_tables,
                "chart_data": chart_data,
                "chart_data_pods": chart_data_pods,
                "raw_excel_table": raw_excel_table,
                "stats": {"total": overall_total, "hit": overall_hits, "miss": overall_total - overall_hits, "hit_pct": overall_hit_pct, "target": 99},
            }

        except Exception as e:
            import traceback

            print(traceback.format_exc())
            return {
                "detail_html": f"<p class='text-danger'>⚠️ Error processing outbound shipments: {e}</p>",
                "sub_tables": [],
                "chart_data": [],
                "chart_data_pods": [],
                "stats": {},
            }

    def filter_b2c_outbound(
        self, request, selected_month=None, selected_months=None
    ):
        """
        B2C Outbound: يقرأ من شيت B2C_Outbound.
        - يُحمّل الشيت كاملاً (حد EXCEL_FULL_MAX_ROWS في settings، أو بدون حد لو None) حتى تظهر
          كل الشهور/الصفوف في الجداول والشارتات دون الاعتماد على معاينة 200 صف.
        - جدول Pick & Peak (Creation / Picked): حسب مدة (Creation → Picked) بالساعات،
          صفوف Cycle (0-9HRS … 105+HRS)، أعمدة بالشهر (YYYY-MM)، عمود %، Status = Delivered فقط.
        - كروت الـ KPI: نفس منطق Hit/Miss حسب الموعد (Deadline) على كل الصفوف الصالحة.
        """
        try:
            import os
            from datetime import time, datetime, timedelta

            excel_path = self.resolve_excel_file_path(request)
            if not excel_path:
                return {
                    "detail_html": "<p class='text-danger'>⚠️ Excel file not found.</p>",
                    "sub_tables": [],
                    "raw_excel_table": None,
                    "stats": {},
                }

            import hashlib

            _b2c_max_rows = _excel_max_rows_for_request(request, force_full=True)
            _b2c_mode = "full" if _b2c_max_rows is None else str(_b2c_max_rows)
            _b2c_key = (
                "b2c_outbound_"
                + hashlib.md5((excel_path or "").encode()).hexdigest()[:16]
                + "_"
                + _b2c_mode
            )
            _b2c_cached = cache.get(_b2c_key)
            if _b2c_cached is not None:
                return _b2c_cached

            xls = pd.ExcelFile(excel_path, engine="openpyxl")
            sheet_name = "B2C_Outbound"
            if sheet_name not in xls.sheet_names:
                b2c_sheet = next(
                    (
                        s
                        for s in xls.sheet_names
                        if s.strip().replace(" ", "_") == sheet_name
                        or s.strip().lower() == sheet_name.lower()
                    ),
                    None,
                )
                if b2c_sheet:
                    sheet_name = b2c_sheet
                else:
                    return {
                        "detail_html": f"<p class='text-warning'>⚠️ Sheet '{sheet_name}' not found.</p>",
                        "sub_tables": [],
                        "raw_excel_table": None,
                        "stats": {},
                    }

            _nr_kw = _read_excel_nrows_kw(_b2c_max_rows)
            df = pd.read_excel(
                excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
            )
            df.columns = df.columns.astype(str).str.strip()
            df_full = df.copy()

            def find_col(d, candidates):
                for c in d.columns:
                    if str(c).strip().lower() in [x.lower() for x in candidates]:
                        return c
                for cand in candidates:
                    for c in d.columns:
                        if cand.lower() in str(c).lower():
                            return c
                return None

            order_col = find_col(
                df,
                [
                    "ORDER / SO",
                    "ORDER/SO",
                    "Order / SO",
                    "SO",
                    "Order Number",
                    "Order Nbr",
                    "Order No",
                ],
            )
            creation_col = find_col(
                df,
                [
                    "CREATTION DATE",
                    "CREATION DATE",
                    "Creation Date & Time",
                    "Creation Date and Time",
                    "Create Timestamp",
                    "Creation DateTime",
                    "Create Date",
                ],
            )
            picked_col = find_col(
                df,
                [
                    "PICKED DATE",
                    "Picked Date",
                    "Picked Date & Time",
                    "Picked DateTime",
                ],
            )
            status_col = find_col(
                df,
                ["Status", "STATUS", "Order Status", "OrderStatus"],
            )
            dispatch_col = find_col(
                df,
                [
                    "Dispatch date & time",
                    "Dispatch Date & Time",
                    "Dispatch Date and Time",
                    "Dispatch DateTime",
                    "Dispatch Date",
                ],
            )
            delivered_col = find_col(
                df,
                [
                    "DELIVERED DATE",
                    "Delivered Date",
                    "Delivered Date & Time",
                    "Delivered DateTime",
                ],
            )

            if not order_col or not creation_col or not picked_col:
                missing = [
                    x
                    for x, c in [
                        ("ORDER / SO", order_col),
                        ("Creation", creation_col),
                        ("Picked", picked_col),
                    ]
                    if not c
                ]
                return {
                    "detail_html": f"<p class='text-danger'>⚠️ B2C_Outbound: missing columns: {', '.join(missing)}.</p>",
                    "sub_tables": [],
                    "raw_excel_table": None,
                    "stats": {},
                }

            df = df.rename(
                columns={
                    order_col: "Order_SO",
                    creation_col: "Creation_Date",
                    picked_col: "Picked_Date",
                }
            )
            df["Creation_Date"] = _excel_dates_to_datetime(df["Creation_Date"], errors="coerce")
            df["Picked_Date"] = _excel_dates_to_datetime(df["Picked_Date"], errors="coerce")
            df = df.dropna(subset=["Order_SO", "Creation_Date", "Picked_Date"])
            if df.empty:
                return {
                    "detail_html": "<p class='text-warning'>⚠️ No rows with valid Order, Creation and Picked dates.</p>",
                    "sub_tables": [],
                    "raw_excel_table": None,
                    "stats": {},
                }

            # تجميع بالطلب: نأخذ أول صف لكل Order (أو يمكن آخر Picked)
            df = df.sort_values(["Order_SO", "Creation_Date"]).drop_duplicates(
                subset=["Order_SO"], keep="first"
            )

            def get_deadline(creation_dt):
                if pd.isna(creation_dt):
                    return pd.NaT
                try:
                    ts = pd.Timestamp(creation_dt)
                    d = ts.date()
                    t = ts.time()
                    t9 = time(9, 0)
                    t15 = time(15, 0)
                    t16 = time(16, 0)
                    t10 = time(10, 0)
                    if t9 <= t < t15:
                        return pd.Timestamp(datetime.combine(d, t16))
                    elif t >= t15:
                        next_d = d + timedelta(days=1)
                        return pd.Timestamp(datetime.combine(next_d, t10))
                    else:
                        return pd.Timestamp(datetime.combine(d, t10))
                except Exception:
                    return pd.NaT

            df["Deadline"] = df["Creation_Date"].apply(get_deadline)
            df["is_hit"] = (df["Picked_Date"] <= df["Deadline"]) & df["Deadline"].notna()
            df["Duration_Hours"] = (
                (df["Picked_Date"] - df["Creation_Date"]).dt.total_seconds() / 3600.0
            )
            df["Duration_Hours"] = df["Duration_Hours"].apply(
                lambda x: int(round(x)) if pd.notna(x) else pd.NA
            )

            def _fmt_dt(x):
                if pd.isna(x) or x is pd.NaT:
                    return ""
                try:
                    return pd.Timestamp(x).strftime("%Y-%m-%d %I:%M %p")
                except Exception:
                    return ""

            def _to_blank(val):
                if val is None:
                    return ""
                if isinstance(val, float) and (pd.isna(val) or (val != val)):
                    return ""
                s = str(val).strip()
                if s.lower() in ("nan", "nat", "none", "<nat>"):
                    return ""
                return s

            pick_peak_cycles = [
                "0-9HRS / SAME DAY",
                "9-24HRS / 1 DAY",
                "24-33HRS / 1.5 DAY",
                "33-48HRS / 2 DAYS",
                "48-57HRS / 2.5 DAYS",
                "57-72HRS / 3 DAYS",
                "72-81HRS / 3.5 DAYS",
                "81-96HRS / 4 DAYS",
                "96-105HRS / 4.5 DAYS",
                "105+HRS / 5+ DAYS",
            ]

            def _pick_peak_cycle_from_hours(hours):
                if hours is None or (
                    isinstance(hours, float) and (hours != hours)
                ):
                    return None
                try:
                    h = float(hours)
                except (TypeError, ValueError):
                    return None
                if h < 0:
                    return None
                if h < 9:
                    return pick_peak_cycles[0]
                if h < 24:
                    return pick_peak_cycles[1]
                if h < 33:
                    return pick_peak_cycles[2]
                if h < 48:
                    return pick_peak_cycles[3]
                if h < 57:
                    return pick_peak_cycles[4]
                if h < 72:
                    return pick_peak_cycles[5]
                if h < 81:
                    return pick_peak_cycles[6]
                if h < 96:
                    return pick_peak_cycles[7]
                if h < 105:
                    return pick_peak_cycles[8]
                return pick_peak_cycles[9]

            df_pp = df.copy()
            if status_col:
                st_series = (
                    df_full.loc[df_pp.index, status_col]
                    .astype(str)
                    .str.strip()
                    .str.upper()
                )
                df_pp = df_pp[st_series == "DELIVERED"].copy()

            df_pp["Duration_Hours_Float"] = (
                df_pp["Picked_Date"] - df_pp["Creation_Date"]
            ).dt.total_seconds() / 3600.0
            df_pp = df_pp[
                df_pp["Duration_Hours_Float"].notna()
                & (df_pp["Duration_Hours_Float"] >= 0)
            ]
            df_pp["Cycle"] = df_pp["Duration_Hours_Float"].apply(
                _pick_peak_cycle_from_hours
            )
            df_pp = df_pp.dropna(subset=["Cycle"])
            df_pp["Month_YM"] = df_pp["Creation_Date"].dt.strftime("%Y-%m")

            month_cols_pp = sorted(
                m
                for m in df_pp["Month_YM"].dropna().unique().tolist()
                if m and str(m).strip().lower() not in ("nan", "nat")
            )
            if df_pp.empty:
                pivot_pp = pd.DataFrame(
                    index=pick_peak_cycles, columns=month_cols_pp, dtype=int
                ).fillna(0)
            else:
                pivot_pp = (
                    df_pp.groupby(["Cycle", "Month_YM"], observed=True)
                    .size()
                    .unstack(fill_value=0)
                )
                if month_cols_pp:
                    pivot_pp = pivot_pp.reindex(
                        columns=month_cols_pp, fill_value=0
                    )
                pivot_pp = pivot_pp.reindex(
                    index=pick_peak_cycles, fill_value=0
                )

            grand_total_pp = int(len(df_pp))

            pick_peak_count_rows = []
            for cyc in pick_peak_cycles:
                row = {"Cycle": cyc}
                row_sum = 0
                for ym in month_cols_pp:
                    c = int(pivot_pp.loc[cyc, ym])
                    row[ym] = c
                    row_sum += c
                rp = (
                    round(100.0 * row_sum / grand_total_pp, 1)
                    if grand_total_pp
                    else 0.0
                )
                row["%"] = f"{rp}%"
                pick_peak_count_rows.append(row)
            total_row_pp = {"Cycle": "TOTAL"}
            for ym in month_cols_pp:
                total_row_pp[ym] = int(df_pp[df_pp["Month_YM"] == ym].shape[0])
            total_row_pp["%"] = "100%" if grand_total_pp else "0%"
            pick_peak_count_rows.append(total_row_pp)

            pick_peak_columns = ["Cycle"] + month_cols_pp + ["%"]
            pick_peak_summary = {}
            if not df_pp.empty:
                cmin = df_pp["Creation_Date"].min()
                cmax = df_pp["Creation_Date"].max()
                try:
                    pick_peak_summary["date_from"] = pd.Timestamp(
                        cmin
                    ).strftime("%d-%b-%y")
                    pick_peak_summary["date_to"] = pd.Timestamp(
                        cmax
                    ).strftime("%d-%b-%y")
                except Exception:
                    pick_peak_summary["date_from"] = ""
                    pick_peak_summary["date_to"] = ""
                y_del = int(df_pp["Creation_Date"].dt.year.max())
                pick_peak_summary["delivered_label"] = f"Delivered {y_del}"
            else:
                pick_peak_summary = {
                    "date_from": "",
                    "date_to": "",
                    "delivered_label": "Delivered",
                }

            sub_tables = [
                {
                    "id": "sub-table-b2c-pick-peak",
                    "title": "Pick & Peak — Creation / Picked",
                    "columns": pick_peak_columns,
                    "data": pick_peak_count_rows,
                    "pick_peak_summary": pick_peak_summary,
                    "full_width": True,
                    "col_12": True,
                }
            ]

            # ——— جدول Dispatch: Status = Delivered، Creation 9am–3pm → قبل 5:30pm، 3pm–9am → قبل 11:30am ———
            if status_col and dispatch_col:
                df_d = df_full[
                    df_full[status_col].astype(str).str.strip().str.upper()
                    == "DELIVERED"
                ][[order_col, creation_col, dispatch_col]].copy()
                df_d = df_d.rename(
                    columns={
                        order_col: "Order_SO",
                        creation_col: "Creation_Date",
                        dispatch_col: "Dispatch_Date",
                    }
                )
                df_d["Creation_Date"] = _excel_dates_to_datetime(
                    df_d["Creation_Date"], errors="coerce"
                )
                df_d["Dispatch_Date"] = _excel_dates_to_datetime(
                    df_d["Dispatch_Date"], errors="coerce"
                )
                df_d = df_d.dropna(
                    subset=["Order_SO", "Creation_Date", "Dispatch_Date"]
                )
                if not df_d.empty:
                    df_d = (
                        df_d.sort_values(["Order_SO", "Creation_Date"])
                        .drop_duplicates(subset=["Order_SO"], keep="first")
                    )

                    def get_dispatch_deadline(creation_dt):
                        if pd.isna(creation_dt):
                            return pd.NaT
                        try:
                            ts = pd.Timestamp(creation_dt)
                            d = ts.date()
                            t = ts.time()
                            t9 = time(9, 0)
                            t15 = time(15, 0)
                            t530 = time(17, 30)
                            t1130 = time(11, 30)
                            if t9 <= t < t15:
                                return pd.Timestamp(
                                    datetime.combine(d, t530)
                                )
                            elif t >= t15:
                                next_d = d + timedelta(days=1)
                                return pd.Timestamp(
                                    datetime.combine(next_d, t1130)
                                )
                            else:
                                return pd.Timestamp(
                                    datetime.combine(d, t1130)
                                )
                        except Exception:
                            return pd.NaT

                    df_d["Deadline"] = df_d["Creation_Date"].apply(
                        get_dispatch_deadline
                    )
                    df_d["is_hit"] = (
                        (df_d["Dispatch_Date"] <= df_d["Deadline"])
                        & df_d["Deadline"].notna()
                    )
                    df_d["Duration_Hours"] = (
                        (
                            df_d["Dispatch_Date"]
                            - df_d["Creation_Date"]
                        ).dt.total_seconds()
                        / 3600.0
                    )
                    df_d["Duration_Hours"] = df_d["Duration_Hours"].apply(
                        lambda x: int(round(x))
                        if pd.notna(x) and not (isinstance(x, float) and (x != x))
                        else pd.NA
                    )

                    df_d["Month"] = df_d["Creation_Date"].dt.strftime("%b")
                    df_d = df_d[df_d["Month"].notna()]

                    month_order = [
                        "Jan",
                        "Feb",
                        "Mar",
                        "Apr",
                        "May",
                        "Jun",
                        "Jul",
                        "Aug",
                        "Sep",
                        "Oct",
                        "Nov",
                        "Dec",
                    ]
                    month_order_value = {m: i for i, m in enumerate(month_order)}

                    dispatch_summary = (
                        df_d.groupby("Month", as_index=False)
                        .agg(
                            Total_Shipments=("Order_SO", "count"),
                            Hits=("is_hit", "sum"),
                        )
                    )
                    dispatch_summary["Misses"] = (
                        dispatch_summary["Total_Shipments"]
                        - dispatch_summary["Hits"]
                    )
                    dispatch_summary = dispatch_summary.sort_values(
                        by="Month",
                        key=lambda c: c.map(
                            lambda m: month_order_value.get(m, 99)
                        ),
                    )
                    ordered_dispatch = dispatch_summary["Month"].tolist()
                    agg_col = None
                    if len(ordered_dispatch) >= 2:
                        agg_col = str(
                            int(df_d["Creation_Date"].dropna().dt.year.max())
                        )
                    pivot_dispatch = ["KPI"] + ordered_dispatch + (
                        [agg_col] if agg_col else []
                    )

                    hit_pct_d = {"KPI": "Hit %"}
                    hit_d = {"KPI": "Hit"}
                    miss_d = {"KPI": "Miss"}
                    total_d = {"KPI": "Total Shipments"}
                    for _, r in dispatch_summary.iterrows():
                        m = r["Month"]
                        t = int(r["Total_Shipments"])
                        h = int(r["Hits"])
                        ms = int(r["Misses"])
                        hp = int(round(h * 100 / t)) if t else 0
                        mp = int(round(ms * 100 / t)) if t else 0
                        hit_pct_d[m] = hp
                        hit_d[m] = f"{h} ({hp}%)"
                        miss_d[m] = f"{ms} ({mp}%)"
                        total_d[m] = t
                    if agg_col:
                        t_all = int(dispatch_summary["Total_Shipments"].sum())
                        h_all = int(dispatch_summary["Hits"].sum())
                        ms_all = t_all - h_all
                        hp_all = int(round(h_all * 100 / t_all)) if t_all else 0
                        mp_all = int(round(ms_all * 100 / t_all)) if t_all else 0
                        hit_pct_d[agg_col] = hp_all
                        hit_d[agg_col] = f"{h_all} ({hp_all}%)"
                        miss_d[agg_col] = f"{ms_all} ({mp_all}%)"
                        total_d[agg_col] = t_all

                    dispatch_chart_data = [
                        {
                            "type": "column",
                            "name": "Dispatch Hit %",
                            "color": "#81613E",
                            "related_table": "sub-table-b2c-dispatch",
                            "dataPoints": [
                                {"label": m, "y": int(hit_pct_d[m])}
                                for m in ordered_dispatch
                                if int(hit_pct_d[m]) > 0
                            ],
                        }
                    ]

                    sub_tables.append(
                        {
                            "id": "sub-table-b2c-dispatch",
                            "title": "Dispatch — Creation to Dispatch",
                            "columns": pivot_dispatch,
                            "data": [
                                hit_pct_d,
                                hit_d,
                                miss_d,
                                total_d,
                            ],
                            "full_width": False,
                            "side_by_side_chart": True,
                            "chart_data": dispatch_chart_data,
                        }
                    )

            # ——— جدول Last Mile KPI: Dispatch → Delivered خلال 48 hours = Hit ———
            if (
                status_col
                and dispatch_col
                and delivered_col
                and creation_col
            ):
                df_lm = df_full[
                    df_full[status_col].astype(str).str.strip().str.upper()
                    == "DELIVERED"
                ][[order_col, creation_col, dispatch_col, delivered_col]].copy()
                df_lm = df_lm.rename(
                    columns={
                        order_col: "Order_SO",
                        creation_col: "Creation_Date",
                        dispatch_col: "Dispatch_Date",
                        delivered_col: "Delivered_Date",
                    }
                )
                df_lm["Creation_Date"] = _excel_dates_to_datetime(
                    df_lm["Creation_Date"], errors="coerce"
                )
                df_lm["Dispatch_Date"] = _excel_dates_to_datetime(
                    df_lm["Dispatch_Date"], errors="coerce"
                )
                df_lm["Delivered_Date"] = _excel_dates_to_datetime(
                    df_lm["Delivered_Date"], errors="coerce"
                )
                df_lm = df_lm.dropna(
                    subset=[
                        "Order_SO",
                        "Creation_Date",
                        "Dispatch_Date",
                        "Delivered_Date",
                    ]
                )
                if not df_lm.empty:
                    df_lm = (
                        df_lm.sort_values(["Order_SO", "Dispatch_Date"])
                        .drop_duplicates(subset=["Order_SO"], keep="first")
                    )
                    hours_48 = (
                        df_lm["Delivered_Date"] - df_lm["Dispatch_Date"]
                    ).dt.total_seconds() / 3600.0
                    df_lm["is_hit"] = (hours_48 <= 48) & hours_48.notna()
                    df_lm["Duration_Hours"] = hours_48
                    df_lm["Duration_Hours"] = df_lm["Duration_Hours"].apply(
                        lambda x: int(round(x))
                        if pd.notna(x) and not (isinstance(x, float) and (x != x))
                        else pd.NA
                    )

                    # Same month buckets as Pick & Peak / Dispatch (creation month), not dispatch month
                    df_lm["Month"] = df_lm["Creation_Date"].dt.strftime("%b")
                    df_lm = df_lm[df_lm["Month"].notna()]

                    month_order_lm = [
                        "Jan",
                        "Feb",
                        "Mar",
                        "Apr",
                        "May",
                        "Jun",
                        "Jul",
                        "Aug",
                        "Sep",
                        "Oct",
                        "Nov",
                        "Dec",
                    ]
                    month_order_value_lm = {
                        m: i for i, m in enumerate(month_order_lm)
                    }

                    lm_summary = (
                        df_lm.groupby("Month", as_index=False)
                        .agg(
                            Total_Shipments=("Order_SO", "count"),
                            Hits=("is_hit", "sum"),
                        )
                    )
                    lm_summary["Misses"] = (
                        lm_summary["Total_Shipments"] - lm_summary["Hits"]
                    )
                    lm_summary = lm_summary.sort_values(
                        by="Month",
                        key=lambda c: c.map(
                            lambda m: month_order_value_lm.get(m, 99)
                        ),
                    )
                    ordered_lm = lm_summary["Month"].tolist()
                    agg_col_lm = None
                    if len(ordered_lm) >= 2:
                        agg_col_lm = str(
                            int(df_lm["Creation_Date"].dropna().dt.year.max())
                        )
                    pivot_lm = ["KPI"] + ordered_lm + (
                        [agg_col_lm] if agg_col_lm else []
                    )

                    hit_pct_lm = {"KPI": "Hit %"}
                    hit_lm = {"KPI": "Hit"}
                    miss_lm = {"KPI": "Miss"}
                    total_lm = {"KPI": "Total Shipments"}
                    for _, r in lm_summary.iterrows():
                        m = r["Month"]
                        t = int(r["Total_Shipments"])
                        h = int(r["Hits"])
                        ms = int(r["Misses"])
                        hp = int(round(h * 100 / t)) if t else 0
                        mp = int(round(ms * 100 / t)) if t else 0
                        hit_pct_lm[m] = hp
                        hit_lm[m] = f"{h} ({hp}%)"
                        miss_lm[m] = f"{ms} ({mp}%)"
                        total_lm[m] = t
                    if agg_col_lm:
                        t_all = int(lm_summary["Total_Shipments"].sum())
                        h_all = int(lm_summary["Hits"].sum())
                        ms_all = t_all - h_all
                        hp_all = int(round(h_all * 100 / t_all)) if t_all else 0
                        mp_all = int(round(ms_all * 100 / t_all)) if t_all else 0
                        hit_pct_lm[agg_col_lm] = hp_all
                        hit_lm[agg_col_lm] = f"{h_all} ({hp_all}%)"
                        miss_lm[agg_col_lm] = f"{ms_all} ({mp_all}%)"
                        total_lm[agg_col_lm] = t_all

                    lastmile_chart_data = [
                        {
                            "type": "column",
                            "name": "Last Mile Hit %",
                            "color": "#9F8170",
                            "related_table": "sub-table-b2c-lastmile",
                            "dataPoints": [
                                {"label": m, "y": int(hit_pct_lm[m])}
                                for m in ordered_lm
                                if int(hit_pct_lm[m]) > 0
                            ],
                        }
                    ]

                    sub_tables.append(
                        {
                            "id": "sub-table-b2c-lastmile",
                            "title": "Last Mile KPI — Dispatch / Delivered (≤48 hours)",
                            "columns": pivot_lm,
                            "data": [
                                hit_pct_lm,
                                hit_lm,
                                miss_lm,
                                total_lm,
                            ],
                            "full_width": False,
                            "side_by_side_chart": True,
                            "chart_data": lastmile_chart_data,
                        }
                    )

            # ——— جدول End to End (Creation / Delivered): خلال 48 hours = Hit ———
            if status_col and creation_col and delivered_col:
                df_ee = df_full[
                    df_full[status_col].astype(str).str.strip().str.upper()
                    == "DELIVERED"
                ][[order_col, creation_col, delivered_col]].copy()
                df_ee = df_ee.rename(
                    columns={
                        order_col: "Order_SO",
                        creation_col: "Creation_Date",
                        delivered_col: "Delivered_Date",
                    }
                )
                df_ee["Creation_Date"] = _excel_dates_to_datetime(
                    df_ee["Creation_Date"], errors="coerce"
                )
                df_ee["Delivered_Date"] = _excel_dates_to_datetime(
                    df_ee["Delivered_Date"], errors="coerce"
                )
                df_ee = df_ee.dropna(
                    subset=["Order_SO", "Creation_Date", "Delivered_Date"]
                )
                if not df_ee.empty:
                    df_ee = (
                        df_ee.sort_values(["Order_SO", "Creation_Date"])
                        .drop_duplicates(subset=["Order_SO"], keep="first")
                    )
                    hours_48_ee = (
                        df_ee["Delivered_Date"] - df_ee["Creation_Date"]
                    ).dt.total_seconds() / 3600.0
                    df_ee["is_hit"] = (hours_48_ee <= 48) & hours_48_ee.notna()
                    df_ee["Duration_Hours"] = hours_48_ee
                    df_ee["Duration_Hours"] = df_ee["Duration_Hours"].apply(
                        lambda x: int(round(x))
                        if pd.notna(x) and not (isinstance(x, float) and (x != x))
                        else pd.NA
                    )

                    df_ee["Month"] = df_ee["Creation_Date"].dt.strftime("%b")
                    df_ee = df_ee[df_ee["Month"].notna()]

                    month_order_ee = [
                        "Jan",
                        "Feb",
                        "Mar",
                        "Apr",
                        "May",
                        "Jun",
                        "Jul",
                        "Aug",
                        "Sep",
                        "Oct",
                        "Nov",
                        "Dec",
                    ]
                    month_order_value_ee = {
                        m: i for i, m in enumerate(month_order_ee)
                    }

                    ee_summary = (
                        df_ee.groupby("Month", as_index=False)
                        .agg(
                            Total_Shipments=("Order_SO", "count"),
                            Hits=("is_hit", "sum"),
                        )
                    )
                    ee_summary["Misses"] = (
                        ee_summary["Total_Shipments"] - ee_summary["Hits"]
                    )
                    ee_summary = ee_summary.sort_values(
                        by="Month",
                        key=lambda c: c.map(
                            lambda m: month_order_value_ee.get(m, 99)
                        ),
                    )
                    ordered_ee = ee_summary["Month"].tolist()
                    agg_col_ee = None
                    if len(ordered_ee) >= 2:
                        agg_col_ee = str(
                            int(df_ee["Creation_Date"].dropna().dt.year.max())
                        )
                    pivot_ee = ["KPI"] + ordered_ee + (
                        [agg_col_ee] if agg_col_ee else []
                    )

                    hit_pct_ee = {"KPI": "Hit %"}
                    hit_ee = {"KPI": "Hit"}
                    miss_ee = {"KPI": "Miss"}
                    total_ee = {"KPI": "Total Shipments"}
                    for _, r in ee_summary.iterrows():
                        m = r["Month"]
                        t = int(r["Total_Shipments"])
                        h = int(r["Hits"])
                        ms = int(r["Misses"])
                        hp = int(round(h * 100 / t)) if t else 0
                        mp = int(round(ms * 100 / t)) if t else 0
                        hit_pct_ee[m] = hp
                        hit_ee[m] = f"{h} ({hp}%)"
                        miss_ee[m] = f"{ms} ({mp}%)"
                        total_ee[m] = t
                    if agg_col_ee:
                        t_all = int(ee_summary["Total_Shipments"].sum())
                        h_all = int(ee_summary["Hits"].sum())
                        ms_all = t_all - h_all
                        hp_all = int(round(h_all * 100 / t_all)) if t_all else 0
                        mp_all = int(round(ms_all * 100 / t_all)) if t_all else 0
                        hit_pct_ee[agg_col_ee] = hp_all
                        hit_ee[agg_col_ee] = f"{h_all} ({hp_all}%)"
                        miss_ee[agg_col_ee] = f"{ms_all} ({mp_all}%)"
                        total_ee[agg_col_ee] = t_all

                    endtoend_chart_data = [
                        {
                            "type": "column",
                            "name": "End to End Hit %",
                            "color": "#81613E",
                            "related_table": "sub-table-b2c-endtoend",
                            "dataPoints": [
                                {"label": m, "y": int(hit_pct_ee[m])}
                                for m in ordered_ee
                                if int(hit_pct_ee[m]) > 0
                            ],
                        }
                    ]

                    sub_tables.append(
                        {
                            "id": "sub-table-b2c-endtoend",
                            "title": "End to End — Creation / Delivered (≤48 hours)",
                            "columns": pivot_ee,
                            "data": [
                                hit_pct_ee,
                                hit_ee,
                                miss_ee,
                                total_ee,
                            ],
                            "full_width": False,
                            "side_by_side_chart": True,
                            "chart_data": endtoend_chart_data,
                        }
                    )

            total_shipments = len(df)
            hits = int(df["is_hit"].sum())
            miss = total_shipments - hits
            hit_pct = round((hits / total_shipments) * 100, 2) if total_shipments else 0

            # جدول الإكسل الخام: نفس حد القراءة الكاملة للتاب (بدون قص إضافي 500 صف)
            _raw_nr = _read_excel_nrows_kw(_b2c_max_rows)
            raw_df_original = pd.read_excel(
                excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_raw_nr
            )
            raw_df_original.columns = raw_df_original.columns.astype(str).str.strip()
            raw_sheet_cols = list(raw_df_original.columns)

            def _raw_cell_val(val):
                if val is None:
                    return ""
                if pd.isna(val) or (isinstance(val, float) and (val != val)):
                    return ""
                if hasattr(val, "strftime"):
                    try:
                        return pd.Timestamp(val).strftime("%Y-%m-%d %I:%M %p")
                    except Exception:
                        return str(val)
                s = str(val).strip()
                if s.lower() in ("nan", "nat", "none", "<nat>"):
                    return ""
                return s

            creation_col_raw = find_col(
                raw_df_original,
                [
                    "CREATTION DATE",
                    "CREATION DATE",
                    "Creation Date & Time",
                    "Creation Date and Time",
                    "Create Timestamp",
                    "Creation DateTime",
                    "Create Date",
                ],
            )
            _month_abbr_order = [
                "Jan",
                "Feb",
                "Mar",
                "Apr",
                "May",
                "Jun",
                "Jul",
                "Aug",
                "Sep",
                "Oct",
                "Nov",
                "Dec",
            ]
            row_month_labels = []
            if creation_col_raw and creation_col_raw in raw_df_original.columns:
                _ts_r = _excel_dates_to_datetime(
                    raw_df_original[creation_col_raw], errors="coerce"
                )
                for i in range(len(raw_df_original)):
                    ts = _ts_r.iloc[i]
                    if pd.isna(ts):
                        row_month_labels.append("")
                    else:
                        try:
                            row_month_labels.append(
                                pd.Timestamp(ts).strftime("%b")
                            )
                        except Exception:
                            row_month_labels.append("")
            else:
                row_month_labels = [""] * len(raw_df_original)

            _seen_m = {m for m in row_month_labels if m}
            month_filter_options = [
                m for m in _month_abbr_order if m in _seen_m
            ]

            raw_records = raw_df_original.to_dict(orient="records")
            raw_excel_rows = []
            for idx, row in enumerate(raw_records):
                rec = {c: _raw_cell_val(row.get(c)) for c in raw_sheet_cols}
                rec["_filter_month"] = (
                    row_month_labels[idx] if idx < len(row_month_labels) else ""
                )
                raw_excel_rows.append(rec)

            raw_excel_table = {
                "id": "sub-table-b2c-raw-sheet",
                "title": "B2C_Outbound (Sheet Data)",
                "columns": [{"name": c, "key": c, "group": "sheet"} for c in raw_sheet_cols],
                "data": raw_excel_rows,
                "full_width": True,
                "month_filter_options": month_filter_options,
            }

            result = {
                "detail_html": "",
                "sub_tables": sub_tables,
                "raw_excel_table": raw_excel_table,
                "stats": {
                    "total": total_shipments,
                    "hit": hits,
                    "miss": miss,
                    "hit_pct": hit_pct,
                    "target": 99,
                },
            }
            cache.set(_b2c_key, result, 120)
            return result

        except Exception as e:
            import traceback

            print(traceback.format_exc())
            return {
                "detail_html": f"<p class='text-danger'>⚠️ Error processing B2C Outbound: {e}</p>",
                "sub_tables": [],
                "raw_excel_table": None,
                "stats": {},
            }

    def _render_safety_kpi_tab(self, request):
        """
        يعرض شيت Safety KPI من ملف الإكسل (اسم الشيت يحتوي safety و kpi، أو مطابقة Safety KPI).
        """
        excel_file_path = self.resolve_excel_file_path(request)
        if not excel_file_path:
            return {
                "detail_html": "<p class='text-danger'>Excel file not found.</p>",
                "chart_data": [],
                "count": 0,
                "hit_pct": 0,
            }
        sheet_names = _get_excel_sheet_names_cached(excel_file_path)
        if not sheet_names:
            return {
                "detail_html": "<p class='text-danger'>No sheets found in the workbook.</p>",
                "chart_data": [],
                "count": 0,
                "hit_pct": 0,
            }

        def _norm_sheet(s):
            return (str(s) or "").lower().replace(" ", "").replace("_", "")

        resolved = None
        for s in sheet_names:
            ns = _norm_sheet(s)
            if "safety" in ns and "kpi" in ns:
                resolved = s
                break
        if not resolved:
            for s in sheet_names:
                if _norm_sheet(s) == _norm_sheet("Safety KPI"):
                    resolved = s
                    break
        if not resolved:
            preview = ", ".join(str(x) for x in sheet_names[:30])
            if len(sheet_names) > 30:
                preview += ", …"
            return {
                "detail_html": (
                    "<div class='alert alert-warning'>No <strong>Safety KPI</strong> sheet found. "
                    "Use a sheet name that includes <em>Safety</em> and <em>KPI</em> (e.g. <code>Safety KPI</code>).</div>"
                    f"<p class='text-muted small mb-0'>Sheets in this file: {preview}</p>"
                ),
                "chart_data": [],
                "count": 0,
                "hit_pct": 0,
            }

        data = self.render_raw_sheet(request, resolved)
        if isinstance(data, dict) and "hit_pct" not in data:
            data = dict(data)
            data.setdefault("hit_pct", 0)
            data.setdefault("chart_data", [])
        return data

    def _placeholder_tab_response(self, tab_name):
        """يرجع استجابة تاب placeholder مع رسالة Loading data."""
        html = (
            "<div class='card p-4 shadow-sm'>"
            "<p class='text-muted text-center mb-0'>Loading data</p>"
            "</div>"
        )
        return {
            "detail_html": html,
            "chart_data": [],
            "count": 0,
            "hit_pct": 0,
        }

    def _norm_col(self, name):
        """تطبيع اسم عمود للمقارنة."""
        if not name or not isinstance(name, str):
            return ""
        return str(name).strip().lower().replace(" ", "").replace("_", "")

    def _find_col(self, cols_map, *candidates):
        """يبحث عن عمود من مرشحين (أسماء ممكنة) مع تطبيع ومرونة في المطابقة."""
        for c in candidates:
            k = self._norm_col(c)
            if not k:
                continue
            for col_key in cols_map:
                col_norm = self._norm_col(col_key)
                # مطابقة تامة بعد التطبيع
                if col_norm == k:
                    return cols_map[col_key]
                # أو أن أحدهما يحتوي الآخر (للأعمدة اللي فيها زيادات مثل Create Timestamp (Local))
                if k in col_norm or col_norm in k:
                    return cols_map[col_key]
        return None

    def _traceability_read_sheets(self, excel_path, request):
        """قراءة شيتي Traceability — دائمًا نطاق كامل (أو EXCEL_FULL_MAX_ROWS) لدقة البحث."""
        sheet_names = _get_excel_sheet_names_cached(excel_path)
        if not sheet_names:
            return None, None
        norm = lambda s: (s or "").lower().replace(" ", "").replace("_", "")

        inbound_name = next(
            (s for s in sheet_names if "traceability" in norm(s) and "inbound" in norm(s)),
            None,
        )
        outbound_name = next(
            (s for s in sheet_names if "traceability" in norm(s) and "outbound" in norm(s)),
            None,
        )
        if not inbound_name:
            return None, None

        df_in = _get_sheet_dataframe(
            excel_path, inbound_name, request=request, force_full=True
        )
        df_out = (
            _get_sheet_dataframe(
                excel_path, outbound_name, request=request, force_full=True
            )
            if outbound_name
            else None
        )
        return df_in, df_out

    def _traceability_search_data(self, request):
        """
        بحث Traceability: فلترة حسب Item Code و/أو Batch Nbr و/أو LPN Nbr من شيت Inbound ثم Outbound.
        يرجع قائمة عناصر كل عنصر: بيانات الوارد + حركات الصادر + الكمية الحالية.
        """
        excel_path = self.resolve_excel_file_path(request)
        if not excel_path:
            return {"error": "Excel file not found.", "results": []}

        item_code = (request.GET.get("item_code") or request.POST.get("item_code") or "").strip()
        batch_nbr = (request.GET.get("batch_nbr") or request.POST.get("batch_nbr") or "").strip()
        lpn_nbr = (request.GET.get("lpn_nbr") or request.POST.get("lpn_nbr") or "").strip()
        if not item_code and not batch_nbr and not lpn_nbr:
            return {
                "error": "Please enter an Item Code, Batch Number, and/or LPN Nbr.",
                "results": [],
            }

        df_in, df_out = self._traceability_read_sheets(excel_path, request)

        if df_in is None or df_in.empty:
            return {"error": "Traceability KPI Inbound sheet not found or is empty.", "results": []}

        cols_in = {c: c for c in df_in.columns}
        lpn_col = self._find_col(cols_in, "LPN Nbr", "LPN Nbr", "LPN")
        item_col = self._find_col(cols_in, "Item Code", "Item Code", "ItemCode")
        batch_col = self._find_col(cols_in, "batch_nbr", "Batch Nbr", "BatchNbr")
        # في شيت Traceability_KPI_Inbound لا يوجد Create Timestamp، لكن يوجد Allocation Mod Timestamp
        create_ts_col = self._find_col(
            cols_in,
            "Create Timestamp",
            "Allocation Mod Timestamp",
            "AllocationModTimestamp",
            "Allocation Timestamp",
        )
        orig_qty_col = self._find_col(cols_in, "Orig Qty", "Orig Qty", "OrigQty")
        expiry_col = self._find_col(cols_in, "Expiry Date", "Expiry Date", "ExpiryDate")
        item_desc_col = self._find_col(
            cols_in, "Item Description", "Item Description", "ItemDescription"
        )
        # عمود الكمية الحالية في الشيت (القيمة المتبقية الآن في المستودع)
        current_qty_in_col = self._find_col(
            cols_in, "Current Qty", "CurrentQty", "Current Quantity"
        )
        # تاريخ استلام الشحنة (من Inbound)
        received_ts_col = self._find_col(
            cols_in,
            "Received Timestamp",
            "Received Timestamp",
            "ReceivedTimestamp",
            "Receipt Timestamp",
            "Receipt Time",
            "Received Date",
            "Goods Receipt",
            "GRN Timestamp",
        )
        received_ts_help = (
            "Received Timestamp is read only from the Traceability KPI Inbound sheet "
            "(«Received Timestamp» or similar receipt column), not from Outbound."
            if received_ts_col
            else (
                "Received Timestamp is empty: no matching receipt column on the Inbound sheet "
                "(this field is not read from Outbound)."
            )
        )

        if not item_col and not batch_col and not lpn_col:
            return {
                "error": "Item Code, Batch Nbr, and LPN Nbr columns were not found in the Inbound sheet.",
                "results": [],
            }
        if item_code and not item_col:
            return {"error": "Item Code column was not found in the Inbound sheet.", "results": []}
        if batch_nbr and not batch_col:
            return {"error": "Batch Nbr column was not found in the Inbound sheet.", "results": []}
        if lpn_nbr and not lpn_col:
            return {"error": "LPN Nbr column was not found in the Inbound sheet.", "results": []}

        def safe_str(v):
            if pd.isna(v):
                return ""
            if isinstance(v, pd.Timestamp):
                return (
                    v.strftime("%Y-%m-%d %H:%M") if hasattr(v, "strftime") else str(v)
                )
            return str(v).strip()

        def normalize_code_value(v):
            """
            يحوّل قيمة كود (Item / Batch) لصيغة نص موحّدة للمقارنة،
            ويتعامل مع الأرقام اللي بتظهر في الإكسل كـ 12345.0 أو scientific notation.
            """
            import numpy as np

            if pd.isna(v):
                return ""
            # أعداد صحيحة
            if isinstance(v, (int, np.integer)):
                return str(v).strip().lower()
            # أعداد عشرية من نوع float (مثلاً 5290378616.0)
            if isinstance(v, (float, np.floating)):
                if np.isfinite(v):
                    if float(v).is_integer():
                        return str(int(v)).strip().lower()
                    # fallback لأعداد عشرية حقيقية
                    return ("%.15g" % float(v)).strip().lower()
                return ""
            s = str(v).strip()
            # قيم مثل "5290378616.0"
            if s.endswith(".0") and s[:-2].isdigit():
                s = s[:-2]
            return s.lower()

        def safe_float(v):
            if pd.isna(v):
                return 0.0
            # لو القيمة رقم أصلاً (int / float / numpy)
            import numpy as np

            if isinstance(v, (int, float, np.integer, np.floating)):
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return 0.0
            # لو القيمة نصية وفيها فواصل آلاف مثلاً "1,234"
            s = str(v).strip().replace(",", "")
            try:
                return float(s)
            except (TypeError, ValueError):
                return 0.0

        # بحث بـ LPN فقط: لازم نفس الـ LPN يظهر في Inbound و Outbound معًا
        lpn_only_search = bool(lpn_nbr and not item_code and not batch_nbr)
        if lpn_only_search:
            q_lpn_match = normalize_code_value(lpn_nbr)
            has_in_lpn = bool(lpn_col) and (
                df_in[lpn_col].map(normalize_code_value) == q_lpn_match
            ).any()
            lpn_out_col = None
            if df_out is not None and not df_out.empty:
                lpn_out_col = self._find_col(
                    {c: c for c in df_out.columns},
                    "LPN Nbr",
                    "LPN Nbr",
                    "LPN",
                )
            msg_base = (
                "LPN Nbr in Traceability KPI Inbound does not match LPN Nbr in "
                "Traceability KPI Outbound."
            )
            if df_out is None or df_out.empty:
                return {
                    "error": msg_base + " (Outbound sheet missing or empty.)",
                    "results": [],
                }
            if not lpn_out_col:
                return {
                    "error": msg_base + " (LPN Nbr column not found on Outbound sheet.)",
                    "results": [],
                }
            has_out_lpn = (
                df_out[lpn_out_col].map(normalize_code_value) == q_lpn_match
            ).any()
            if has_in_lpn and not has_out_lpn:
                return {
                    "error": msg_base + " (This LPN was not found on Outbound.)",
                    "results": [],
                }
            if not has_in_lpn and has_out_lpn:
                return {
                    "error": msg_base + " (This LPN was not found on Inbound.)",
                    "results": [],
                }
            if not has_in_lpn and not has_out_lpn:
                return {"error": "", "results": []}

        # فلترة (AND): Item Code، Batch Nbr، LPN Nbr — حسب ما أدخل المستخدم
        if item_col and item_code:
            q_item = normalize_code_value(item_code)
            df_in = df_in[df_in[item_col].map(normalize_code_value) == q_item]
        if not df_in.empty and batch_col and batch_nbr:
            q_batch = normalize_code_value(batch_nbr)
            df_in = df_in[df_in[batch_col].map(normalize_code_value) == q_batch]
        if not df_in.empty and lpn_col and lpn_nbr:
            q_lpn = normalize_code_value(lpn_nbr)
            df_in = df_in[df_in[lpn_col].map(normalize_code_value) == q_lpn]

        if df_in.empty:
            return {"error": "", "results": []}

        cols_out = {}
        df_out_filtered = None
        if df_out is not None and not df_out.empty:
            cols_out = {c: c for c in df_out.columns}
            lpn_out = self._find_col(cols_out, "LPN Nbr", "LPN Nbr", "LPN")
            item_out = self._find_col(cols_out, "Item Code", "Item Code", "ItemCode")
            batch_out = self._find_col(cols_out, "batch_nbr", "Batch Nbr", "BatchNbr")
            packed_qty_col = self._find_col(
                cols_out,
                "Packed Qty",
                "PackedQty",
                "Packed Quantity",
                "Quantity",
                "Qty",
            )
            customer_col = self._find_col(
                cols_out, "Customer Name", "Customer Name", "CustomerName"
            )
            picked_time_col = self._find_col(
                cols_out,
                "Detail Picked Time",
                "Detail Picked Time",
                "DetailPickedTime",
            )
            packed_ts_col = self._find_col(
                cols_out, "Packed Timestamp", "Packed Timestamp", "PackedTimestamp"
            )
            current_qty_col = self._find_col(
                cols_out, "Current Qty", "Current Qty", "CurrentQty"
            )

            mask_out = pd.Series(False, index=df_out.index)
            if item_out and item_code:
                q_item = normalize_code_value(item_code)
                mask_out = mask_out | (
                    df_out[item_out].map(normalize_code_value) == q_item
                )
            if batch_out and batch_nbr:
                q_batch = normalize_code_value(batch_nbr)
                mask_out = mask_out | (
                    df_out[batch_out].map(normalize_code_value) == q_batch
                )
            if lpn_out and lpn_nbr:
                q_lpn_o = normalize_code_value(lpn_nbr)
                mask_out = mask_out | (
                    df_out[lpn_out].map(normalize_code_value) == q_lpn_o
                )
            df_out_filtered = df_out[mask_out].copy() if mask_out.any() else None
        else:
            packed_qty_col = customer_col = picked_time_col = packed_ts_col = current_qty_col = lpn_out = item_out = batch_out = None

        # بناء قائمة Outbound مرة واحدة (مع Item Code بدل LPN)
        outbound_list = []
        if df_out_filtered is not None and not df_out_filtered.empty:
            _out_cols = [
                c
                for c in (
                    packed_qty_col,
                    item_out,
                    batch_out,
                    lpn_out,
                    customer_col,
                    picked_time_col,
                    packed_ts_col,
                )
                if c
            ]
            _sub_out = (
                df_out_filtered[_out_cols] if _out_cols else df_out_filtered
            )
            for rec in _sub_out.to_dict("records"):
                pq = (
                    safe_float(rec.get(packed_qty_col)) if packed_qty_col else 0.0
                )
                outbound_list.append(
                    {
                        "item_code": safe_str(rec.get(item_out)) if item_out else "",
                        "batch_nbr": safe_str(rec.get(batch_out)) if batch_out else "",
                        "lpn_nbr": safe_str(rec.get(lpn_out)) if lpn_out else "",
                        "packed_qty": pq,
                        "customer_name": (
                            safe_str(rec.get(customer_col)) if customer_col else ""
                        ),
                        "detail_picked_time": (
                            safe_str(rec.get(picked_time_col))
                            if picked_time_col
                            else ""
                        ),
                        "packed_timestamp": (
                            safe_str(rec.get(packed_ts_col)) if packed_ts_col else ""
                        ),
                    }
                )
            def _parse_dt(s):
                if s is None or (isinstance(s, str) and not s.strip()):
                    return pd.Timestamp.min
                if hasattr(s, "to_pydatetime"):
                    return s
                try:
                    return _excel_dates_to_datetime(s, errors="coerce")
                except Exception:
                    return pd.Timestamp.min

            outbound_list.sort(
                key=lambda o: _parse_dt(o.get("packed_timestamp") or o.get("detail_picked_time")),
                reverse=False,
            )

        # تجميع حسب التاريخ فقط (نفس اليوم): صفين نفس التاريخ → صف واحد + جمع Orig Qty
        def _norm_ts(v):
            if pd.isna(v):
                return ""
            if hasattr(v, "strftime"):
                return v.strftime("%Y-%m-%d")
            s = str(v).strip()
            try:
                dt = _excel_dates_to_datetime(s, errors="coerce")
                return dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else s
            except Exception:
                return s

        _create_ts_col_orig = create_ts_col
        if not create_ts_col:
            create_ts_col = "__dummy__"
            df_in = df_in.copy()
            df_in["__dummy__"] = ""

        df_in = df_in.copy()
        df_in["_group_ts"] = df_in[create_ts_col].map(_norm_ts)

        results = []
        for _group_ts, grp in df_in.groupby("_group_ts", sort=False):
            if not _create_ts_col_orig and _group_ts == "":
                pass
            elif not _group_ts:
                continue
            create_ts_display = grp[create_ts_col].iloc[0]
            create_ts_str = safe_str(create_ts_display)
            if orig_qty_col and orig_qty_col in grp.columns:
                orig_qty = grp[orig_qty_col].apply(safe_float).sum()
            else:
                inbound_current_sum = (
                    grp[current_qty_in_col].apply(safe_float).sum()
                    if current_qty_in_col and current_qty_in_col in grp.columns
                    else 0.0
                )
                packed_total = sum(safe_float(o.get("packed_qty")) for o in outbound_list) if outbound_list else 0.0
                orig_qty = inbound_current_sum + packed_total
            expiry = safe_str(grp[expiry_col].iloc[0]) if expiry_col and expiry_col in grp.columns else ""
            item_desc_raw = safe_str(grp[item_desc_col].iloc[0]) if item_desc_col and item_desc_col in grp.columns else ""
            item_desc = _clean_traceability_item_description(item_desc_raw)
            received_ts_str = safe_str(grp[received_ts_col].iloc[0]) if received_ts_col and received_ts_col in grp.columns else ""
            item_code_display = safe_str(grp[item_col].iloc[0]) if item_col and item_col in grp.columns else (item_code or "")
            batch_display = (
                safe_str(grp[batch_col].iloc[0])
                if batch_col and batch_col in grp.columns
                else ""
            )
            lpn_display = (
                safe_str(grp[lpn_col].iloc[0])
                if lpn_col and lpn_col in grp.columns
                else ""
            )
            current_qty = (
                grp[current_qty_in_col].apply(safe_float).sum()
                if current_qty_in_col and current_qty_in_col in grp.columns
                else 0.0
            )
            results.append(
                {
                    "item_code": item_code_display,
                    "batch_nbr": batch_display,
                    "lpn_nbr": lpn_display,
                    "create_timestamp": create_ts_str,
                    "orig_qty": orig_qty,
                    "expiry_date": expiry,
                    "item_description": item_desc,
                    "received_timestamp": received_ts_str,
                    "outbounds": outbound_list,
                    "current_qty": current_qty,
                }
            )

        # ترتيب صفوف الـ Inbound من الأقدم للأحدث
        def _parse_create_ts(r):
            ts = r.get("create_timestamp") or ""
            if not ts:
                return pd.Timestamp.min
            try:
                return _excel_dates_to_datetime(ts, errors="coerce")
            except Exception:
                return pd.Timestamp.min

        results.sort(key=lambda r: _parse_create_ts(r), reverse=False)

        def _coerce_ts(val):
            if val is None:
                return pd.NaT
            if isinstance(val, str) and not val.strip():
                return pd.NaT
            try:
                t = _excel_dates_to_datetime(val, errors="coerce")
                return t if pd.notna(t) else pd.NaT
            except Exception:
                return pd.NaT

        def _outbound_event_ts(o):
            for key in ("packed_timestamp", "detail_picked_time"):
                t = _coerce_ts(o.get(key))
                if pd.notna(t):
                    return t
            return pd.Timestamp.min

        timeline = []
        for r in results:
            ir = {k: v for k, v in r.items() if k != "outbounds"}
            timeline.append(
                {
                    "kind": "inbound",
                    "display_date": ir.get("create_timestamp") or "",
                    "inbound": ir,
                }
            )
        for o in outbound_list:
            od = {
                "item_code": o.get("item_code"),
                "batch_nbr": o.get("batch_nbr"),
                "lpn_nbr": o.get("lpn_nbr"),
                "packed_qty": o.get("packed_qty"),
                "customer_name": o.get("customer_name"),
                "detail_picked_time": o.get("detail_picked_time"),
                "packed_timestamp": o.get("packed_timestamp"),
            }
            label = od.get("packed_timestamp") or od.get("detail_picked_time") or ""
            timeline.append(
                {
                    "kind": "outbound",
                    "display_date": label,
                    "outbound": od,
                }
            )

        def _ev_sort_key(ev):
            if ev["kind"] == "inbound":
                t = _coerce_ts(ev["inbound"].get("create_timestamp"))
                sub = 0
            else:
                t = _outbound_event_ts(ev["outbound"])
                sub = 1
            if pd.isna(t) or t == pd.Timestamp.min:
                t = pd.Timestamp.max
            return (t, sub)

        timeline.sort(key=_ev_sort_key)

        return {
            "error": "",
            "results": results,
            "timeline": timeline,
            "received_timestamp_help": received_ts_help,
        }

    def traceability_search(self, request):
        """استجابة AJAX لبحث Traceability (Item Code / Batch Nbr / LPN Nbr)."""
        if getattr(settings, "DEBUG", False):
            print(
                "[Traceability] search item_code=%s batch_nbr=%s lpn_nbr=%s"
                % (
                    request.GET.get("item_code", ""),
                    request.GET.get("batch_nbr", ""),
                    request.GET.get("lpn_nbr", ""),
                )
            )
        data = self._traceability_search_data(request)
        return JsonResponse(data, safe=False)

    def _traceability_kpi_tab_response(self, request):
        """يرجع HTML تاب Traceability KPI: حقلان بحث + تحميل + نتائج. البحث يعمل عبر event delegation في الصفحة الرئيسية (بدون سكربت هنا لتجنب خطأ appendChild/replaceChild)."""
        html = (
            '<div class="card p-4 shadow-sm mx-auto" style="border-color: #C3B091; max-width: 100%;">'
            '<div class="card-header mb-3" style="background-color: #E3DAC9; color: #81613E; border-color: #C3B091;">'
            "<h5 class='mb-0'>Traceability KPI — Shipment Traceability</h5>"
            "</div>"
            "<div class='mb-3'>"
            "<label class='form-label fw-semibold' style='color: #81613E;'>Search by Item Code, Batch Nbr, and/or LPN Nbr</label>"
            "<div class='d-flex flex-wrap align-items-center justify-content-center gap-2'>"
            "<input type='text' id='traceability-item-code' class='form-control' placeholder='Item Code' style='border-color: #C3B091; max-width: 180px;' />"
            "<input type='text' id='traceability-batch-nbr' class='form-control' placeholder='Batch Nbr' style='border-color: #C3B091; max-width: 180px;' />"
            "<input type='text' id='traceability-lpn-nbr' class='form-control' placeholder='LPN Nbr' style='border-color: #C3B091; max-width: 180px;' />"
            "<button type='button' id='traceability-search-btn' class='btn text-white' style='background-color: #9F8170; border-color: #9F8170;'>Search</button>"
            "</div>"
            "<small class='text-muted'>Enter at least one field, then Search or press Enter.</small>"
            "</div>"
            "<div id='traceability-loading' class='mb-3' style='display:none;'>"
            "<p class='text-muted mb-0'><span class='spinner-border spinner-border-sm me-2' role='status'></span>Searching...</p>"
            "</div>"
            
            "<div id='traceability-results-container'></div>"
        )
        return {
            "detail_html": html,
            "chart_data": [],
            "count": 0,
            "hit_pct": 0,
        }

    def _render_b2c_outbound_tab(
        self, request, selected_month=None, selected_months=None
    ):
        """يرجع HTML تاب B2C Outbound للاستجابة AJAX (نفس أسلوب B2B Outbound)."""
        res = self.filter_b2c_outbound(
            request,
            selected_month=selected_month,
            selected_months=selected_months,
        )
        stats = res.get("stats") or {}
        hit_pct = stats.get("hit_pct", 0)
        try:
            hit_pct = round(float(hit_pct), 2)
        except (TypeError, ValueError):
            hit_pct = 0
        count = stats.get("total", 0) or 0
        b2c_tab = {
            "name": "B2C Outbound",
            "stats": stats,
            "sub_tables": res.get("sub_tables", []),
            "raw_excel_table": res.get("raw_excel_table"),
        }
        html = render_to_string(
            "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
            {
                "tab": b2c_tab,
                "selected_month": selected_month,
                "selected_months": selected_months,
            },
        )
        return {
            "detail_html": html,
            "chart_data": [],
            "count": count,
            "hit_pct": hit_pct,
        }

    def filter_inbound(self, request, selected_month=None, selected_months=None):
        """
        تاب Inbound: يقرأ من ملف all_sheet_nespresso.xlsx، شيت "inbound_tab".
        - الشيت الجديد بدون مناطق: KPI واحد + شارت (Hit/Miss بالشهر)، ثم جدول تفاصيل شيت الإكسل تحته.
        - من Create Timestamp إلى Last LPN Rcv TS: يوم أو أقل = Hit، أكثر = Miss.
        - أعمدة مضافة في جدول التفاصيل: Days (رقم صحيح)، HIT or MISS، Within 24h.
        """
        try:
            import os

            # تاب Inbound يقرأ من all_sheet_nespresso.xlsx (كل التابات ما عدا Dashboard)
            excel_path = _get_excel_path_for_request(request)
            if not excel_path or not os.path.exists(excel_path):
                return {
                    "detail_html": "<p class='text-danger'>⚠️ Excel file not found.</p>",
                    "sub_tables": [],
                    "chart_data": [],
                }

            xls = pd.ExcelFile(excel_path, engine="openpyxl")
            sheet_name = None

            def _norm(s):
                return (str(s).strip().lower().replace(" ", "").replace("_", "") if s else "")

            # أولوية: شيت inbound_tab (يقبل "inbound_tab" أو "inbound tab" أو "Inbound Tab")
            for s in xls.sheet_names:
                if _norm(s) == "inboundtab":
                    sheet_name = s
                    break

            # ثم شيت ARAMCO Inbound Report القديم إن وجد
            if not sheet_name:
                for s in xls.sheet_names:
                    if "ARAMCO Inbound Report" in (s or "").strip():
                        sheet_name = s
                        break

            # وأخيراً أي شيت يحتوي على كلمة inbound (للتوافق)
            if not sheet_name:
                sheet_name = next((s for s in xls.sheet_names if "inbound" in (s or "").lower()), None)
            if not sheet_name:
                available = ", ".join(str(s) for s in (xls.sheet_names or [])[:20])
                if len(xls.sheet_names or []) > 20:
                    available += ", …"
                return {
                    "detail_html": (
                        "<p class='text-warning'>⚠️ الشيت <strong>inbound_tab</strong> غير موجود داخل الملف.</p>"
                        "<p class='text-muted small'>المفروض: الملف = <strong>all_sheet_nespresso.xlsx</strong>، والشيت جواه = <strong>inbound_tab</strong>.</p>"
                        "<p class='text-muted small'>تأكد أن الملف اللي بيُقرأ هو all_sheet_nespresso.xlsx (رفعه أو ضعه في مجلد الرفع)، وأن بداخله شيت اسمه بالظبط <strong>inbound_tab</strong>.</p>"
                        f"<p class='text-muted small'>الشيتات الموجودة حالياً في الملف: {available}</p>"
                    ),
                    "sub_tables": [],
                    "chart_data": [],
                }

            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
            df = pd.read_excel(
                excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
            )
            if df.empty:
                return {
                    "detail_html": "<p class='text-warning'>⚠️ Inbound sheet is empty.</p>",
                    "sub_tables": [],
                    "chart_data": [],
                }

            df.columns = df.columns.astype(str).str.strip()

            # إذا الصف الأول عنوان (مثل "ARAMCO Inbound Report") والرؤوس في صف تالي، نكتشف صف الرؤوس
            first_col = str(df.columns[0]).strip() if len(df.columns) else ""
            if (
                first_col.startswith("Unnamed:")
                or first_col == "ARAMCO Inbound Report"
                or (first_col and "inbound report" in first_col.lower())
            ):
                raw = pd.read_excel(
                    excel_path,
                    sheet_name=sheet_name,
                    engine="openpyxl",
                    header=None,
                    **_nr_kw,
                )
                if raw.empty or raw.shape[0] < 2:
                    raw.columns = raw.columns.astype(str).str.strip()
                else:
                    header_row_idx = None
                    for idx in range(min(10, raw.shape[0])):
                        row = raw.iloc[idx]
                        cells = " ".join(str(c).strip().lower() for c in row.dropna().astype(str))
                        if (
                            "facility" in cells
                            and ("shipment" in cells or "shipment_nbr" in cells)
                            and ("create" in cells or "creation" in cells)
                            and ("received" in cells or "lpn" in cells)
                        ):
                            header_row_idx = idx
                            break
                    if header_row_idx is not None:
                        df = raw.iloc[header_row_idx + 1 :].copy()
                        headers = [str(c).strip() if pd.notna(c) and str(c).strip() else f"Col_{i}" for i, c in enumerate(raw.iloc[header_row_idx].values)]
                        df.columns = headers
                        df = df.reset_index(drop=True)
                    else:
                        df = raw.copy()
                        headers = [str(c).strip() if pd.notna(c) and str(c).strip() else f"Col_{i}" for i, c in enumerate(raw.iloc[0].values)]
                        df.columns = headers
                        df = df.iloc[1:].reset_index(drop=True)

            df.columns = df.columns.astype(str).str.strip()

            def normalize_name(val):
                return re.sub(r"[^a-z0-9]", "", str(val).strip().lower())

            def find_column(possible_names):
                normalized_map = {normalize_name(col): col for col in df.columns}
                for name in possible_names:
                    norm = normalize_name(name)
                    if norm in normalized_map:
                        return normalized_map[norm]
                for col in df.columns:
                    col_norm = normalize_name(col)
                    if any(normalize_name(n) in col_norm for n in possible_names):
                        return col
                return None

            # مرادفات كثيرة لأن أسماء الأعمدة في الإكسل تختلف (مسافات، شرطات، رموز)
            col_facility = find_column([
                "Facility", "facility", "Facility Code", "facility code", "FacilityCode",
                "Site", "Warehouse", "Location", "Facility Name", "facility name",
                # شيت ARAMCO Inbound Report الجديد يستخدم Region بدل Facility
                "Region",
            ])
            col_shipment = find_column([
                "Shipment_nbr", "Shipment nbr", "Shipment Nbr", "Shipment No", "Shipment Number",
                "shipment number", "Shipment ID", "ShipmentID", "Shipment #", "Shipment#",
                "Shipment No.", "ShipmentNbr", "Shipment Nbr.",
                # شيت ARAMCO Inbound Report: Shipment_ID
                "Shipment_ID",
            ])
            col_create = find_column([
                "Create shipment D&T", "Create Shipment D&T", "Create shipment D&T", "Create Shipment D&T",
                "Create Timestamp", "create timestamp", "Creation Date", "Create Date", "Created Date",
                "Shipment Create Date", "Create Date & Time", "Create D&T", "Create DT",
                "Create shipement D&T", "Create shipemnt D&T", "Create Shipement D&T",
                # شيت ARAMCO Inbound Report: نعتبر Ship_Date هو تاريخ الإنشاء
                "Ship_Date",
            ])
            col_received = find_column([
                "Received LPN D&T", "Received LPN D&T", "Last LPN Rcv TS", "last lpn rcv ts",
                "Received LPN Date", "LPN Received Date", "Receipt Date", "Received Date",
                "Last LPN Receive", "LPN Rcv TS", "Received D&T", "Received DT",
                "Received LPN D&T", "Received LPN DT",
                # شيت ARAMCO Inbound Report: نستخدم Receiving_Complete_Date أو Verified_Date
                "Receiving_Complete_Date", "Verified_Date",
            ])
            col_first_rcv = find_column([
                "First LPN Rcv TS", "first lpn rcv ts", "First LPN Rcv", "First LPN Receive",
                "First Received LPN", "First LPN Receive Date", "First Received D&T",
            ])
            col_status = find_column(["Status", "status", "Shipment Status", "State"])
            col_lpn = find_column(["LPN", "LPN Nbr", "LPN_nbr", "LPNs", "LPN Number", "LPN No"])

            # بحث ثاني: أي عمود اسمه يحتوي على الكلمات المفتاحية (لو الأسماء مختلفة جداً)
            def find_column_containing(*keywords):
                k_norm = [normalize_name(k) for k in keywords]
                for col in df.columns:
                    c = normalize_name(col)
                    if all(k in c for k in k_norm):
                        return col
                return None

            if not col_facility:
                col_facility = find_column_containing("facility")
            if not col_shipment:
                col_shipment = find_column_containing("shipment", "nbr") or find_column_containing("shipment", "no") or find_column_containing("shipment", "number") or find_column_containing("shipment", "id")
            if not col_create:
                col_create = find_column_containing("create", "shipment") or find_column_containing("create", "timestamp") or find_column_containing("create", "date") or find_column_containing("creation", "date")
            if not col_received:
                col_received = find_column_containing("received", "lpn") or find_column_containing("lpn", "rcv") or find_column_containing("lpn", "receive") or find_column_containing("last", "lpn")

            if not col_shipment or not col_create or not col_received:
                missing = []
                if not col_shipment:
                    missing.append("Shipment Nbr")
                if not col_create:
                    missing.append("Create Timestamp")
                if not col_received:
                    missing.append("Last LPN Rcv TS")
                actual_cols = ", ".join(str(c) for c in df.columns.tolist()[:20])
                if len(df.columns) > 20:
                    actual_cols += ", …"
                return {
                    "detail_html": (
                        f"<p class='text-danger'>⚠️ Missing required columns: {', '.join(missing)}</p>"
                        f"<p class='text-muted small mt-2'>الأعمدة الموجودة في الشيت: {actual_cols}</p>"
                    ),
                    "sub_tables": [],
                    "chart_data": [],
                }

            renames = {
                col_shipment: "Shipment_nbr",
                col_create: "Create_shipment_DT",
                col_received: "Received_LPN_DT",
            }
            if col_facility:
                renames[col_facility] = "Facility"
            if col_first_rcv:
                renames[col_first_rcv] = "First_LPN_Rcv_DT"
            df = df.rename(columns=renames)
            if col_status:
                df = df.rename(columns={col_status: "Status"})
            else:
                df["Status"] = ""
            if col_lpn:
                df = df.rename(columns={col_lpn: "LPN"})
            else:
                df["LPN"] = range(len(df))

            if col_facility:
                df["Facility"] = df["Facility"].astype(str).str.strip()
            df["Shipment_nbr"] = df["Shipment_nbr"].astype(str).str.strip()
            df["Create_shipment_DT"] = _excel_dates_to_datetime(
                df["Create_shipment_DT"], errors="coerce"
            )
            df["Received_LPN_DT"] = _excel_dates_to_datetime(
                df["Received_LPN_DT"], errors="coerce"
            )
            if "First_LPN_Rcv_DT" in df.columns:
                df["First_LPN_Rcv_DT"] = _excel_dates_to_datetime(
                    df["First_LPN_Rcv_DT"], errors="coerce"
                )
            df["Month"] = df["Create_shipment_DT"].dt.strftime("%b").fillna("")

            if df.empty:
                return {
                    "detail_html": "<p class='text-warning'>⚠️ Inbound sheet has no valid rows.</p>",
                    "sub_tables": [],
                    "chart_data": [],
                }

            def month_order_value(label):
                if not label:
                    return 999
                label = (label or "").strip()[:3].title()
                for idx in range(1, 13):
                    if month_abbr[idx] == label:
                        return idx
                return 999

            # KPI واحد لكل الشحنات (بدون تقسيم مناطق — الشيت الجديد مفيهوش مناطق)
            create_min = df.groupby("Shipment_nbr")["Create_shipment_DT"].min()
            received_max = df.groupby("Shipment_nbr")["Received_LPN_DT"].max()

            hits = 0
            misses = 0
            rows = []
            ship_month_hit = []
            # ⚠️ لتقليل الحمل على السيرفر، نقيّد عدد الشحنات المفصّلة (مثلاً أول 500 شحنة فقط)
            max_shipments_for_detail = 500
            unique_shipments = list(df["Shipment_nbr"].unique())
            limited_shipments = unique_shipments[:max_shipments_for_detail]

            for ship_id in limited_shipments:
                if not ship_id:
                    continue
                create_ts = create_min.get(ship_id)
                received_ts = received_max.get(ship_id)
                if pd.isna(create_ts) or pd.isna(received_ts):
                    continue
                month = create_ts.strftime("%b") if pd.notna(create_ts) else ""
                total_seconds = (received_ts - create_ts).total_seconds()
                delta_days = total_seconds / (24 * 3600)
                days_used = int(round(delta_days)) if delta_days >= 0 else 0
                delta_hours = total_seconds / 3600 if total_seconds >= 0 else None
                within_24h = None
                if delta_hours is not None:
                    within_24h = "≤24h" if delta_hours <= 24 else ">24h"
                is_hit = delta_days <= 1
                if is_hit:
                    hits += 1
                else:
                    misses += 1
                ship_month_hit.append((month, is_hit))
                rows.append({
                    "Shipment_nbr": ship_id,
                    "Days": days_used,
                    "HIT or MISS": "Hit" if is_hit else "Miss",
                    "Within 24h": within_24h,
                    "Month": month,
                })

            by_month = {}
            for month, is_hit in ship_month_hit:
                if not month:
                    continue
                if month not in by_month:
                    by_month[month] = {"total": 0, "hit": 0, "miss": 0}
                by_month[month]["total"] += 1
                if is_hit:
                    by_month[month]["hit"] += 1
                else:
                    by_month[month]["miss"] += 1
            for m in by_month:
                by_month[m]["hit_pct"] = round((by_month[m]["hit"] / by_month[m]["total"]) * 100, 2) if by_month[m]["total"] else 0

            ordered_months = sorted(by_month.keys(), key=lambda x: month_order_value(x))
            shipment_metrics = {r["Shipment_nbr"]: {"Days": r["Days"], "HIT or MISS": r["HIT or MISS"], "Within 24h": r["Within 24h"]} for r in rows}

            # الشارت: Hit و Miss بالشهر فقط (بدون رقم Facility أو مناطق بجانب الشهر)
            chart_data = [{
                "type": "column",
                "name": "Hit",
                "color": "#9F8170",
                "valueSuffix": "",
                "related_table": "inbound-aggregated-kpi",
                "dataPoints": [{"label": m, "y": by_month.get(m, {}).get("hit", 0)} for m in ordered_months],
            }, {
                "type": "column",
                "name": "Miss",
                "color": "#81613E",
                "valueSuffix": "",
                "related_table": "inbound-aggregated-kpi",
                "dataPoints": [{"label": m, "y": by_month.get(m, {}).get("miss", 0)} for m in ordered_months],
            }]

            overall_total = hits + misses
            overall_hits = hits
            overall_miss = misses
            overall_hit_pct = round((overall_hits / overall_total) * 100, 2) if overall_total else 0

            # جدول KPI واحد (بالشهور + 2025) مع الشارت
            pivot_cols = ["KPI"] + ordered_months + ["2025"]
            hit_pct_row = {"KPI": "Hit %"}
            hit_row = {"KPI": "Hit"}
            miss_row = {"KPI": "Miss"}
            total_row = {"KPI": "Total Shipments"}
            for m in ordered_months:
                b = by_month.get(m, {})
                hit_pct_row[m] = int(round(b.get("hit_pct", 0)))
                hit_row[m] = b.get("hit", 0)
                miss_row[m] = b.get("miss", 0)
                total_row[m] = b.get("total", 0)
            hit_pct_row["2025"] = int(round(overall_hit_pct))
            hit_row["2025"] = overall_hits
            miss_row["2025"] = overall_miss
            total_row["2025"] = overall_total
            aggregated_kpi_table = {
                "id": "inbound-aggregated-kpi",
                "title": "Inbound KPI ≤ 24h",
                "columns": pivot_cols,
                "data": [hit_pct_row, hit_row, miss_row, total_row],
                "chart_data": chart_data,
                "canvas_id": "chart-inbound-24h",
                "full_width": False,
            }

            # جدول ثاني: First LPN Rcv TS → Last LPN Rcv TS، استبعاد Cancelled، < 18 ساعة = Hit
            second_kpi_table = None
            if "First_LPN_Rcv_DT" in df.columns:
                df2 = df[df["Status"].astype(str).str.strip().str.lower() != "cancelled"].copy()
                if not df2.empty:
                    first_min = df2.groupby("Shipment_nbr")["First_LPN_Rcv_DT"].min()
                    last_max = df2.groupby("Shipment_nbr")["Received_LPN_DT"].max()
                    facility_first = df2.groupby("Shipment_nbr")["Facility"].first() if "Facility" in df2.columns else pd.Series(dtype=object)
                    status_first = df2.groupby("Shipment_nbr")["Status"].first()

                    hits_18 = 0
                    misses_18 = 0
                    rows_18h = []
                    ship_month_hit_18 = []
                    for ship_id in df2["Shipment_nbr"].unique():
                        if not ship_id:
                            continue
                        first_ts = first_min.get(ship_id)
                        last_ts = last_max.get(ship_id)
                        if pd.isna(first_ts) or pd.isna(last_ts):
                            continue
                        total_seconds = (last_ts - first_ts).total_seconds()
                        hours_val = round(total_seconds / 3600, 2) if total_seconds >= 0 else 0
                        is_hit_18 = hours_val < 18
                        if is_hit_18:
                            hits_18 += 1
                        else:
                            misses_18 += 1
                        month_18 = first_ts.strftime("%b") if pd.notna(first_ts) else ""
                        ship_month_hit_18.append((month_18, is_hit_18))
                        fac = facility_first.get(ship_id, "")
                        st = status_first.get(ship_id, "")
                        rows_18h.append({
                            "Facility Code": fac,
                            "Shipment Nbr": ship_id,
                            "Status": st,
                            "First LPN Rcv TS": first_ts.strftime("%Y-%m-%d %H:%M") if pd.notna(first_ts) else "",
                            "Last LPN Rcv TS": last_ts.strftime("%Y-%m-%d %H:%M") if pd.notna(last_ts) else "",
                            "Hours (ساعات)": hours_val,
                            "HIT or MISS": "Hit" if is_hit_18 else "Miss",
                            "_month": month_18,
                        })

                    by_month_18 = {}
                    for month_18, is_hit in ship_month_hit_18:
                        if not month_18:
                            continue
                        if month_18 not in by_month_18:
                            by_month_18[month_18] = {"total": 0, "hit": 0, "miss": 0}
                        by_month_18[month_18]["total"] += 1
                        if is_hit:
                            by_month_18[month_18]["hit"] += 1
                        else:
                            by_month_18[month_18]["miss"] += 1
                    for m in by_month_18:
                        by_month_18[m]["hit_pct"] = round((by_month_18[m]["hit"] / by_month_18[m]["total"]) * 100, 2) if by_month_18[m]["total"] else 0
                    ordered_months_18 = sorted(by_month_18.keys(), key=lambda x: month_order_value(x))

                    chart_data_18h = [{
                        "type": "column",
                        "name": "Hit",
                        "color": "#9F8170",
                        "valueSuffix": "",
                        "related_table": "inbound-aggregated-kpi-18h",
                        "dataPoints": [{"label": m, "y": by_month_18.get(m, {}).get("hit", 0)} for m in ordered_months_18],
                    }, {
                        "type": "column",
                        "name": "Miss",
                        "color": "#81613E",
                        "valueSuffix": "",
                        "related_table": "inbound-aggregated-kpi-18h",
                        "dataPoints": [{"label": m, "y": by_month_18.get(m, {}).get("miss", 0)} for m in ordered_months_18],
                    }]

                    total_18 = hits_18 + misses_18
                    hit_pct_18 = round((hits_18 / total_18) * 100, 2) if total_18 else 0
                    pivot_cols_18 = ["KPI"] + ordered_months_18 + ["2025"]
                    hit_pct_row_18 = {"KPI": "Hit %"}
                    hit_row_18 = {"KPI": "Hit"}
                    miss_row_18 = {"KPI": "Miss"}
                    total_row_18 = {"KPI": "Total Shipments"}
                    for m in ordered_months_18:
                        b = by_month_18.get(m, {})
                        hit_pct_row_18[m] = int(round(b.get("hit_pct", 0)))
                        hit_row_18[m] = b.get("hit", 0)
                        miss_row_18[m] = b.get("miss", 0)
                        total_row_18[m] = b.get("total", 0)
                    hit_pct_row_18["2025"] = int(round(hit_pct_18))
                    hit_row_18["2025"] = hits_18
                    miss_row_18["2025"] = misses_18
                    total_row_18["2025"] = total_18

                    second_kpi_table = {
                        "id": "inbound-aggregated-kpi-18h",
                        "title": "Inbound KPI ≤ 18h (First → Last LPN)",
                        "columns": pivot_cols_18,
                        "data": [hit_pct_row_18, hit_row_18, miss_row_18, total_row_18],
                        "chart_data": chart_data_18h,
                        "canvas_id": "chart-inbound-18h",
                        "full_width": False,
                    }

            # استبعاد أعمدة الأسماء التلقائية (مثل Col_4) اللي مش موجودة فعلياً في الإكسل
            def _is_auto_col(name):
                return bool(re.match(r"^Col_\d+$", str(name).strip()))
            raw_columns = [c for c in df.columns if not c.startswith("_") and not _is_auto_col(c)]
            if "Facility" in raw_columns and "Facility Code" not in raw_columns:
                raw_columns = ["Facility Code" if c == "Facility" else c for c in raw_columns]
            if "Shipment_nbr" in raw_columns:
                raw_columns = ["Shipment Nbr" if c == "Shipment_nbr" else c for c in raw_columns]
            if "Create_shipment_DT" in raw_columns:
                raw_columns = ["Create Timestamp" if c == "Create_shipment_DT" else c for c in raw_columns]
            if "Received_LPN_DT" in raw_columns:
                raw_columns = ["Last LPN Rcv TS" if c == "Received_LPN_DT" else c for c in raw_columns]
            drop_cols = [c for c in df.columns if c.startswith("_") or _is_auto_col(c)]
            raw_df = df.drop(columns=drop_cols, errors="ignore").copy()
            if "Facility" in raw_df.columns and "Facility Code" not in raw_df.columns:
                raw_df["Facility Code"] = raw_df["Facility"]
            if "Shipment_nbr" in raw_df.columns:
                raw_df["Shipment Nbr"] = raw_df["Shipment_nbr"]
            if "Create_shipment_DT" in raw_df.columns:
                raw_df["Create Timestamp"] = raw_df["Create_shipment_DT"]
            if "Received_LPN_DT" in raw_df.columns:
                raw_df["Last LPN Rcv TS"] = raw_df["Received_LPN_DT"]

            # إضافة الأعمدة المحسوبة (عدد الأيام صحيح، Hit/Miss، وحالة 24 ساعة) بجانب تواريخ الإنشاء والاستلام
            def _days_int(v):
                if v is None or (isinstance(v, float) and (pd.isna(v) or v != v)):
                    return ""
                try:
                    return int(round(float(v)))
                except (TypeError, ValueError):
                    return v

            if shipment_metrics and ("Shipment_nbr" in raw_df.columns or "Shipment Nbr" in raw_df.columns):
                ship_col = "Shipment Nbr" if "Shipment Nbr" in raw_df.columns else "Shipment_nbr"
                raw_df["Days"] = raw_df[ship_col].map(
                    lambda s: _days_int(shipment_metrics.get(s, {}).get("Days"))
                )
                raw_df["HIT or MISS"] = raw_df[ship_col].map(
                    lambda s: shipment_metrics.get(s, {}).get("HIT or MISS")
                )
                raw_df["Within 24h"] = raw_df[ship_col].map(
                    lambda s: shipment_metrics.get(s, {}).get("Within 24h")
                )

                # وضع الأعمدة الجديدة بعد عمود الاستلام (Last LPN Rcv TS) أو Create Timestamp
                insert_after = None
                if "Last LPN Rcv TS" in raw_columns:
                    insert_after = "Last LPN Rcv TS"
                elif "Create Timestamp" in raw_columns:
                    insert_after = "Create Timestamp"

                new_cols = ["Days", "HIT or MISS", "Within 24h"]
                if insert_after and insert_after in raw_columns:
                    idx = raw_columns.index(insert_after) + 1
                    for c in new_cols:
                        if c not in raw_columns:
                            raw_columns.insert(idx, c)
                            idx += 1
                else:
                    for c in new_cols:
                        if c not in raw_columns:
                            raw_columns.append(c)
            # إضافة عمود الشهر للفلتر (من تاريخ إنشاء الشحنة)
            if "Create_shipment_DT" in raw_df.columns:
                raw_df["Month"] = _excel_dates_to_datetime(raw_df["Create_shipment_DT"], errors="coerce").dt.strftime("%b")
                if "Month" not in raw_columns:
                    raw_columns = list(raw_columns) + ["Month"]

            def _to_blank(val):
                if val is None or (isinstance(val, float) and (pd.isna(val) or val != val)):
                    return ""
                s = str(val).strip()
                if s.lower() in ("nan", "nat", "none", "<nat>"):
                    return ""
                return s

            for col in raw_df.columns:
                if pd.api.types.is_datetime64_any_dtype(raw_df[col]):
                    raw_df[col] = raw_df[col].apply(lambda x: x.strftime("%Y-%m-%d %H:%M") if pd.notna(x) else "")

            detail_rows = [{k: _to_blank(v) for k, v in row.items()} for row in raw_df.head(500).to_dict(orient="records")]

            facility_options = sorted(df["Facility"].dropna().unique().astype(str).tolist()) if "Facility" in df.columns else []
            month_options = sorted(raw_df["Month"].dropna().unique().tolist()) if "Month" in raw_df.columns else []
            # المناطق حسب الشهر (لتاب Inbound فقط)
            regions_by_month = []
            if "Facility" in df.columns and "Month" in df.columns:
                g = df.dropna(subset=["Month", "Facility"]).groupby("Month")["Facility"].apply(
                    lambda x: sorted(x.astype(str).str.strip().unique().tolist())
                )
                for month in ordered_months:
                    regions = g.get(month, [])
                    if regions:
                        regions_by_month.append({"month": month, "regions": regions})
            detail_table = {
                "id": "sub-table-inbound-detail",
                "title": "Inbound Shipments Detail",
                "columns": raw_columns,
                "data": detail_rows,
                "chart_data": [],
                "full_width": True,
                "filter_options": {
                    "facility_codes": facility_options,
                    "months": month_options,
                },
            }

            # شارت موحّد كبير: داتا الجدولين (24h + 18h) — كل عمود/سلسلة = داتا جدول
            all_months_combined = sorted(
                set(ordered_months) | (set(ordered_months_18) if second_kpi_table else set()),
                key=lambda x: month_order_value(x),
            )
            combined_chart_data = [
                {
                    "type": "column",
                    "name": "≤24h Hit",
                    "color": "#9F8170",
                    "valueSuffix": "",
                    "dataPoints": [{"label": m, "y": by_month.get(m, {}).get("hit", 0)} for m in all_months_combined],
                },
                {
                    "type": "column",
                    "name": "≤24h Miss",
                    "color": "#81613E",
                    "valueSuffix": "",
                    "dataPoints": [{"label": m, "y": by_month.get(m, {}).get("miss", 0)} for m in all_months_combined],
                },
            ]
            if second_kpi_table:
                combined_chart_data.extend([
                    {
                        "type": "column",
                        "name": "≤18h Hit",
                        "color": "#A0785A",
                        "valueSuffix": "",
                        "dataPoints": [{"label": m, "y": by_month_18.get(m, {}).get("hit", 0)} for m in all_months_combined],
                    },
                    {
                        "type": "column",
                        "name": "≤18h Miss",
                        "color": "#EDC9AF",
                        "valueSuffix": "",
                        "dataPoints": [{"label": m, "y": by_month_18.get(m, {}).get("miss", 0)} for m in all_months_combined],
                    },
                ])
            combined_chart_sub = {
                "id": "inbound-combined-chart",
                "title": "Inbound KPI (≤24h & ≤18h)",
                "columns": [],
                "data": [],
                "chart_data": combined_chart_data,
                "canvas_id": "chart-inbound-combined",
                "full_width": True,
            }

            # جدول KPI + جدول 18h (إن وُجد) + تفاصيله، ثم جدول تفاصيل شيت الإكسل — الشارت الموحّد يُعرض أولاً في التمبلت
            sub_tables = [combined_chart_sub, aggregated_kpi_table]
            if second_kpi_table:
                sub_tables.append(second_kpi_table)
            sub_tables.append(detail_table)

            # بناء هيكل التاب لاستخدامه مع تمبلت excel-sheet-table
            inbound_tab = {
                "name": "Inbound",
                "stats": {
                    "total": overall_total,
                    "hit": overall_hits,
                    "miss": overall_miss,
                    "hit_pct": overall_hit_pct,
                },
                "sub_tables": sub_tables,
            }

            from django.template.loader import render_to_string

            html = render_to_string(
                "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
                {
                    "tab": inbound_tab,
                    "selected_month": selected_month,
                    "selected_months": selected_months,
                },
            )

            return {
                "detail_html": html,
                "sub_tables": sub_tables,
                "chart_data": chart_data,
                "stats": inbound_tab["stats"],
                "regions_by_month": regions_by_month,
            }

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return {
                "detail_html": f"<p class='text-danger'>⚠️ Error while processing inbound data: {e}</p>",
                "sub_tables": [],
                "chart_data": [],
            }

    def filter_capacity_expiry(self, request, selected_month=None, selected_months=None):
        """
        تاب Capacity + Expiry: يقرأ من ملف all_sheet_nespresso.xlsx (أو الملف المرفوع)، شيت "Capacity + Expiry_tab".
        - فلترة: Facility، Order Nbr، Status = Allocated فقط.
        - جدول Capacity: عد الوكيشنات (From Location) للحالة Allocated.
        - جدول Expiry: تجميع حسب batch_nbr و Expiry Date؛ فترات 1–3، 3–6، 6–9 أشهر؛ تحذير أحمر للقريب من الانتهاء.
        - جدول التفاصيل: شيت الإكسل مع الفلاتر.
        """
        import os
        from datetime import datetime, timedelta
        from django.template.loader import render_to_string

        try:
            excel_path = _get_excel_path_for_request(request)
            if not excel_path or not os.path.exists(excel_path):
                return {
                    "detail_html": "<p class='text-danger'>⚠️ Excel file not found.</p>",
                    "sub_tables": [],
                    "chart_data": [],
                    "count": 0,
                }

            xls = pd.ExcelFile(excel_path, engine="openpyxl")

            def _norm(s):
                return (str(s).strip().lower().replace(" ", "").replace("+", "").replace("_", "") if s else "")

            # أولوية: شيت Capacity + Expiry_tab (يقبل "Capacity + Expiry_tab" أو "Capacity + Expiry tab")
            sheet_name = None
            for s in xls.sheet_names:
                if _norm(s) == "capacityexpirytab":
                    sheet_name = s
                    break
            if not sheet_name:
                for s in xls.sheet_names:
                    if "capacity" in _norm(s) and "expiry" in _norm(s):
                        sheet_name = s
                        break
            # بديل: شيت "Expiry" لو الملف القديم ما فيهوش Capacity + Expiry_tab
            if not sheet_name:
                for s in xls.sheet_names:
                    if (s or "").strip().lower() == "expiry":
                        sheet_name = s
                        break
            if not sheet_name:
                available = ", ".join(str(x) for x in (xls.sheet_names or [])[:15])
                if len(xls.sheet_names or []) > 15:
                    available += ", …"
                return {
                    "detail_html": (
                        "<p class='text-warning'>⚠️ الشيت <strong>Capacity + Expiry_tab</strong> غير موجود داخل الملف.</p>"
                        "<p class='text-muted small'>المفروض: الملف = <strong>all_sheet_nespresso.xlsx</strong> (أو الملف المرفوع)، والشيت جواه = <strong>Capacity + Expiry_tab</strong>.</p>"
                        "<p class='text-muted small'>لو عندك ملف قديم، ضيف شيت باسم <strong>Capacity + Expiry_tab</strong> أو ارفع الملف الجديد.</p>"
                        f"<p class='text-muted small'>الشيتات الموجودة حالياً: {available}</p>"
                    ),
                    "sub_tables": [],
                    "chart_data": [],
                    "count": 0,
                }

            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
            df = pd.read_excel(
                excel_path, sheet_name=sheet_name, engine="openpyxl", header=0, **_nr_kw
            )
            if df.empty:
                return {
                    "detail_html": "<p class='text-warning'>⚠️ الشيت فاضي.</p>",
                    "sub_tables": [],
                    "chart_data": [],
                    "count": 0,
                }

            df.columns = df.columns.astype(str).str.strip()

            def _col(name_candidates):
                for c in df.columns:
                    for n in name_candidates:
                        if n.lower() in str(c).lower() or str(c).lower().replace(" ", "") == n.lower().replace(" ", ""):
                            return c
                return None

            facility_col = _col(["Facility", "Facility Code"])
            order_col = _col(["Order Nbr", "Order Nbr.", "Order_Nbr", "OrderNbr"])
            status_col = _col(["Status", "status"])
            from_loc_col = _col(["From Location", "From_Location", "FromLocation"])
            batch_col = _col(["batch_nbr", "Batch Nbr", "Batch_Nbr", "BatchNbr"])
            expiry_col = _col(["Expiry Date", "Expiry_Date", "ExpiryDate", "Expiry"])

            if not status_col:
                return {
                    "detail_html": "<p class='text-danger'>⚠️ عمود Status غير موجود في الشيت.</p>",
                    "sub_tables": [],
                    "chart_data": [],
                    "count": 0,
                }

            facility_filter = (request.GET.get("facility") or "").strip()
            order_filter = (request.GET.get("order_nbr") or "").strip()
            df_filtered = df.copy()
            if facility_filter and facility_col:
                df_filtered = df_filtered[df_filtered[facility_col].astype(str).str.strip() == facility_filter]
            if order_filter and order_col:
                df_filtered = df_filtered[df_filtered[order_col].astype(str).str.strip().str.lower().str.contains(order_filter.lower(), na=False)]

            # Capacity: Used = عدد الوكيشنات (From Location) اللي Status = Allocated، Available = الباقي
            total_locations = int(df_filtered[from_loc_col].nunique()) if from_loc_col and from_loc_col in df_filtered.columns else 0
            df_allocated = df_filtered[df_filtered[status_col].astype(str).str.strip().str.lower() == "allocated"].copy()
            used_locations = int(df_allocated[from_loc_col].nunique()) if from_loc_col and from_loc_col in df_allocated.columns and not df_allocated.empty else 0
            available_locations = max(0, total_locations - used_locations)
            used_pct = round((used_locations / total_locations) * 100, 1) if total_locations else 0
            available_pct = round(100 - used_pct, 1)

            capacity_counts = {
                "locations_allocated": used_locations,
                "locations_available": available_locations,
                "total_locations": total_locations,
            }
            if df_allocated.empty:
                capacity_chart = {"utilization_pct_list": []}
                # جدول: Capacity | Utilization | Empty | Percentage (بدون Pending، بدون مدن)
                cap, util, empty = total_locations, 0, total_locations
                pct = "—"
                capacity_rows = [{"Capacity": cap, "Utilization": util, "Empty": empty, "Percentage": pct}]
                capacity_rows.append({"Capacity": cap, "Utilization": util, "Empty": empty, "Percentage": pct, "_is_total": True})
                capacity_table = {
                    "id": "capacity-summary",
                    "title": "Capacity",
                    "columns": ["Capacity", "Utilization", "Empty", "Percentage"],
                    "data": capacity_rows,
                    "chart_data": [],
                    "full_width": False,
                }
                expiry_table = {
                    "id": "expiry-summary",
                    "title": "Expiry (Allocated by batch — near expiry أولاً مع تحذير)",
                    "columns": ["Batch Nbr", "Expiry Date", "1-3 months", "3-6 months", "6-9 months", "Near expiry"],
                    "data": [],
                    "chart_data": [],
                    "full_width": False,
                }
                raw_columns = [c for c in df.columns if not str(c).startswith("_")]
                def _to_blank(v):
                    if v is None or (isinstance(v, float) and pd.isna(v)):
                        return ""
                    s = str(v).strip()
                    if s.lower() in ("nan", "nat", "none", ""):
                        return ""
                    return s
                detail_df = df_filtered.copy()
                today_empty = pd.Timestamp.now().normalize()
                if expiry_col and expiry_col in detail_df.columns:
                    detail_df["_expiry_dt"] = _excel_dates_to_datetime(detail_df[expiry_col], errors="coerce")
                    detail_df["_days_to_expiry"] = (detail_df["_expiry_dt"] - today_empty).dt.days
                    def _days_lbl(d):
                        if pd.isna(d):
                            return ""
                        d = int(d)
                        if d == 0:
                            return "Today"
                        if d > 0:
                            return f"In {d} day{'s' if d != 1 else ''}"
                        return f"Expired {abs(d)} day{'s' if d != -1 else ''} ago"
                    detail_df["Days left to expiry"] = detail_df["_days_to_expiry"].apply(_days_lbl)
                    detail_df["_near_expiry"] = detail_df["_days_to_expiry"].notna() & (detail_df["_days_to_expiry"] <= 30)
                    idx = raw_columns.index(expiry_col) + 1 if expiry_col in raw_columns else len(raw_columns)
                    raw_columns = raw_columns[:idx] + ["Days left to expiry"] + [c for c in raw_columns[idx:] if c != "Days left to expiry"]
                else:
                    detail_df["Days left to expiry"] = ""
                    detail_df["_near_expiry"] = False
                    raw_columns = raw_columns + ["Days left to expiry"]
                for col in detail_df.columns:
                    if col.startswith("_"):
                        continue
                    if pd.api.types.is_datetime64_any_dtype(detail_df[col]):
                        detail_df[col] = detail_df[col].apply(lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) and hasattr(x, "strftime") else "")
                detail_df = detail_df.sort_values(by="_days_to_expiry", ascending=True, na_position="last") if "_days_to_expiry" in detail_df.columns else detail_df
                detail_rows = detail_df.head(500).to_dict(orient="records")
                for r in detail_rows:
                    r.pop("_expiry_dt", None)
                    r.pop("_days_to_expiry", None)
                detail_rows = [{k: _to_blank(v) if k != "_near_expiry" else v for k, v in row.items()} for row in detail_rows]
                expiry_counts = {"within_1_3": 0, "within_3_6": 0, "within_6_9": 0}
            else:
                expiry_counts = {"within_1_3": 0, "within_3_6": 0, "within_6_9": 0}
                # جدول: Capacity | Utilization | Empty | Percentage (بدون Pending، بدون مدن)
                capacity_rows = []
                if from_loc_col:
                    if facility_col:
                        # لكل facility: Capacity = إجمالي locations في df_filtered، Utilization = allocated
                        facilities = df_filtered[facility_col].astype(str).str.strip().unique()
                        for fac in facilities:
                            cap = int(df_filtered[df_filtered[facility_col].astype(str).str.strip() == fac][from_loc_col].nunique())
                            util = int(df_allocated[df_allocated[facility_col].astype(str).str.strip() == fac][from_loc_col].nunique()) if not df_allocated.empty else 0
                            empty = max(0, cap - util)
                            capacity_rows.append({"Capacity": cap, "Utilization": util, "Empty": empty})
                    else:
                        cap = int(df_filtered[from_loc_col].nunique())
                        util = int(df_allocated[from_loc_col].nunique())
                        empty = max(0, cap - util)
                        capacity_rows.append({"Capacity": cap, "Utilization": util, "Empty": empty})
                else:
                    capacity_rows = [{"Capacity": 0, "Utilization": 0, "Empty": 0}]
                total_util = sum(r["Utilization"] for r in capacity_rows)
                utilization_pct_list = []
                for r in capacity_rows:
                    pct_val = round(r["Utilization"] / total_util * 100, 1) if total_util else 0
                    utilization_pct_list.append(pct_val)
                    r["Percentage"] = f"{pct_val}%" if total_util else "—"
                capacity_chart = {"utilization_pct_list": utilization_pct_list}
                cap_tot = sum(r["Capacity"] for r in capacity_rows)
                util_tot = sum(r["Utilization"] for r in capacity_rows)
                empty_tot = sum(r["Empty"] for r in capacity_rows)
                pct_tot = "100.0%" if total_util else "—"
                capacity_rows.append({"Capacity": cap_tot, "Utilization": util_tot, "Empty": empty_tot, "Percentage": pct_tot, "_is_total": True})
                capacity_table = {
                    "id": "capacity-summary",
                    "title": "Capacity",
                    "columns": ["Capacity", "Utilization", "Empty", "Percentage"],
                    "data": capacity_rows,
                    "chart_data": [],
                    "full_width": False,
                }

                today = pd.Timestamp.now().normalize()
                expiry_buckets = []
                count_1_3 = count_3_6 = count_6_9 = 0
                if batch_col and expiry_col:
                    df_allocated[expiry_col] = _excel_dates_to_datetime(df_allocated[expiry_col], errors="coerce")
                    df_exp = df_allocated.dropna(subset=[expiry_col]).copy()
                    df_exp["_months_to_expiry"] = (df_exp[expiry_col] - today).dt.days / 30.44
                    df_exp["_days_to_expiry"] = (df_exp[expiry_col] - today).dt.days
                    for _, row in df_exp.drop_duplicates(subset=[batch_col] if batch_col else [expiry_col]).iterrows():
                        batch_val = row.get(batch_col, row.get(expiry_col))
                        exp_date = row[expiry_col]
                        months = row["_months_to_expiry"]
                        days_int = int(row["_days_to_expiry"]) if not pd.isna(row["_days_to_expiry"]) else None
                        if pd.isna(months):
                            continue
                        m1_3 = "✓" if 1 <= months <= 3 else ""
                        m3_6 = "✓" if 3 < months <= 6 else ""
                        m6_9 = "✓" if 6 < months <= 9 else ""
                        near = "⚠️" if months < 1 else ""
                        if m1_3:
                            count_1_3 += 1
                        if m3_6:
                            count_3_6 += 1
                        if m6_9:
                            count_6_9 += 1
                        if days_int is not None and days_int >= 0:
                            days_label = "Today" if days_int == 0 else f"In {days_int} day{'s' if days_int != 1 else ''}"
                        elif days_int is not None and days_int < 0:
                            days_label = f"Expired {abs(days_int)} day{'s' if days_int != -1 else ''} ago"
                        else:
                            days_label = ""
                        expiry_buckets.append({
                            "Batch Nbr": batch_val,
                            "Expiry Date": exp_date.strftime("%Y-%m-%d") if hasattr(exp_date, "strftime") else str(exp_date),
                            "1-3 months": m1_3,
                            "3-6 months": m3_6,
                            "6-9 months": m6_9,
                            "Near expiry": near,
                            "_months": months,
                            "_days": days_int if days_int is not None else 9999,
                            "_is_near": 1 if near else 0,
                        })
                    expiry_buckets.sort(key=lambda x: (-x["_is_near"], x["_days"], str(x.get("Batch Nbr", ""))))
                    for r in expiry_buckets:
                        del r["_months"]
                        del r["_days"]
                        del r["_is_near"]
                else:
                    expiry_buckets = []

                expiry_cols = ["Batch Nbr", "Expiry Date", "1-3 months", "3-6 months", "6-9 months", "Near expiry"]
                expiry_table = {
                    "id": "expiry-summary",
                    "title": "Expiry (Allocated by batch — 1-3 / 3-6 / 6-9 months; near expiry أولاً مع تحذير)",
                    "columns": expiry_cols,
                    "data": expiry_buckets,
                    "chart_data": [],
                    "full_width": False,
                }
                expiry_counts = {"within_1_3": count_1_3, "within_3_6": count_3_6, "within_6_9": count_6_9}

                raw_columns = [c for c in df.columns if not str(c).startswith("_")]
                def _to_blank(v):
                    if v is None or (isinstance(v, float) and pd.isna(v)):
                        return ""
                    s = str(v).strip()
                    if s.lower() in ("nan", "nat", "none", ""):
                        return ""
                    return s
                detail_df = df_filtered.copy()
                today = pd.Timestamp.now().normalize()
                if expiry_col and expiry_col in detail_df.columns:
                    detail_df["_expiry_dt"] = _excel_dates_to_datetime(detail_df[expiry_col], errors="coerce")
                    detail_df["_days_to_expiry"] = (detail_df["_expiry_dt"] - today).dt.days
                    def _days_label(days):
                        if pd.isna(days):
                            return ""
                        d = int(days)
                        if d == 0:
                            return "Today"
                        if d > 0:
                            return f"In {d} day{'s' if d != 1 else ''}"
                        return f"Expired {abs(d)} day{'s' if d != -1 else ''} ago"
                    detail_df["Days left to expiry"] = detail_df["_days_to_expiry"].apply(_days_label)
                    detail_df["_near_expiry"] = detail_df["_days_to_expiry"].notna() & (detail_df["_days_to_expiry"] <= 30)
                    insert_idx = raw_columns.index(expiry_col) + 1 if expiry_col in raw_columns else len(raw_columns)
                    raw_columns = raw_columns[:insert_idx] + ["Days left to expiry"] + [c for c in raw_columns[insert_idx:] if c != "Days left to expiry"]
                else:
                    detail_df["Days left to expiry"] = ""
                    detail_df["_near_expiry"] = False
                    raw_columns = raw_columns + ["Days left to expiry"]
                for col in detail_df.columns:
                    if col.startswith("_"):
                        continue
                    if pd.api.types.is_datetime64_any_dtype(detail_df[col]):
                        detail_df[col] = detail_df[col].apply(lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) and hasattr(x, "strftime") else "")
                detail_df = detail_df.sort_values(by="_days_to_expiry", ascending=True, na_position="last") if "_days_to_expiry" in detail_df.columns else detail_df
                detail_rows = detail_df.head(500).to_dict(orient="records")
                for r in detail_rows:
                    if "_expiry_dt" in r:
                        del r["_expiry_dt"]
                    if "_days_to_expiry" in r:
                        del r["_days_to_expiry"]
                detail_rows = [{k: _to_blank(v) if k != "_near_expiry" else v for k, v in row.items()} for row in detail_rows]

            facility_options = sorted(df[facility_col].dropna().unique().astype(str).tolist()) if facility_col and facility_col in df.columns else []
            status_options = sorted(df[status_col].dropna().astype(str).str.strip().unique().tolist()) if status_col in df.columns else []
            expiry_options = []
            if expiry_col and expiry_col in df.columns:
                exp_ser = _excel_dates_to_datetime(df[expiry_col], errors="coerce")
                expiry_options = exp_ser.dropna().dt.strftime("%Y-%m-%d").unique().tolist()
                expiry_options = sorted([str(x) for x in expiry_options if str(x) and str(x) not in ("nan", "NaT")])

            detail_table = {
                "id": "capacity-expiry-detail",
                "title": "Capacity + Expiry (Excel sheet)",
                "columns": raw_columns if not df.empty else df.columns.tolist(),
                "data": detail_rows if not df.empty else [],
                "chart_data": [],
                "full_width": True,
                "filter_options": {
                    "facility_codes": facility_options,
                    "statuses": status_options,
                    "expiry_dates": expiry_options,
                    "facility_column": facility_col or "Facility",
                    "status_column": status_col or "Status",
                    "expiry_column": expiry_col or "Expiry Date",
                },
            }

            sub_tables = [capacity_table, detail_table]

            tab_data = {
                "name": "Capacity + Expiry",
                "sub_tables": sub_tables,
                "chart_data": [],
                "capacity_counts": capacity_counts,
                "capacity_chart": capacity_chart,
                "expiry_counts": expiry_counts,
            }

            html = render_to_string(
                "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
                {"tab": tab_data, "selected_month": selected_month},
            )

            return {
                "detail_html": html,
                "sub_tables": sub_tables,
                "chart_data": [],
                "count": len(df),
                "tab_data": tab_data,
                "hit_pct": 0,
                "target_pct": 100,
            }

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return {
                "detail_html": f"<p class='text-danger'>⚠️ Error: {e}</p>",
                "sub_tables": [],
                "chart_data": [],
                "count": 0,
                "hit_pct": 0,
                "target_pct": 100,
            }

    # Merge sheets from Excel
    def filter_pods_update(self, request, selected_month=None, selected_months=None):
        """
        تاب PODs: قراءة من شيت PODs.
        - فلترة بـ W.HNAME، Created on، PGI Date.
        - حساب الأيام بين Created on و PGI Date (باستثناء يوم الجمعة).
        - Hit = استلام خلال 7 أيام أو أقل، Miss = أكثر من 7 أيام.
        - الجدول العلوي: KPI + صفوف المدن لكل شهر.
        - الجدول السفلي: التفاصيل مع فلتر W.HNAME، الشهور، Hit/Miss؛ البادجات على المدن و Hit/Miss.
        """
        import pandas as pd
        from django.template.loader import render_to_string
        import os
        from datetime import datetime, timedelta

        try:
            excel_path = self.resolve_excel_file_path(request)
            if not excel_path:
                return {"error": "⚠️ Excel file not found."}

            xls = pd.ExcelFile(excel_path, engine="openpyxl")
            sheet_name = next(
                (
                    s
                    for s in xls.sheet_names
                    if "pod" in s.lower() and "update" not in s.lower()
                ),
                None,
            )
            if not sheet_name:
                return {"error": "⚠️ Sheet 'PODs' was not found."}

            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
            df = pd.read_excel(
                excel_path,
                sheet_name=sheet_name,
                engine="openpyxl",
                dtype=str,
                header=0,
                **_nr_kw,
            ).fillna("")
            df.columns = df.columns.astype(str).str.strip()

            def _norm(val):
                return re.sub(r"[^a-z0-9]", "", str(val).strip().lower())

            def _find_col(dframe, names):
                nmap = {_norm(c): c for c in dframe.columns}
                for name in names:
                    n = _norm(name)
                    if n in nmap:
                        return nmap[n]
                for col in dframe.columns:
                    if any(_norm(n) in _norm(col) for n in names):
                        return col
                return None

            col_created = _find_col(df, ["created on", "createdon", "created"])
            col_pgi = _find_col(df, ["pgi date", "pgidate", "pgi"])
            col_whname = _find_col(
                df, ["w.hname", "whname", "warehouse name", "warehouse"]
            )
            col_shpng = _find_col(df, ["shpng pnt", "shpngpnt", "shipping point"])
            col_plant = _find_col(df, ["plant"])
            col_whno = _find_col(df, ["wh no", "whno", "warehouse no"])
            col_delivery = _find_col(df, ["delivery"])
            col_inv = _find_col(df, ["inv", "invoice"])
            col_shipto = _find_col(df, ["ship-to party", "shiptoparty", "ship to"])
            col_shipto_name = _find_col(
                df, ["name of the ship-to party", "ship-to party name"]
            )
            col_qty = _find_col(df, ["qty", "quantity"])
            col_unit = _find_col(df, ["unit"])
            col_city = _find_col(df, ["city"])

            if not col_created or not col_pgi or not col_whname:
                return {
                    "error": "⚠️ Required columns (Created on, PGI Date, W.HNAME) not found."
                }

            # تحويل التواريخ
            df["_created_dt"] = _excel_dates_to_datetime(df[col_created], errors="coerce")
            df["_pgi_dt"] = _excel_dates_to_datetime(df[col_pgi], errors="coerce")

            # حساب Days (باستثناء يوم الجمعة) - الفرق بين Created on و PGI Date
            def business_days_between(start, end):
                if pd.isna(start) or pd.isna(end):
                    return None
                if start > end:
                    return None
                days = 0
                current = start.date()
                end_date = end.date()
                # نحسب الأيام بدون الجمعة (من Created on إلى PGI Date)
                while current < end_date:  # < وليس <= لأننا لا نحسب اليوم الأخير
                    if current.weekday() != 4:  # 4 = Friday
                        days += 1
                    current += timedelta(days=1)
                return days

            df["Days"] = df.apply(
                lambda row: business_days_between(row["_created_dt"], row["_pgi_dt"]),
                axis=1,
            )
            # Hit = استلام خلال 7 أيام أو أقل (بدون الجمعة)، Miss = أكثر من 7 أيام
            df["Hit or Miss"] = df["Days"].apply(
                lambda d: (
                    "Hit"
                    if d is not None and d <= 7
                    else ("Miss" if d is not None else "Pending")
                )
            )
            df["Days"] = df["Days"].apply(
                lambda d: str(int(d)) if d is not None else ""
            )

            # استخراج الشهر من Created on (نحتفظ بـ _created_dt و _pgi_dt لجدول التفاصيل)
            df["Month"] = df["_created_dt"].dt.strftime("%b").fillna("")

            # فلترة الشهر
            selected_months_norm = []
            if selected_months:
                if isinstance(selected_months, str):
                    selected_months = [selected_months]
                seen = set()
                for m in selected_months:
                    norm = self.normalize_month_label(m)
                    if norm and norm.lower() not in seen:
                        seen.add(norm.lower())
                        selected_months_norm.append(norm)

            if selected_months_norm:
                df = df[
                    df["Month"]
                    .str.lower()
                    .isin([m.lower() for m in selected_months_norm])
                ]
            elif selected_month:
                month_norm = self.normalize_month_label(selected_month)
                if month_norm:
                    df = df[df["Month"].str.lower() == month_norm.lower()]

            if df.empty:
                return {
                    "detail_html": "<p class='text-warning text-center p-4'>⚠️ No data available for selected period.</p>",
                    "count": 0,
                    "hit_pct": 0,
                }

            # إحصائيات عامة (لكروت KPI): عدد الشحنات الكل = Hit + Miss فقط
            hit_count = len(df[df["Hit or Miss"] == "Hit"])
            miss_count = len(df[df["Hit or Miss"] == "Miss"])
            total_shipments = hit_count + miss_count
            hit_pct = (
                round((hit_count / total_shipments * 100), 2)
                if total_shipments > 0
                else 0
            )
            miss_pct = (
                round((miss_count / total_shipments * 100), 2)
                if total_shipments > 0
                else 0
            )

            # تجميع كل المدن مع بعض (جدول واحد وشارت واحد)
            month_order = [
                "Jan",
                "Feb",
                "Mar",
                "Apr",
                "May",
                "Jun",
                "Jul",
                "Aug",
                "Sep",
                "Oct",
                "Nov",
                "Dec",
            ]
            months_raw = df["Month"].dropna().unique().tolist()
            months = sorted(
                months_raw,
                key=lambda m: month_order.index(m) if m in month_order else 999,
            )

            # الحصول على قائمة المدن
            cities = sorted(
                df[col_whname]
                .astype(str)
                .str.strip()
                .replace("", pd.NA)
                .dropna()
                .unique()
                .tolist()
            )

            if not months:
                return {
                    "detail_html": "<p class='text-warning text-center p-4'>⚠️ No months found in data.</p>",
                    "count": 0,
                    "hit_pct": 0,
                }

            # تجميع حسب المدينة والشهر: Hit (Closed), Miss (Pending), Total
            # لكل مدينة: Closed, Pending, Total
            city_data = {}
            for city in cities:
                df_city = df[df[col_whname].astype(str).str.strip() == city].copy()
                if df_city.empty:
                    continue

                closed_by_month_city = []
                pending_by_month_city = []
                total_by_month_city = []

                for month in months:
                    df_month_city = df_city[df_city["Month"] == month]
                    hit_month = len(
                        df_month_city[df_month_city["Hit or Miss"] == "Hit"]
                    )
                    miss_month = len(
                        df_month_city[df_month_city["Hit or Miss"] == "Miss"]
                    )
                    closed_by_month_city.append(hit_month)
                    pending_by_month_city.append(miss_month)
                    total_by_month_city.append(hit_month + miss_month)

                # YTD لكل مدينة
                closed_ytd_city = sum(closed_by_month_city)
                pending_ytd_city = sum(pending_by_month_city)
                total_ytd_city = sum(total_by_month_city)

                city_data[city] = {
                    "closed": closed_by_month_city + [closed_ytd_city],
                    "pending": pending_by_month_city + [pending_ytd_city],
                    "total": total_by_month_city + [total_ytd_city],
                }

            # تجميع حسب الشهر: Hit (Closed), Miss (Pending), Total (كل المدن مجمعة)
            closed_by_month = []
            pending_by_month = []
            total_by_month = []

            for month in months:
                df_month = df[df["Month"] == month]
                hit_month = len(df_month[df_month["Hit or Miss"] == "Hit"])
                miss_month = len(df_month[df_month["Hit or Miss"] == "Miss"])
                closed_by_month.append(hit_month)
                pending_by_month.append(miss_month)
                total_by_month.append(hit_month + miss_month)

            # إضافة YTD
            closed_ytd = sum(closed_by_month)
            pending_ytd = sum(pending_by_month)
            total_ytd = sum(total_by_month)

            months_display = months + ["YTD"]
            closed_by_month.append(closed_ytd)
            pending_by_month.append(pending_ytd)
            total_by_month.append(total_ytd)

            # حساب النسب المئوية
            closed_pct = [
                round((c / t * 100), 2) if t > 0 else 0
                for c, t in zip(closed_by_month, total_by_month)
            ]

            # بناء الجدول: KPI، ثم أعمدة المدن جانب أعمدة الشهور، ثم الشهور + YTD
            # الأعمدة: KPI, City1, City2, ..., Jan, Feb, ..., YTD
            columns = ["KPI"] + cities + months_display
            table_rows = []

            # صف Closed (الإجمالي): لكل مدينة نعرض YTD، ثم لكل شهر نعرض الإجمالي
            closed_row = {"KPI": "Closed"}
            for city in cities:
                closed_row[city] = int(city_data.get(city, {}).get("closed", [0])[-1])
            for i, month in enumerate(months_display):
                closed_row[month] = int(closed_by_month[i])
            table_rows.append(closed_row)

            # صف Pending (الإجمالي)
            pending_row = {"KPI": "Pending"}
            for city in cities:
                pending_row[city] = int(city_data.get(city, {}).get("pending", [0])[-1])
            for i, month in enumerate(months_display):
                pending_row[month] = int(pending_by_month[i])
            table_rows.append(pending_row)

            # صف Total (الإجمالي)
            total_row = {"KPI": "Total"}
            for city in cities:
                total_row[city] = int(city_data.get(city, {}).get("total", [0])[-1])
            for i, month in enumerate(months_display):
                total_row[month] = int(total_by_month[i])
            table_rows.append(total_row)

            # شارت واحد (كل المدن مجمعة)
            chart_data = [
                {
                    "type": "column",
                    "name": "Closed %",
                    "color": "#9fc0e4",
                    "showInLegend": True,
                    "indexLabel": "{y}%",
                    "related_table": "PODs YTD",
                    "dataPoints": [
                        {"label": m, "y": closed_pct[i]}
                        for i, m in enumerate(months_display)
                    ],
                },
                {
                    "type": "line",
                    "name": "Target 100%",
                    "color": "red",
                    "showInLegend": True,
                    "related_table": "PODs YTD",
                    "dataPoints": [{"label": m, "y": 100} for m in months_display],
                },
            ]

            sub_tables = [
                {
                    "id": "sub-table-pods-ytd",
                    "title": "PODs YTD",
                    "columns": columns,
                    "data": table_rows,
                    "chart_data": chart_data,
                }
            ]

            # ✅ بناء جدول التفاصيل الكامل (مثل Outbound و Inbound)
            detail_columns = [
                col_shpng if col_shpng else "Shpng Pnt",
                col_whname if col_whname else "W.HNAME",
                col_plant if col_plant else "PLANT",
                col_whno if col_whno else "WH No",
                col_created if col_created else "Created on",
                col_pgi if col_pgi else "PGI Date",
                col_delivery if col_delivery else "Delivery",
                col_inv if col_inv else "INV",
                col_shipto if col_shipto else "Ship-to party",
                col_shipto_name if col_shipto_name else "Name of the ship-to party",
                col_qty if col_qty else "QTY",
                col_unit if col_unit else "Unit",
                col_city if col_city else "City",
                "Days",
                "Hit or Miss",
                "Month",
            ]

            # تنظيف الأعمدة (إزالة None)
            detail_columns = [c for c in detail_columns if c]

            # إعداد البيانات للجدول التفصيلي
            detail_df = df.copy()

            # حفظ عمود الترتيب قبل التحويل
            if "_created_dt" in detail_df.columns:
                detail_df["_sort_ts"] = detail_df["_created_dt"]

            def _fmt_date(x):
                if pd.isna(x) or x is pd.NaT:
                    return ""
                try:
                    return pd.Timestamp(x).strftime("%Y-%m-%d %H:%M")
                except Exception:
                    return ""

            # تحويل التواريخ إلى نص
            if col_created in detail_df.columns and "_created_dt" in detail_df.columns:
                detail_df[col_created] = detail_df["_created_dt"].apply(_fmt_date)
            if col_pgi in detail_df.columns and "_pgi_dt" in detail_df.columns:
                detail_df[col_pgi] = detail_df["_pgi_dt"].apply(_fmt_date)

            # ترتيب البيانات قبل إزالة الأعمدة المؤقتة
            if "_sort_ts" in detail_df.columns:
                detail_df = detail_df.sort_values("_sort_ts", ascending=False)

            # إزالة الأعمدة المؤقتة
            drop_cols = ["_created_dt", "_pgi_dt", "_sort_ts"]
            detail_df = detail_df.drop(
                columns=[c for c in drop_cols if c in detail_df.columns],
                errors="ignore",
            )

            # استخراج البيانات
            detail_rows_raw = detail_df.head(500)[detail_columns].to_dict(
                orient="records"
            )

            def _to_blank(val):
                if val is None:
                    return ""
                if isinstance(val, float) and (pd.isna(val) or (val != val)):
                    return ""
                s = str(val).strip()
                if s.lower() in ("nan", "nat", "none", "<nat>"):
                    return ""
                return s

            detail_rows = [
                {k: _to_blank(v) for k, v in row.items()} for row in detail_rows_raw
            ]

            # بناء قائمة الفلاتر
            detail_df_for_options = detail_df.copy()
            whname_options = (
                sorted(
                    detail_df_for_options[col_whname]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                    .replace("", None)
                    .dropna()
                    .unique()
                    .tolist()
                )
                if col_whname in detail_df_for_options.columns
                else []
            )

            city_options = (
                sorted(
                    detail_df_for_options[col_city]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                    .replace("", None)
                    .dropna()
                    .unique()
                    .tolist()
                )
                if col_city in detail_df_for_options.columns
                else []
            )

            status_options = (
                sorted(
                    detail_df_for_options["Hit or Miss"]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                    .replace("", None)
                    .dropna()
                    .unique()
                    .tolist()
                )
                if "Hit or Miss" in detail_df_for_options.columns
                else []
            )

            month_options = (
                sorted(
                    detail_df_for_options["Month"]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                    .replace("", None)
                    .dropna()
                    .unique()
                    .tolist()
                )
                if "Month" in detail_df_for_options.columns
                else []
            )

            # إضافة جدول التفاصيل
            detail_table = {
                "id": "sub-table-pods-detail",
                "title": "PODs Shipments Detail",
                "columns": detail_columns,
                "data": detail_rows,
                "chart_data": [],
                "full_width": True,
                "filter_options": {
                    "whnames": whname_options,
                    "statuses": status_options,
                    "months": month_options,
                },
            }

            sub_tables.append(detail_table)

            # كروت KPI
            stats = {
                "total_shipments": total_shipments,
                "hit_pct": hit_pct,
                "miss_pct": miss_pct,
                "target": 100,
            }

            tab_data = {
                "name": "PODs Update",
                "sub_tables": sub_tables,
                "chart_data": chart_data,
                "chart_title": "PODs Closed % Performance",
                "hit_pct": hit_pct,
                "target_pct": 100,
                "stats": stats,
            }

            month_norm_tab = self.apply_month_filter_to_tab(
                tab_data, selected_month, selected_months_norm or None
            )
            html = render_to_string(
                "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
                {"tab": tab_data, "selected_month": month_norm_tab},
            )

            return {
                "detail_html": html,
                "chart_data": chart_data,
                "chart_title": "PODs Closed % Performance",
                "hit_pct": hit_pct,
                "target_pct": 100,
                "count": total_shipments,
                "tab_data": tab_data,
            }

        except Exception as e:
            import traceback

            print(traceback.format_exc())
            return {"error": f"⚠️ Error processing PODs: {e}"}

    def filter_rejections_combined(
        self, request, selected_month=None, selected_months=None
    ):
        """
        تاب Return & Refusal: عرض جدول Return فقط من شيت Inbound (Shipment Type = RMA).
        بدون Rejection / Rejection breakdown / شارت — جدول فقط بعرض الصفحة.
        """
        import pandas as pd
        import os
        from django.template.loader import render_to_string

        try:
            excel_path = self.resolve_excel_file_path(request)
            if not excel_path:
                return {
                    "detail_html": "<p class='text-danger'>⚠️ Excel file not found.</p>",
                    "chart_data": [],
                    "count": 0,
                }

            xls = pd.ExcelFile(excel_path, engine="openpyxl")
            sheet_names = [s.strip() for s in xls.sheet_names]
            sub_tables = []
            chart_data = []

            selected_months_norm = []
            if selected_months:
                if isinstance(selected_months, str):
                    selected_months = [selected_months]
                seen = set()
                for m in selected_months:
                    norm = self.normalize_month_label(m)
                    if norm and norm not in seen:
                        seen.add(norm)
                        selected_months_norm.append(norm)

            # ✅ جدول Return فقط من شيت Inbound: فلتر Shipment Type = RMA
            return_columns_display = [
                "Shipment Nbr",
                "Shipment Type",
                "Status",
                "Create Timestamp",
                "Arrival Date",
                "Offloading Date",
                "Last LPN Rcv TS",
            ]

            def _normalize_col(val):
                return re.sub(r"[^a-z0-9]", "", str(val).strip().lower())

            def _find_col(df, possible_names):
                norm_map = {_normalize_col(c): c for c in df.columns}
                for name in possible_names:
                    n = _normalize_col(name)
                    if n in norm_map:
                        return norm_map[n]
                for col in df.columns:
                    if any(
                        _normalize_col(name) in _normalize_col(col)
                        for name in possible_names
                    ):
                        return col
                return None

            inbound_sheet = next(
                (s for s in sheet_names if "inbound" in s.lower()), None
            )
            if inbound_sheet:
                try:
                    _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
                    df_in = pd.read_excel(
                        excel_path,
                        sheet_name=inbound_sheet,
                        engine="openpyxl",
                        dtype=str,
                        header=0,
                        **_nr_kw,
                    ).fillna("")
                    df_in.columns = df_in.columns.astype(str).str.strip()

                    col_ship_nbr = _find_col(
                        df_in, ["shipment nbr", "shipment number", "shipment no"]
                    )
                    col_ship_type = _find_col(
                        df_in, ["shipment type", "shipmenttype", "type"]
                    )
                    col_status = _find_col(df_in, ["status", "shipment status"])
                    col_create = _find_col(
                        df_in, ["create timestamp", "created timestamp"]
                    )
                    col_arrival = _find_col(
                        df_in, ["arrival date", "arrival timestamp"]
                    )
                    col_offload = _find_col(df_in, ["offloading date", "offload date"])
                    col_last_lpn = _find_col(
                        df_in, ["last lpn rcv ts", "last lpn receive ts"]
                    )

                    if col_ship_type is not None:
                        df_in = df_in[
                            df_in[col_ship_type].astype(str).str.strip().str.upper()
                            == "RMA"
                        ]
                    else:
                        df_in = df_in.iloc[0:0]

                    if not df_in.empty and all(
                        [
                            col_ship_nbr,
                            col_status,
                            col_create,
                            col_arrival,
                            col_offload,
                            col_last_lpn,
                        ]
                    ):
                        rename = {
                            col_ship_nbr: "Shipment Nbr",
                            col_status: "Status",
                            col_create: "Create Timestamp",
                            col_arrival: "Arrival Date",
                            col_offload: "Offloading Date",
                            col_last_lpn: "Last LPN Rcv TS",
                        }
                        if col_ship_type is not None:
                            rename[col_ship_type] = "Shipment Type"
                        df_in = df_in.rename(columns=rename)

                        for c in return_columns_display:
                            if c not in df_in.columns:
                                df_in[c] = ""

                        df_in = df_in[return_columns_display]
                        if selected_month or selected_months_norm:
                            month_col = _find_col(df_in, ["month", "create timestamp"])
                            if month_col and month_col in df_in.columns:
                                if selected_months_norm:
                                    active = {
                                        self.normalize_month_label(m)
                                        for m in selected_months_norm
                                    }
                                else:
                                    active = {
                                        self.normalize_month_label(selected_month)
                                    }
                                if "Create Timestamp" in df_in.columns:
                                    try:
                                        ts = _excel_dates_to_datetime(
                                            df_in["Create Timestamp"],
                                            errors="coerce",
                                        )
                                        df_in["_month"] = ts.dt.strftime("%b")
                                        df_in = df_in[
                                            df_in["_month"]
                                            .fillna("")
                                            .str.lower()
                                            .isin([m.lower() for m in active])
                                        ]
                                        df_in = df_in.drop(
                                            columns=["_month"], errors="ignore"
                                        )
                                    except Exception:
                                        pass

                        # حساب Hit/Miss للـ Return (≤24h بين Create Timestamp و Last LPN Rcv TS)
                        return_kpi = None
                        try:
                            ts_create = _excel_dates_to_datetime(
                                df_in["Create Timestamp"], errors="coerce"
                            )
                            ts_last = _excel_dates_to_datetime(
                                df_in["Last LPN Rcv TS"], errors="coerce"
                            )
                            hours = (ts_last - ts_create).dt.total_seconds() / 3600.0
                            df_in["_is_hit"] = (hours <= 24) & (hours.notna())
                            total_ret = len(df_in)
                            successful_ret = int(df_in["_is_hit"].sum())
                            failed_ret = total_ret - successful_ret
                            hit_pct_ret = (
                                round(100.0 * successful_ret / total_ret, 2)
                                if total_ret else 0
                            )
                            return_kpi = {
                                "total_shipments": total_ret,
                                "successful": successful_ret,
                                "failed": failed_ret,
                                "target": 99,
                                "hit_pct": hit_pct_ret,
                            }
                            df_in = df_in.drop(columns=["_is_hit"], errors="ignore")
                        except Exception:
                            total_ret = len(df_in)
                            return_kpi = {
                                "total_shipments": total_ret,
                                "successful": total_ret,
                                "failed": 0,
                                "target": 99,
                                "hit_pct": 100.0 if total_ret else 0,
                            }

                        sub_tables.append(
                            {
                                "title": "Return",
                                "columns": return_columns_display,
                                "data": df_in.to_dict(orient="records"),
                                "return_kpi": return_kpi,
                            }
                        )
                    else:
                        sub_tables.append(
                            {
                                "title": "Return",
                                "columns": return_columns_display,
                                "data": [],
                                "error": (
                                    "Inbound sheet missing required columns or no RMA rows."
                                    if col_ship_type is not None
                                    else "Column 'Shipment Type' not found in Inbound."
                                ),
                            }
                        )
                except Exception as e_in:
                    import traceback

                    print(traceback.format_exc())
                    sub_tables.append(
                        {
                            "title": "Return",
                            "columns": return_columns_display,
                            "data": [],
                            "error": str(e_in),
                        }
                    )
            else:
                sub_tables.append(
                    {
                        "title": "Return",
                        "columns": return_columns_display,
                        "data": [],
                        "error": "Sheet containing 'Inbound' was not found.",
                    }
                )

            # ✅ التحقق من وجود بيانات بعد الفلترة
            total_count = sum(len(st["data"]) for st in sub_tables)
            if (selected_month or selected_months_norm) and total_count == 0:
                if selected_months_norm:
                    msg = ", ".join(selected_months_norm)
                else:
                    msg = str(selected_month).strip().capitalize()
                return {
                    "detail_html": f"<p class='text-warning text-center p-4'>⚠️ No data available for {msg} in Return & Refusal.</p>",
                    "chart_data": [],
                    "count": 0,
                }

            # 🧩 بناء الـ HTML — نمرّر return_kpi من أول sub_table للكروت فوق الجدول
            return_kpi_for_tab = None
            if sub_tables:
                return_kpi_for_tab = sub_tables[0].get("return_kpi")
            tab_data = {
                "name": "Return & Refusal",
                "sub_tables": sub_tables,
                "chart_data": chart_data,
                "chart_title": "Return & Refusal Overview",
                "return_kpi": return_kpi_for_tab,
            }
            month_norm_tab = self.apply_month_filter_to_tab(
                tab_data,
                selected_month if not selected_months_norm else None,
                selected_months_norm or None,
            )
            html = render_to_string(
                "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
                {"tab": tab_data, "selected_month": month_norm_tab},
            )

            # 🧮 حساب hit% من متوسط القيم في % of Rejection (بالنسبة المئوية)
            hit_values = []
            for st in sub_tables:
                if "rejection" in st["title"].lower():
                    for row in st["data"]:
                        val = row.get("% of Rejection", "")
                        try:
                            num = to_percentage_number(val)
                            if num is not None:
                                hit_values.append(num)
                        except:
                            pass

            hit_pct = round(sum(hit_values) / len(hit_values), 2) if hit_values else 0

            result = {
                "detail_html": html,
                "chart_data": chart_data,
                "chart_title": "Return & Refusal Overview",
                "count": total_count,
                "hit_pct": hit_pct,
                "tab_data": tab_data,
            }

            return result

        except Exception as e:
            import traceback

            print(traceback.format_exc())
            return {
                "detail_html": f"<p class='text-danger'>⚠️ Error while processing Return & Refusal data: {e}</p>",
                "chart_data": [],
                "count": 0,
            }

    def filter_expiry(self, request, selected_month=None, selected_months=None):
        """
        تاب Expiry: قراءة من شيت Expiry.
        - فلتر Status: Located, Allocated, Partly Allocated فقط.
        - أعمدة: Facility, Company, LPN Nbr, Status, Item Code, Item Description, Current Qty, batch_nbr, Expiry Date.
        - تحذير: اللي ينتهي خلال 3 شهور = قريب، خلال 6 شهور = warning، يعرض تحت الجدول في Bootstrap 5 alert.
        """
        import pandas as pd
        import os
        from datetime import datetime, timedelta
        from django.template.loader import render_to_string

        try:
            excel_path = self.resolve_excel_file_path(request)
            if not excel_path:
                return {
                    "detail_html": "<p class='text-danger'>⚠️ Excel file not found.</p>",
                    "chart_data": [],
                    "count": 0,
                }

            xls = pd.ExcelFile(excel_path, engine="openpyxl")
            sheet_names = [s.strip() for s in xls.sheet_names]
            expiry_sheet = next((s for s in sheet_names if "expiry" in s.lower()), None)
            if not expiry_sheet:
                return {
                    "detail_html": "<p class='text-warning'>⚠️ Sheet containing 'Expiry' was not found.</p>",
                    "chart_data": [],
                    "count": 0,
                }

            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))
            df = pd.read_excel(
                excel_path,
                sheet_name=expiry_sheet,
                engine="openpyxl",
                dtype=str,
                header=0,
                **_nr_kw,
            ).fillna("")
            df.columns = df.columns.astype(str).str.strip()

            def _norm(val):
                return re.sub(r"[^a-z0-9]", "", str(val).strip().lower())

            def _find_col(dframe, names):
                nmap = {_norm(c): c for c in dframe.columns}
                for name in names:
                    n = _norm(name)
                    if n in nmap:
                        return nmap[n]
                for col in dframe.columns:
                    if any(_norm(n) in _norm(col) for n in names):
                        return col
                return None

            col_facility = _find_col(df, ["facility", "facility code"])
            col_company = _find_col(df, ["company"])
            col_lpn = _find_col(df, ["lpn nbr", "lpn", "lpn nbr"])
            col_status = _find_col(df, ["status"])
            col_item_code = _find_col(df, ["item code", "itemcode"])
            col_item_desc = _find_col(df, ["item description", "item desc"])
            col_qty = _find_col(df, ["current qty", "currentqty", "qty"])
            col_batch = _find_col(df, ["batch_nbr", "batch nbr", "batch"])
            col_expiry = _find_col(df, ["expiry date", "expirydate", "expiry"])

            if not col_status:
                return {
                    "detail_html": "<p class='text-danger'>⚠️ Column 'Status' not found in Expiry sheet.</p>",
                    "chart_data": [],
                    "count": 0,
                }

            # فلتر Status: Located, Allocated, Partly Allocated
            status_vals = {"located", "allocated", "partly allocated"}
            df = df[
                df[col_status].astype(str).str.strip().str.lower().isin(status_vals)
            ]

            display_columns = [
                "Facility",
                "Company",
                "LPN Nbr",
                "Status",
                "Item Code",
                "Item Description",
                "Current Qty",
                "batch_nbr",
                "Expiry Date",
            ]
            rename_map = {}
            if col_facility:
                rename_map[col_facility] = "Facility"
            if col_company:
                rename_map[col_company] = "Company"
            if col_lpn:
                rename_map[col_lpn] = "LPN Nbr"
            if col_status:
                rename_map[col_status] = "Status"
            if col_item_code:
                rename_map[col_item_code] = "Item Code"
            if col_item_desc:
                rename_map[col_item_desc] = "Item Description"
            if col_qty:
                rename_map[col_qty] = "Current Qty"
            if col_batch:
                rename_map[col_batch] = "batch_nbr"
            if col_expiry:
                rename_map[col_expiry] = "Expiry Date"

            df = df.rename(columns=rename_map)
            for c in display_columns:
                if c not in df.columns:
                    df[c] = ""

            df = df[display_columns]

            # تحويل Expiry Date وتحديد نطاقات: 1–3، 3–6، 6–9 شهور
            today = pd.Timestamp(datetime.now().date())
            three_months = today + pd.DateOffset(months=3)
            six_months = today + pd.DateOffset(months=6)
            nine_months = today + pd.DateOffset(months=9)

            expiry_ser = _excel_dates_to_datetime(df["Expiry Date"], errors="coerce")
            df["_expiry_dt"] = expiry_ser
            df["Expiry Date"] = expiry_ser.dt.strftime("%Y-%m-%d").fillna("")

            within_1_3 = (
                (df["_expiry_dt"].notna())
                & (df["_expiry_dt"] >= today)
                & (df["_expiry_dt"] <= three_months)
            )
            within_3_6 = (
                (df["_expiry_dt"].notna())
                & (df["_expiry_dt"] > three_months)
                & (df["_expiry_dt"] <= six_months)
            )
            within_6_9 = (
                (df["_expiry_dt"].notna())
                & (df["_expiry_dt"] > six_months)
                & (df["_expiry_dt"] <= nine_months)
            )
            df = df.drop(columns=["_expiry_dt"], errors="ignore")

            table_data = df[display_columns].to_dict(orient="records")

            # أعداد المنتجات لكل نطاق
            expiry_counts = {
                "within_1_3": int(within_1_3.sum()),
                "within_3_6": int(within_3_6.sum()),
                "within_6_9": int(within_6_9.sum()),
            }

            # خيارات الفلاتر: Facility, Company, Status, Expiry Date
            facility_codes = (
                sorted(
                    df["Facility"]
                    .astype(str)
                    .str.strip()
                    .replace("", pd.NA)
                    .dropna()
                    .unique()
                    .tolist()
                )
                if "Facility" in df.columns
                else []
            )
            companies = (
                sorted(
                    df["Company"]
                    .astype(str)
                    .str.strip()
                    .replace("", pd.NA)
                    .dropna()
                    .unique()
                    .tolist()
                )
                if "Company" in df.columns
                else []
            )
            statuses = (
                sorted(
                    df["Status"]
                    .astype(str)
                    .str.strip()
                    .replace("", pd.NA)
                    .dropna()
                    .unique()
                    .tolist()
                )
                if "Status" in df.columns
                else []
            )
            expiry_dates = (
                sorted(
                    df["Expiry Date"]
                    .astype(str)
                    .str.strip()
                    .replace("", pd.NA)
                    .dropna()
                    .unique()
                    .tolist()
                )
                if "Expiry Date" in df.columns
                else []
            )

            filter_options = {
                "facility_codes": facility_codes,
                "companies": companies,
                "statuses": statuses,
                "expiry_dates": expiry_dates,
            }

            sub_tables = [
                {
                    "id": "sub-table-expiry-detail",
                    "title": "Expiry",
                    "columns": display_columns,
                    "data": table_data,
                    "filter_options": filter_options,
                }
            ]
            tab_data = {
                "name": "Expiry",
                "sub_tables": sub_tables,
                "chart_data": [],
                "expiry_counts": expiry_counts,
            }
            month_norm = self.apply_month_filter_to_tab(tab_data, selected_month, None)
            html = render_to_string(
                "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
                {"tab": tab_data, "selected_month": month_norm},
            )
            return {
                "detail_html": html,
                "chart_data": [],
                "count": len(table_data),
                "tab_data": tab_data,
            }
        except Exception as e:
            import traceback

            print(traceback.format_exc())
            return {
                "detail_html": f"<p class='text-danger'>⚠️ Error processing Expiry: {e}</p>",
                "chart_data": [],
                "count": 0,
            }

    def filter_total_lead_time_performance(
        self, request, selected_month=None, selected_months=None
    ):
        """
        🔹 عرض جدول Miss Breakdown (3PL و Roche كل واحد منفصل)
        🔹 عرض الشارت الخاص بـ 3PL On-Time Delivery
        🔹 عرض خطوات Outbound في الأسفل
        🔹 تم إضافة كاش على مستوى الدالة لتسريع التحميل لأول مرة
        """
        try:
            excel_path = self.resolve_excel_file_path(request)
            if not excel_path:
                return {
                    "detail_html": "<p class='text-danger text-center'>⚠️ Excel file not found for display.</p>",
                    "chart_data": [],
                    "count": 0,
                }

            # ✅ كاش على مستوى Total Lead Time Performance (يعتمد على مسار الملف والشهور المختارة)
            import hashlib

            _path_hash = hashlib.md5((excel_path or "").encode()).hexdigest()[:12]
            _month_part = (
                str(selected_month).strip() if selected_month is not None else ""
            )
            _months_list = (
                ",".join(map(str, selected_months))
                if selected_months is not None
                else ""
            )
            _full_flag = "1" if _excel_full_data_requested(request) else "0"
            # v2: chart_data includes outbound POD charts for All-in-One overview
            _cache_key = f"tlp_total_lead_time_{_path_hash}_{_month_part}_{_months_list}_{_full_flag}_v2"
            cached_result = cache.get(_cache_key)
            if cached_result is not None:
                return cached_result

            xls = pd.ExcelFile(excel_path, engine="openpyxl")
            sub_tables = []
            chart_data = []
            selected_month_norm = None
            selected_months_norm = []
            actual_target = 0  # يُحدَّث من الشيت الرئيسي إن وُجد

            if selected_month:
                raw_month = str(selected_month).strip()
                parsed = _excel_dates_to_datetime(raw_month, errors="coerce")
                if pd.isna(parsed):
                    selected_month_norm = raw_month[:3].capitalize()
                else:
                    selected_month_norm = parsed.strftime("%b")

            if selected_months:
                if isinstance(selected_months, str):
                    selected_months = [selected_months]
                for m in selected_months:
                    norm = self.normalize_month_label(m)
                    if norm:
                        selected_months_norm.append(norm)
                # إزالة التكرارات مع الحفاظ على الترتيب
                seen = set()
                selected_months_norm = [
                    m for m in selected_months_norm if not (m in seen or seen.add(m))
                ]

            # ----------------------------
            # 🟦 جدول 3PL SIDE
            # ----------------------------
            sheet_3pl = next(
                (
                    s
                    for s in xls.sheet_names
                    if "total lead time preformance" in s.lower()
                    and "-r" not in s.lower()
                ),
                None,
            )

            final_df_3pl = None
            _nr_kw = _read_excel_nrows_kw(_excel_max_rows_for_request(request))

            if sheet_3pl:
                df = pd.read_excel(
                    excel_path, sheet_name=sheet_3pl, engine="openpyxl", header=0, **_nr_kw
                )
                df.columns = df.columns.str.strip().str.lower()

                required_cols = [
                    "month",
                    "outbound delivery",
                    "kpi",
                    "reason group",
                    "miss reason",
                ]
                if all(col in df.columns for col in required_cols):
                    df["year"] = _excel_dates_to_datetime(df["month"], errors="coerce").dt.year
                    df = df[df["year"] == 2025]

                    if "month" in df.columns:
                        # نحاول تحويل القيم في عمود Month إلى تاريخ، ثم استخراج اسم الشهر المختصر
                        df["month"] = _excel_dates_to_datetime(
                            df["month"], errors="coerce"
                        ).dt.strftime("%b")
                    else:
                        # fallback لو مفيش عمود Month
                        df["month"] = _excel_dates_to_datetime(
                            df["ob distribution date"], errors="coerce"
                        ).dt.strftime("%b")

                    # ترتيب الشهور
                    month_order = [
                        "Jan",
                        "Feb",
                        "Mar",
                        "Apr",
                        "May",
                        "Jun",
                        "Jul",
                        "Aug",
                        "Sep",
                        "Oct",
                        "Nov",
                        "Dec",
                    ]

                    df["month"] = pd.Categorical(
                        df["month"], categories=month_order, ordered=True
                    )
                    missing_months = []
                    if selected_month_norm:
                        df = df[df["month"] == selected_month_norm]
                        if df.empty:
                            result = {
                                "detail_html": f"<p class='text-warning text-center p-4'>⚠️ No data available for month {selected_month_norm} in Total Lead Time Performance.</p>",
                                "chart_data": [],
                                "count": 0,
                                "hit_pct": 0,
                            }
                            cache.set(_cache_key, result, 300)
                            return result
                        existing_months = [selected_month_norm]
                    elif selected_months_norm:
                        df = df[df["month"].isin(selected_months_norm)]
                        available_months = [
                            m
                            for m in selected_months_norm
                            if m in df["month"].dropna().unique()
                        ]
                        missing_months = [
                            m for m in selected_months_norm if m not in available_months
                        ]
                        if df.empty:
                            result = {
                                "detail_html": "<p class='text-warning text-center p-4'>⚠️ No data available for the selected quarter months in Total Lead Time Performance.</p>",
                                "chart_data": [],
                                "count": 0,
                                "hit_pct": 0,
                            }
                            cache.set(_cache_key, result, 300)
                            return result
                        existing_months = selected_months_norm
                    else:
                        existing_months = [
                            m for m in month_order if m in df["month"].dropna().unique()
                        ]

                    df["reason group"] = (
                        df["reason group"].astype(str).str.strip().str.lower()
                    )
                    df["kpi"] = df["kpi"].astype(str).str.strip().str.lower()
                    df["miss reason"] = (
                        df["miss reason"].astype(str).str.strip().str.lower()
                    )

                    df_hit = df[df["kpi"].str.lower() == "hit"].copy()
                    hit_counts = (
                        df_hit.groupby("month")["outbound delivery"]
                        .nunique()
                        .reindex(existing_months, fill_value=0)
                    )

                    df_3pl_miss = df[
                        (df["kpi"].str.lower() == "miss")
                        & (df["reason group"] == "3pl")
                    ].copy()

                    miss_grouped = (
                        df_3pl_miss.groupby(["miss reason", "month"])[
                            "outbound delivery"
                        ]
                        .nunique()
                        .reset_index(name="count")
                        .pivot_table(
                            index="miss reason",
                            columns="month",
                            values="count",
                            fill_value=0,
                        )
                    )

                    for m in existing_months:
                        if m not in miss_grouped.columns:
                            miss_grouped[m] = 0
                    miss_grouped = miss_grouped[existing_months]

                    final_df_3pl = miss_grouped.copy()
                    final_df_3pl.loc["on time delivery"] = hit_counts
                    final_df_3pl = final_df_3pl.fillna(0).astype(int)
                    final_df_3pl["2025"] = final_df_3pl.sum(axis=1)

                    total_row = final_df_3pl.sum(numeric_only=True)
                    total_row.name = "total"
                    final_df_3pl = pd.concat([final_df_3pl, pd.DataFrame([total_row])])

                    final_df_3pl.reset_index(inplace=True)
                    final_df_3pl.rename(columns={"index": "KPI"}, inplace=True)
                    final_df_3pl["KPI"] = final_df_3pl["KPI"].str.title()

                    desired_order = [
                        "On Time Delivery",
                        "Late Arrive To The Customer",
                        "Customer Close On Arrive",
                        "Remote Area",
                    ]
                    final_df_3pl["order_key"] = final_df_3pl["KPI"].apply(
                        lambda x: (
                            desired_order.index(x)
                            if x in desired_order
                            else len(desired_order) + 1
                        )
                    )
                    final_df_3pl = final_df_3pl.sort_values(
                        by=["order_key", "KPI"]
                    ).drop(columns=["order_key"])
                    # final_df_3pl.insert(1, "Reason Group", "3PL")
                    #
                    # # ✅ حذف عمود Reason Group قبل الإرسال
                    # if "Reason Group" in final_df_3pl.columns:
                    #     final_df_3pl = final_df_3pl.drop(columns=["Reason Group"])

                    # ✅ حساب التارجت الفعلي لكل شهر (On Time ÷ Total × 100)
                    percent_hit = []
                    existing_months = [
                        m
                        for m in final_df_3pl.columns
                        if m not in ["KPI", "Reason Group", "2025", "Total"]
                    ]

                    on_time_row = final_df_3pl.loc[
                        final_df_3pl["KPI"].str.lower() == "on time delivery"
                    ].iloc[0]
                    total_row = final_df_3pl.loc[
                        final_df_3pl["KPI"].str.lower() == "total"
                    ].iloc[0]

                    for m in existing_months:
                        on_time_val = float(on_time_row.get(m, 0))
                        total_val = float(total_row.get(m, 0))

                        # ✅ لو الشهر فيه صفر فعلاً، خليه 0 في الشارت كمان
                        if total_val == 0 or on_time_val == 0:
                            percent = 0
                        else:
                            percent = int(round((on_time_val / total_val) * 100))

                        percent_hit.append(percent)

                    try:
                        total_year_val = total_row["2025"]
                        on_time_year_val = on_time_row["2025"]
                        actual_target = (
                            int(round((on_time_year_val / total_year_val) * 100))
                            if total_year_val > 0
                            else 0
                        )
                    except Exception:
                        actual_target = 100

                    # ✅ إنشاء قائمة بالشهور اللي فيها قيم غير صفرية (فقط للشارت)
                    nonzero_months = [
                        m for i, m in enumerate(existing_months) if percent_hit[i] > 0
                    ]
                    nonzero_percents = [
                        percent_hit[i]
                        for i, m in enumerate(existing_months)
                        if percent_hit[i] > 0
                    ]
                    if not nonzero_months:
                        nonzero_months = existing_months
                        nonzero_percents = [
                            percent_hit[i] for i in range(len(existing_months))
                        ]

                    chart_data.append(
                        {
                            "type": "column",
                            "name": "On-Time Delivery (%)",
                            "color": "#9fc0e4",
                            "showInLegend": True,
                            "related_table": "Miss Breakdown – 3PL Side",  # ✅ ربط الشارت بالجدول
                            "dataPoints": [
                                {"label": m, "y": nonzero_percents[i]}
                                for i, m in enumerate(nonzero_months)
                            ],
                        }
                    )
                    chart_data.append(
                        {
                            "type": "line",
                            "name": f"Target ({actual_target}%)",
                            "color": "red",
                            "showInLegend": True,
                            "related_table": "Miss Breakdown – 3PL Side",  # ✅ ربط الشارت بالجدول
                            "dataPoints": [
                                {"label": m, "y": actual_target} for m in nonzero_months
                            ],
                        }
                    )

                    sub_tables.append(
                        {
                            "title": "Miss Breakdown – 3PL Side",
                            "columns": list(final_df_3pl.columns),
                            "data": final_df_3pl.to_dict(orient="records"),
                        }
                    )
                    # لم نعد نضيف جدول Missing Months هنا، يتم التعامل معه لاحقًا عبر apply_month_filter_to_tab

            # ----------------------------
            # 🟥 جدول ROCHE SIDE
            # ----------------------------
            sheet_roche = next(
                (s for s in xls.sheet_names if "preformance -r" in s.lower()), None
            )
            if sheet_roche:
                df = pd.read_excel(
                    excel_path, sheet_name=sheet_roche, engine="openpyxl", header=0, **_nr_kw
                )
                df.columns = df.columns.str.strip()
                if "Month" in df.columns:
                    month_order = [
                        "Jan",
                        "Feb",
                        "Mar",
                        "Apr",
                        "May",
                        "Jun",
                        "Jul",
                        "Aug",
                        "Sep",
                        "Oct",
                        "Nov",
                        "Dec",
                    ]
                    df["Month"] = pd.Categorical(
                        df["Month"], categories=month_order, ordered=True
                    )
                    df = df.sort_values("Month")

                    if selected_month_norm:
                        df_filtered = df[
                            df["Month"].astype(str).str.lower()
                            == selected_month_norm.lower()
                        ]
                        if df_filtered.empty:
                            sub_tables.append(
                                {
                                    "title": "Miss Breakdown – Roche Side",
                                    "columns": [],
                                    "data": [],
                                    "message": f"⚠️ لا توجد بيانات متاحة للشهر {selected_month_norm}.",
                                }
                            )
                        else:
                            df_melted = df_filtered.melt(
                                id_vars=["Month"], var_name="KPI", value_name="Count"
                            )
                            pivot_df = (
                                df_melted.groupby(["KPI", "Month"])["Count"]
                                .sum()
                                .unstack(fill_value=0)
                            )
                            pivot_df["2025"] = pivot_df.sum(axis=1)
                            total_row = pivot_df.sum(numeric_only=True)
                            total_row.name = "TOTAL"
                            pivot_df = pd.concat([pivot_df, pd.DataFrame([total_row])])
                            pivot_df.reset_index(inplace=True)
                            pivot_df.rename(columns={"index": "KPI"}, inplace=True)
                            keep_cols = [
                                col
                                for col in ["KPI", selected_month_norm]
                                if col in pivot_df.columns
                            ]
                            pivot_df = pivot_df[keep_cols]
                            sub_tables.append(
                                {
                                    "title": "Miss Breakdown – Roche Side",
                                    "columns": list(pivot_df.columns),
                                    "data": pivot_df.to_dict(orient="records"),
                                }
                            )
                    elif selected_months_norm:
                        df_filtered = df[
                            df["Month"]
                            .astype(str)
                            .str.lower()
                            .isin([m.lower() for m in selected_months_norm])
                        ]
                        if df_filtered.empty:
                            sub_tables.append(
                                {
                                    "title": "Miss Breakdown – Roche Side",
                                    "columns": [],
                                    "data": [],
                                    "message": "⚠️ No data available for the selected quarter months.",
                                }
                            )
                        else:
                            df_melted = df_filtered.melt(
                                id_vars=["Month"], var_name="KPI", value_name="Count"
                            )
                            pivot_df = (
                                df_melted.groupby(["KPI", "Month"])["Count"]
                                .sum()
                                .unstack(fill_value=0)
                            )
                            ordered_months = [
                                m for m in selected_months_norm if m in pivot_df.columns
                            ]
                            for m in selected_months_norm:
                                if m not in pivot_df.columns:
                                    pivot_df[m] = 0
                            pivot_df = pivot_df[selected_months_norm]
                            pivot_df["2025"] = pivot_df.sum(axis=1)
                            total_row = pivot_df.sum(numeric_only=True)
                            total_row.name = "TOTAL"
                            pivot_df = pd.concat([pivot_df, pd.DataFrame([total_row])])
                            pivot_df.reset_index(inplace=True)
                            pivot_df.rename(columns={"index": "KPI"}, inplace=True)
                            sub_tables.append(
                                {
                                    "title": "Miss Breakdown – Roche Side",
                                    "columns": list(pivot_df.columns),
                                    "data": pivot_df.to_dict(orient="records"),
                                }
                            )
                    else:
                        df_melted = df.melt(
                            id_vars=["Month"], var_name="KPI", value_name="Count"
                        )
                        pivot_df = (
                            df_melted.groupby(["KPI", "Month"])["Count"]
                            .sum()
                            .unstack(fill_value=0)
                            .reindex(columns=month_order, fill_value=0)
                        )
                        pivot_df["2025"] = pivot_df.sum(axis=1)
                        total_row = pivot_df.sum(numeric_only=True)
                        total_row.name = "TOTAL"
                        pivot_df = pd.concat([pivot_df, pd.DataFrame([total_row])])
                        pivot_df.reset_index(inplace=True)
                        pivot_df.rename(columns={"index": "KPI"}, inplace=True)
                        pivot_df = pivot_df.loc[:, (pivot_df != 0).any(axis=0)]

                        sub_tables.append(
                            {
                                "title": "Miss Breakdown – Roche Side",
                                "columns": list(pivot_df.columns),
                                "data": pivot_df.to_dict(orient="records"),
                            }
                        )

            # Outbound Shipments (Outbound1 + Outbound2, Hit/Miss) — نفس فكرة Inbound
            outbound_result = self.filter_outbound_shipments(
                request,
                selected_month if not selected_months_norm else None,
                selected_months_norm if selected_months_norm else None,
            )
            # ✅ جلب نسبة الـ Hit من Outbound (هي اللي هنستخدمها كـ KPI للتاب ده)
            outbound_stats = outbound_result.get("stats", {}) or {}
            outbound_hit_pct = outbound_stats.get("hit_pct", 0) or 0
            # ✅ إذا لم تكن موجودة في stats، نحاول جلبها مباشرة من outbound_result
            if not outbound_hit_pct:
                outbound_hit_pct = outbound_result.get("hit_pct", 0) or 0
            print(
                f"🔍 Total Lead Time Performance - Outbound hit_pct: {outbound_hit_pct}% (from stats: {outbound_stats.get('hit_pct', 'N/A')})"
            )

            if outbound_result.get("sub_tables"):
                outbound_tab = {
                    "name": "B2B Outbound",
                    "stats": outbound_result.get("stats", {}),
                    "sub_tables": outbound_result["sub_tables"],
                    "chart_data": outbound_result.get("chart_data", []),
                    "chart_data_pods": outbound_result.get("chart_data_pods", []),
                    "raw_excel_table": outbound_result.get("raw_excel_table"),
                }
                outbound_html = render_to_string(
                    "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
                    {
                        "tab": outbound_tab,
                        "selected_month": selected_month
                        or (selected_months_norm[0] if selected_months_norm else None),
                    },
                )
            else:
                outbound_html = outbound_result.get("detail_html", "")

            # لا نرجع "لا توجد بيانات" إلا لو مفيش جداول رئيسية ومفيش محتوى Outbound
            has_outbound = bool(outbound_html and str(outbound_html).strip())
            if not sub_tables and not has_outbound:
                return {
                    "detail_html": "<p class='text-muted'>⚠️ No valid data was found in any sheets.</p>",
                    "chart_data": [],
                    "count": 0,
                }

            # ✅ لا نحتاج لتعيين related_table هنا لأنه تم تعيينه بالفعل لكل dataset
            # if chart_data:
            #     for dataset in chart_data:
            #         dataset.setdefault("related_table", "Total Lead Time Performance")

            tab_data = {
                "name": "B2B Outbound",
                "sub_tables": sub_tables,
                "outbound_html": outbound_html,
                "chart_data": chart_data,
            }
            month_norm_tab = self.apply_month_filter_to_tab(
                tab_data,
                selected_month_norm if not selected_months_norm else None,
                selected_months_norm or None,
            )

            html = render_to_string(
                "forms-table/table/bootstrap-table/basic-table/components/excel-sheet-table.html",
                {
                    "tab": tab_data,
                    "selected_month": month_norm_tab,
                    "selected_months": selected_months_norm,
                },
            )

            total_count = sum(len(st["data"]) for st in sub_tables)

            # ✅ نسبة الـ Hit الخاصة بالـ Outbound (هي اللي هنستخدمها كـ KPI للتاب ده)
            try:
                hit_pct_calculated = (
                    float(outbound_hit_pct) if outbound_hit_pct else 0.0
                )
                hit_pct_calculated = round(hit_pct_calculated, 2)  # تقريب لرقمين عشريين
            except (ValueError, TypeError):
                hit_pct_calculated = 0.0
            print(
                f"✅ Total Lead Time Performance - Using Outbound hit_pct: {hit_pct_calculated}%"
            )

            # ✅ إذا لم يكن هناك chart_data من 3PL، نستخدم chart_data من Outbound
            if not chart_data:
                outbound_chart_data = outbound_result.get("chart_data", []) or []
                if outbound_chart_data:
                    chart_data = outbound_chart_data
                    print(
                        f"✅ Total Lead Time Performance - Using Outbound chart_data: {len(chart_data)} datasets"
                    )

            print(
                f"✅ Total Lead Time Performance - Final chart_data: {len(chart_data)} datasets"
            )

            # All-in-One / overview: append POD column charts (B2B/BTQ POD %) after main outbound charts
            if outbound_result:
                _pods_ch = outbound_result.get("chart_data_pods") or []
                if _pods_ch:
                    chart_data = list(chart_data or [])
                    chart_data.extend(_pods_ch)

            return {
                "detail_html": html,
                "chart_data": chart_data,
                "chart_title": "Total Lead Time Performance – On-Time Delivery",
                "count": total_count,
                "hit_pct": hit_pct_calculated,  # ✅ نسبة الـ Hit من Outbound
                "tab_data": tab_data,
            }

        except Exception as e:
            import traceback

            print(traceback.format_exc())
            return {
                "detail_html": f"<p class='text-danger'>⚠️ Error while processing data: {e}</p>",
                "chart_data": [],
                "count": 0,
                "hit_pct": 0,  # ✅ إضافة hit_pct في حالة الخطأ
            }

    def filter_dock_to_stock_combined(
        self, request, selected_month=None, selected_months=None, from_all_in_one=False
    ):
        """
        🔹 يعرض تاب Dock to stock بالاعتماد على تحليل Inbound (KPI ≤24h).
        ⚡️ سرعة تحميل عالية للتاب.
        """
        import hashlib
        from django.template.loader import render_to_string

        # --- Improvement: Faster cache, less code, non-blocking data retrieval ---

        # Key for caching based on source file and filters
        excel_path = _get_excel_path_for_request(request)
        _path_hash = hashlib.md5((excel_path or "").encode()).hexdigest()[:12]
        _month_part = str(selected_month).strip() if selected_month is not None else ""
        _months_list = (
            ",".join(map(str, selected_months)) if selected_months is not None else ""
        )
        _cache_key = f"tlp_dock_to_stock_fast_{_path_hash}_{_month_part}_{_months_list}"

        # Try cache very first (fast/minimal or full depending on mode)
        is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"
        requested_tab = (request.GET.get("tab") or "").strip().lower()
        wants_full = is_ajax and ("inbound" in requested_tab) and (not from_all_in_one)

        cache_key = _cache_key + ("::full" if wants_full else "::mini")
        cached = cache.get(cache_key)
        if cached is not None:
            return cached

        # --- Full mode: when user opens Inbound tab (AJAX), return full details ---
        if wants_full:
            try:
                full_res = self.filter_inbound(
                    request, selected_month=selected_month, selected_months=selected_months
                )
                # Cache full payload longer (details are heavier)
                cache.set(cache_key, _sanitize_for_json(full_res), 300)
                return full_res
            except Exception as e:
                import traceback

                print(traceback.format_exc())
                return {
                    "chart_data": [],
                    "detail_html": f"<p class='text-danger'>⚠️ Error: {e}</p>",
                    "count": 0,
                }

        # --- Minimal mode: used داخل All-in-One/Overview لتسريع التحميل ---
        try:
            # Fetch only the KPI summary from inbound
            inbound_result = self.filter_inbound(request, selected_month, selected_months)
            # sub_tables could be heavy, chart_data/statistics are small
            stats = inbound_result.get("stats", {})
            chart_data = inbound_result.get("chart_data", [])
            hit_pct = stats.get("hit_pct", 0)
            total_count = stats.get("total", 0)

            # Prepare minimal result
            result = {
                "chart_data": chart_data,
                # في وضع All-in-One لا نعرض تفاصيل التاب هنا (الـ overview بيعرض progress فقط)
                "detail_html": "",
                "count": total_count,
                "canvas_id": "chart-inbound-kpi",
                "hit_pct": hit_pct,
                "target_pct": 100,
                "tab_data": {
                    "name": "Inbound",
                    "stats": stats,
                    "canvas_id": "chart-inbound-kpi",
                },
            }
            # Put the fast minimal response in cache (short)
            cache.set(cache_key, _sanitize_for_json(result), 60)
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return {
                "chart_data": [],
                "detail_html": f"<p class='text-danger'>⚠️ Error: {e}</p>",
                "count": 0,
            }

        return result

    def overview_tab(
        self,
        request=None,
        selected_month=None,
        selected_months=None,
        from_all_in_one=False,
    ):
        from concurrent.futures import ThreadPoolExecutor, as_completed

        tab_cards = []

        target_manual = {
            "inbound": 99,
            "outbound": 98,
            "b2b outbound": 98,
            "b2c outbound": 99,
            "total lead time performance": 98,
            "return & refusal": 100,
            "safety kpi": 100,
            "traceability kpi": 100,
        }

        def process_tab(tab_name):
            detail_html, count, hit_pct = "", 0, 0
            try:
                res = {}
                tab_lower = tab_name.lower()
                month_for_filters = selected_month if not selected_months else None

                if tab_lower in ["rejections", "return & refusal"]:
                    res = self.filter_rejections_combined(
                        request,
                        month_for_filters,
                        selected_months=selected_months,
                    )
                elif tab_lower == "inbound":
                    res = self.filter_dock_to_stock_combined(
                        request,
                        month_for_filters,
                        selected_months=selected_months,
                        from_all_in_one=from_all_in_one,
                    )
                elif tab_lower == "capacity + expiry":
                    res = self.filter_capacity_expiry(
                        request,
                        month_for_filters,
                        selected_months=selected_months,
                    )
                elif tab_lower == "safety kpi":
                    res = self._render_safety_kpi_tab(request)
                elif tab_lower == "traceability kpi":
                    res = {"detail_html": "<p class='text-muted text-center p-4'>Loading data</p>", "count": 0, "hit_pct": 0}
                elif tab_lower == "b2c outbound":
                    res = self.filter_b2c_outbound(
                        request,
                        month_for_filters,
                        selected_months=selected_months,
                    )
                elif tab_lower in ("outbound", "b2b outbound") or "total lead time performance" in tab_lower:
                    res = self.filter_total_lead_time_performance(
                        request,
                        month_for_filters,
                        selected_months=selected_months,
                    )

                # النسبة الحقيقية زي ما راجعة من الدالة (أو من stats للتابات مثل B2C Outbound)
                hit_pct = res.get("hit_pct") or (res.get("stats") or {}).get("hit_pct", 0)
                if isinstance(hit_pct, dict):
                    if selected_month and selected_month.capitalize() in hit_pct:
                        hit_pct_val = hit_pct[selected_month.capitalize()]
                    else:
                        # نحسب المتوسط
                        hit_pct_val = int(round(sum(hit_pct.values()) / len(hit_pct)))
                else:
                    try:
                        hit_pct_val = int(round(float(hit_pct)))
                    except:
                        hit_pct_val = 0

                hit_pct_val = max(0, min(hit_pct_val, 100))

                target_pct = target_manual.get(tab_lower, 100)
                color_class = "bg-success" if hit_pct_val >= target_pct else "bg-danger"
                chart_data = list(res.get("chart_data", []) or [])
                _pods_extra = res.get("chart_data_pods")
                if isinstance(_pods_extra, list) and _pods_extra:
                    chart_data.extend(_pods_extra)
                chart_type = res.get("chart_type", "bar") or "bar"

                progress_html = f"""
                    <div class='mb-3'>
                        <div class='d-flex justify-content-between align-items-center mb-1'>
                            <strong class='text-capitalize'>{tab_name}</strong>
                            <small>{hit_pct_val}% / Target: {target_pct}%</small>
                        </div>
                        <div class='progress' style='height: 20px;'>
                            <div class='progress-bar {color_class}' role='progressbar'
                                 style='width: {hit_pct_val}%;' aria-valuenow='{hit_pct_val}'
                                 aria-valuemin='0' aria-valuemax='100'>
                                 {hit_pct_val}%
                            </div>
                        </div>
                    </div>
                """

                detail_html = progress_html + (res.get("detail_html", "") or "")
                count = res.get("count", 0)

            except Exception:
                detail_html = "<p class='text-muted'>No data available.</p>"
                hit_pct_val = 0
                target_pct = target_manual.get(tab_name.lower(), 100)
                chart_data = []
                chart_type = "bar"

            return {
                "name": tab_name,
                "hit_pct": hit_pct_val,
                "target_pct": target_pct,
                # في All-in-One نكتفي بعرض كروت الـ KPI بدون تفاصيل ثقيلة
                "detail_html": progress_html if from_all_in_one else detail_html,
                "count": count,
                "chart_data": chart_data,
                "chart_type": chart_type,
            }

        tabs_order = [
            "Inbound",
            "B2B Outbound",
            "B2C Outbound",
            "Capacity + Expiry",
            "Return & Refusal",
            "Safety KPI",
            "Traceability KPI",
        ]

        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {executor.submit(process_tab, t): t for t in tabs_order}
            for future in as_completed(futures):
                tab_cards.append(future.result())

        tab_cards.sort(key=lambda x: tabs_order.index(x["name"]))

        if not from_all_in_one:
            tab_cards = [
                t
                for t in tab_cards
                if t.get("name", "").strip().lower()
                not in ["rejections", "return & refusal"]
            ]

        all_progress_html = "<div class='card p-4 shadow-sm rounded-4 mb-4'>"
        all_progress_html += (
            "<h5 class='fw-bold text-primary mb-3'>📈 نسب الأداء لكل التابات</h5>"
        )
        for tab in tab_cards:
            color_class = (
                "bg-success" if tab["hit_pct"] >= tab["target_pct"] else "bg-danger"
            )
            all_progress_html += f"""
                <div class='mb-3'>
                    <div class='d-flex justify-content-between align-items-center mb-1'>
                        <strong>{tab['name']}</strong>
                        <small>{tab['hit_pct']}% / Target: {tab['target_pct']}%</small>
                    </div>
                    <div class='progress' style='height: 20px;'>
                        <div class='progress-bar {color_class}' role='progressbar'
                             style='width: {tab['hit_pct']}%;' aria-valuenow='{tab['hit_pct']}'
                             aria-valuemin='0' aria-valuemax='100'>
                             {tab['hit_pct']}%
                        </div>
                    </div>
                </div>
            """
        all_progress_html += "</div>"

        return {"tab_cards": tab_cards, "detail_html": all_progress_html}

    def _get_dashboard_include_context(self, request):
        """
        يُرجع سياق الداشبورد (نفس منطق dashboard_tab) لاستخدامه عند include
        container-fluid-dashboard في التمبلت حتى تورث base واللينكات والشارتس.
        """
        context = get_dashboard_tab_context(request)
        context["title"] = self.DASHBOARD_TAB_NAME
        # تاب Dashboard يقرأ من ملفه فقط (Aramco_Tamer3PL_KPI_Dashboard.xlsx)، وليس من all_sheet_nespresso
        excel_path = _get_dashboard_excel_path(request)

        # كل الداتا من الشيت فقط — لا قيم يدوية. لو مفيش ملف أو الشيت فاضي نستخدم قيم فارغة/صفر.
        if excel_path:
            inbound_data = _read_inbound_data_from_excel(excel_path, request)
            if inbound_data:
                context["inbound_kpi"] = inbound_data["inbound_kpi"]
                context["pending_shipments"] = inbound_data["pending_shipments"]

            charts_from_excel = _read_dashboard_charts_from_excel(excel_path, request)
            for key, value in charts_from_excel.items():
                if value is not None:
                    context[key] = value

            outbound_data = _read_outbound_data_from_excel(excel_path, request)
            if outbound_data and "outbound_kpi" in outbound_data:
                context["outbound_kpi"] = outbound_data["outbound_kpi"]
                context["outbound_kpi_keys_from_sheet"] = outbound_data.get("outbound_kpi_keys_from_sheet", [])

            pods_data = _read_pods_data_from_excel(excel_path, request)
            if pods_data:
                context["pod_compliance_chart_data"] = {
                    "categories": pods_data.get("categories", []),
                    "series": pods_data.get("series", []),
                }
                if "pod_status_breakdown" in pods_data:
                    context["pod_status_breakdown"] = pods_data["pod_status_breakdown"]

            returns_data = _read_returns_data_from_excel(excel_path, request)
            if returns_data:
                context["returns_kpi"] = returns_data.get("returns_kpi", {})
                if "returns_chart_data" in returns_data:
                    context["returns_chart_data"] = returns_data["returns_chart_data"]

            inventory_data = _read_inventory_data_from_excel(excel_path, request)
            if inventory_data:
                context["inventory_kpi"] = inventory_data.get("inventory_kpi", {})

            capacity_data = _read_inventory_snapshot_capacity_from_excel(excel_path, request)
            if capacity_data:
                context["inventory_capacity_data"] = capacity_data.get("inventory_capacity_data", {})

            warehouse_table = _read_inventory_warehouse_table_from_excel(excel_path, request)
            if warehouse_table:
                context["inventory_warehouse_table"] = warehouse_table.get("inventory_warehouse_table", [])

            returns_region = _read_returns_region_table_from_excel(excel_path, request)
            if returns_region:
                context["returns_region_table"] = returns_region.get("returns_region_table", [])

        # قيم فارغة/صفر فقط عند غياب الملف أو فشل القراءة (حتى لا يكسر القالب)
        _empty_inbound_kpi = {
            "number_of_vehicles": 0,
            "number_of_shipments": 0,
            "number_of_pallets": 0,
            "total_quantity": 0,
            "total_quantity_display": "0",
        }
        _empty_outbound_kpi = {
            "released_orders": 0,
            "picked_orders": 0,
            "number_of_pallets": 0,
        }
        _empty_pod_chart = {"categories": [], "series": []}
        _empty_pod_breakdown = [
            {"label": "On Time", "pct": 0, "color": "#7FB7A6"},
            {"label": "Pending", "pct": 0, "color": "#A8C8EB"},
            {"label": "Late", "pct": 0, "color": "#E8A8A2"},
        ]

        # تتبع أي أقسام تعرض قيماً افتراضية (صفر) لعرض تنبيه "ارفع الملف مرة أخرى"
        missing_by_section = {}
        if "inbound_kpi" not in context:
            missing_by_section["Dashboard – Inbound"] = ["Number of Shipments", "Number of Pallets (LPNs)", "Total Quantity", "Pending Shipments"]
        context.setdefault("inbound_kpi", _empty_inbound_kpi)
        context.setdefault("pending_shipments", [])

        _outbound_card_names = {"released_orders": "Released Orders", "picked_orders": "Picked Orders", "number_of_pallets": "Number of Pallets (LPNs)"}
        if "outbound_kpi" not in context:
            missing_by_section.setdefault("Dashboard – Outbound", []).extend(["Released Orders", "Picked Orders", "Number of Pallets (LPNs)"])
        else:
            keys_from_sheet = context.get("outbound_kpi_keys_from_sheet") or []
            for key, card_name in _outbound_card_names.items():
                if key not in keys_from_sheet:
                    missing_by_section.setdefault("Dashboard – Outbound", []).append(card_name)
        context.setdefault("outbound_kpi", _empty_outbound_kpi)
        if "pod_compliance_chart_data" not in context:
            missing_by_section.setdefault("Dashboard – Outbound", []).append("PODs Compliance (chart)")
        context.setdefault("outbound_chart_data", _empty_pod_chart)
        context.setdefault("pod_compliance_chart_data", _empty_pod_chart)
        context.setdefault("pod_status_breakdown", _empty_pod_breakdown)

        if "returns_kpi" not in context:
            missing_by_section["Dashboard – Returns"] = ["Total SKUs", "Total LPNs", "Returns chart"]
        context.setdefault("returns_kpi", {"total_skus": 0, "total_lpns": 0})
        context.setdefault("returns_chart_data", _empty_pod_chart)
        context.setdefault("returns_region_table", [])

        if "inventory_kpi" not in context:
            missing_by_section["Dashboard – Inventory"] = ["Total SKUs", "Total LPNs", "Utilization %", "Capacity chart", "Warehouse table"]
        context.setdefault("inventory_kpi", {"total_skus": 0, "total_lpns": 0, "utilization_pct": ""})
        context.setdefault("inventory_capacity_data", {"used": 0, "available": 0})
        context.setdefault("inventory_warehouse_table", [])

        context["dashboard_missing_data"] = [{"section": k, "cards": v} for k, v in missing_by_section.items()]
        return context

    def dashboard_tab(self, request):
        """
        🔹 تاب Dashboard: يعرض تصميم الداشبورد (container-fluid-dashboard).
        التمبلت منفصل عن excel-sheet-table ويُحمّل داخل منطقة المحتوى عند اختيار تاب Dashboard.
        نفس فكرة rejection: نرجع detail_html + chart_data + chart_title عشان الشارتات تبقى دينامك.
        """
        try:
            context = self._get_dashboard_include_context(request)
            html = render_to_string(
                "container-fluid-dashboard.html",
                context,
                request=request,
            )
            # نفس شكل الـ rejection: chart_data و chart_title للشارتات الدينامك
            outbound_chart = context.get("outbound_chart_data")
            chart_data = []
            if outbound_chart and isinstance(outbound_chart, dict):
                categories = outbound_chart.get("categories", [])
                series = outbound_chart.get("series", [])
                if categories and series is not None:
                    chart_data.append({
                        "type": "line",
                        "name": "POD Compliance",
                        "dataPoints": [{"label": c, "y": float(s)} for c, s in zip(categories, series)],
                    })
            return {
                "detail_html": html,
                "chart_data": chart_data,
                "chart_title": "Dashboard – POD Compliance",
                "dashboard_charts": {
                    "outbound": context.get("outbound_chart_data"),
                    "returns": context.get("returns_chart_data"),
                    "inventory": context.get("inventory_capacity_data"),
                },
            }
        except Exception as e:
            import traceback

            traceback.print_exc()
            return {"error": f"An error occurred while loading Dashboard: {e}"}

    def meeting_points_tab(self, request):
        """
        🔹 عرض تاب Meeting Points & Action مع إمكانية الفلترة حسب الحالة (منتهية / غير منتهية)
        """
        try:
            # ✅ جلب الحالة من الـ GET parameter
            status_filter = request.GET.get(
                "status"
            )  # القيم الممكنة: done / pending / all

            # ✅ استرجاع كل النقاط بالترتيب
            meeting_points = MeetingPoint.objects.all().order_by(
                "is_done", "-created_at"
            )

            # ✅ تطبيق الفلترة بناءً على الحالة
            if status_filter == "done":
                meeting_points = meeting_points.filter(is_done=True)
            elif status_filter == "pending":
                meeting_points = meeting_points.filter(is_done=False)
            # 'all' يعرض كل النقاط (done + pending)
            # لا حاجة لفلترة إضافية لأنه استرجعنا كل النقاط في البداية

            # ✅ إحصائيات
            done_count = meeting_points.filter(is_done=True).count()
            total_count = meeting_points.count()

            # ✅ تجهيز البيانات للتمبلت مع assigned_to
            meeting_data = [
                {
                    "id": p.id,
                    "description": p.description,
                    "assigned_to": getattr(
                        p, "assigned_to", ""
                    ),  # ✅ الاسم ممكن يكون فاضي
                    "status": "Done" if p.is_done else "Pending",
                    "created_at": p.created_at,
                    "target_date": p.target_date,
                }
                for p in meeting_points
            ]

            context = {
                "meeting_points": meeting_points,
                "meeting_data": meeting_data,  # لو حابة تستخدمي البيانات مباشرة في JS
                "done_count": done_count,
                "total_count": total_count,
                "status_filter": status_filter,
            }

            # ✅ بناء HTML من التمبلت
            html = render_to_string("meeting_points.html", context, request=request)

            # ✅ إرجاع النتيجة
            return JsonResponse(
                {
                    "detail_html": html,
                    "count": meeting_points.count(),
                    "done_count": done_count,
                    "total_count": total_count,
                },
                safe=False,
            )

        except Exception as e:
            import traceback

            traceback.print_exc()
            return JsonResponse(
                {"error": f"An error occurred while loading data: {e}"}, status=500
            )


class MeetingPointListCreateView(View):
    template_name = "meeting_points.html"

    def get(self, request, *args, **kwargs):
        status_filter = request.GET.get("status")  # "done" أو "pending" أو None

        today = date.today()
        current_month, current_year = today.month, today.year

        # حساب الشهر السابق
        if current_month == 1:
            prev_month = 12
            prev_year = current_year - 1
        else:
            prev_month = current_month - 1
            prev_year = current_year

        # ✅ جلب كل النقاط (الشهر الحالي كله + pending من الشهر السابق)
        meeting_points = MeetingPoint.objects.filter(
            Q(created_at__year=current_year, created_at__month=current_month)
            | Q(created_at__year=prev_year, created_at__month=prev_month, is_done=False)
        ).order_by("is_done", "-created_at")

        # ✅ تطبيق الفلتر لو المستخدم اختار حاجة
        if status_filter == "done":
            meeting_points = meeting_points.filter(is_done=True)
        elif status_filter == "pending":
            meeting_points = meeting_points.filter(is_done=False)

        done_count = meeting_points.filter(is_done=True).count()
        total_count = meeting_points.count()

        return render(
            request,
            self.template_name,
            {
                "meeting_points": meeting_points,
                "done_count": done_count,
                "total_count": total_count,
                "status_filter": status_filter,
            },
        )

    def post(self, request, *args, **kwargs):
        description = request.POST.get("description", "").strip()
        target_date = request.POST.get("target_date", "").strip() or None
        assigned_to = request.POST.get("assigned_to", "").strip() or None

        if description:
            point = MeetingPoint.objects.create(
                description=description,
                target_date=target_date,
                assigned_to=assigned_to if assigned_to else None,
            )

            return JsonResponse(
                {
                    "id": point.id,
                    "description": point.description,
                    "assigned_to": point.assigned_to,
                    "created_at": str(point.created_at),
                    "target_date": str(point.target_date),
                    "is_done": point.is_done,
                }
            )

        return JsonResponse({"error": "Empty description"}, status=400)


class ToggleMeetingPointView(View):
    def post(self, request, pk, *args, **kwargs):
        point = get_object_or_404(MeetingPoint, pk=pk)
        point.is_done = not point.is_done
        point.save()
        return JsonResponse({"is_done": point.is_done})


class DoneMeetingPointView(View):
    def post(self, request, pk, *args, **kwargs):
        point = get_object_or_404(MeetingPoint, pk=pk)
        point.is_done = not point.is_done
        point.save()
        return JsonResponse({"is_done": point.is_done})
